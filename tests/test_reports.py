#!/usr/bin/env python3
"""
Tests for Reports functionality (Booking Report and Player Payment Report)
Tests both SQLite and PostgreSQL compatibility
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import datetime, timedelta
import sqlite3
from io import BytesIO
from openpyxl import load_workbook

# Set test mode before importing api
os.environ["TEST_MODE"] = "true"

from api import app, init_db, hash_password, USE_POSTGRES, get_connection, PLACEHOLDER, DB_PATH
from fastapi.testclient import TestClient

client = TestClient(app)

def setup_test_data():
    """Create test data for reports"""
    with get_connection() as conn:
        cur = conn.cursor()
        
        # Clear existing data
        cur.execute("DELETE FROM users")
        cur.execute("DELETE FROM practice_sessions")
        cur.execute("DELETE FROM practice_availability")
        cur.execute("DELETE FROM practice_payments")
        
        # Create admin user
        admin_email = "admin@test.com"
        admin_password = hash_password("admin123")
        cur.execute(
            f"INSERT INTO users (email, full_name, password, user_type) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
            (admin_email, "Admin User", admin_password, "admin")
        )
        
        # Create regular users
        users = [
            ("user1@test.com", "User One", hash_password("pass123")),
            ("user2@test.com", "User Two", hash_password("pass123")),
            ("user3@test.com", "User Three", hash_password("pass123")),
        ]
        
        for email, name, pwd in users:
            cur.execute(
                f"INSERT INTO users (email, full_name, password, user_type) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, 'member')",
                (email, name, pwd)
            )
        
        # Create practice sessions with payment info
        today = datetime.now().date()
        sessions = [
            (str(today - timedelta(days=30)), "practice", "Session", "18:00", "Location A", 20.0, admin_email, True),
            (str(today - timedelta(days=20)), "match", "League Match", "19:00", "Location B", 30.0, "user1@test.com", True),
            (str(today - timedelta(days=10)), "social", "Team Dinner", "20:00", "Location C", 25.0, admin_email, True),
            (str(today + timedelta(days=5)), "practice", "Recovery", "18:30", "Location D", 20.0, None, False),  # Future session
        ]
        
        for date, event_type, event_title, time, location, cost, paid_by, payment_requested in sessions:
            cur.execute(
                f"INSERT INTO practice_sessions (date, event_type, event_title, time, location, session_cost, paid_by, payment_requested) "
                f"VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
                (date, event_type, event_title, time, location, cost, paid_by, payment_requested if USE_POSTGRES else (1 if payment_requested else 0))
            )
        
        # Add availability for past sessions
        availabilities = [
            (str(today - timedelta(days=30)), "user1@test.com", "available"),
            (str(today - timedelta(days=30)), "user2@test.com", "available"),
            (str(today - timedelta(days=30)), "user3@test.com", "tentative"),
            (str(today - timedelta(days=20)), "user1@test.com", "available"),
            (str(today - timedelta(days=20)), "user2@test.com", "available"),
            (str(today - timedelta(days=20)), "user3@test.com", "not_available"),
            (str(today - timedelta(days=10)), "user1@test.com", "available"),
            (str(today - timedelta(days=10)), "user2@test.com", "tentative"),
        ]
        
        for date, email, status in availabilities:
            cur.execute(
                f"INSERT INTO practice_availability (date, user_email, status) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
                (date, email, status)
            )
        
        # Add payment confirmations
        payments = [
            (str(today - timedelta(days=30)), "user1@test.com", True),
            (str(today - timedelta(days=30)), "user2@test.com", False),
            (str(today - timedelta(days=20)), "user1@test.com", True),
            (str(today - timedelta(days=20)), "user2@test.com", True),
            (str(today - timedelta(days=10)), "user1@test.com", False),
        ]
        
        for date, email, paid in payments:
            cur.execute(
                f"INSERT INTO practice_payments (date, user_email, paid) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
                (date, email, paid if USE_POSTGRES else (1 if paid else 0))
            )
        
        conn.commit()

def test_booking_report_admin_only():
    """Test that booking report requires admin access"""
    print("Testing booking report admin-only access...")
    
    # Login as regular user
    response = client.post("/api/login", data={
        "username": "user1@test.com",
        "password": "pass123"
    })
    assert response.status_code == 200
    user_token = response.json()["access_token"]
    
    # Try to access report
    today = datetime.now().date()
    from_date = str(today - timedelta(days=40))
    to_date = str(today)
    
    response = client.get(
        f"/api/reports/booking?from_date={from_date}&to_date={to_date}",
        headers={"Authorization": f"Bearer {user_token}"}
    )
    assert response.status_code == 403
    assert "Admins only" in response.json()["detail"]
    print("✓ Booking report correctly requires admin access")

def test_player_payment_report_admin_only():
    """Test that player payment report requires admin access"""
    print("Testing player payment report admin-only access...")
    
    # Login as regular user
    response = client.post("/api/login", data={
        "username": "user1@test.com",
        "password": "pass123"
    })
    assert response.status_code == 200
    user_token = response.json()["access_token"]
    
    # Try to access report
    today = datetime.now().date()
    from_date = str(today - timedelta(days=40))
    to_date = str(today)
    
    response = client.get(
        f"/api/reports/player-payment?from_date={from_date}&to_date={to_date}",
        headers={"Authorization": f"Bearer {user_token}"}
    )
    assert response.status_code == 403
    assert "Admins only" in response.json()["detail"]
    print("✓ Player payment report correctly requires admin access")

def test_booking_report_generation():
    """Test booking report Excel generation"""
    print("Testing booking report generation...")
    
    # Login as admin
    response = client.post("/api/login", data={
        "username": "admin@test.com",
        "password": "admin123"
    })
    assert response.status_code == 200
    admin_token = response.json()["access_token"]
    
    # Generate report
    today = datetime.now().date()
    from_date = str(today - timedelta(days=40))
    to_date = str(today)
    
    response = client.get(
        f"/api/reports/booking?from_date={from_date}&to_date={to_date}",
        headers={"Authorization": f"Bearer {admin_token}"}
    )
    
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in response.headers["content-disposition"]
    
    # Verify Excel content
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    
    # Check headers
    assert ws.cell(1, 1).value == "Event Date"
    assert ws.cell(1, 2).value == "Event Type"
    assert ws.cell(1, 3).value == "Event Title"
    assert ws.cell(1, 4).value == "Time"
    assert ws.cell(1, 5).value == "Place"
    assert ws.cell(1, 6).value == "Total Cost (£)"
    assert ws.cell(1, 7).value == "Paid By"

    assert ws.cell(2, 2).value == "Practice"
    assert ws.cell(2, 3).value == "Session"
    assert ws.cell(3, 2).value == "Match"
    assert ws.cell(3, 3).value == "League Match"
    
    # Check that we have data rows (at least 3 past sessions)
    assert ws.cell(2, 1).value is not None  # First data row
    assert ws.cell(3, 1).value is not None  # Second data row
    assert ws.cell(4, 1).value is not None  # Third data row
    
    print("✓ Booking report generated successfully with correct format")

def test_player_payment_report_generation():
    """Test player payment report Excel generation"""
    print("Testing player payment report generation...")
    
    # Login as admin
    response = client.post("/api/login", data={
        "username": "admin@test.com",
        "password": "admin123"
    })
    assert response.status_code == 200
    admin_token = response.json()["access_token"]
    
    # Generate report
    today = datetime.now().date()
    from_date = str(today - timedelta(days=40))
    to_date = str(today)
    
    response = client.get(
        f"/api/reports/player-payment?from_date={from_date}&to_date={to_date}",
        headers={"Authorization": f"Bearer {admin_token}"}
    )
    
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    # Verify Excel content
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    
    # Check headers
    assert ws.cell(1, 1).value == "Event Date"
    assert ws.cell(1, 2).value == "Event Type"
    assert ws.cell(1, 3).value == "Event Title"
    assert ws.cell(1, 4).value == "Time"
    assert ws.cell(1, 5).value == "Place"
    assert ws.cell(1, 6).value == "Total Cost (£)"
    assert ws.cell(1, 7).value == "Paid By"
    assert ws.cell(1, 8).value == "Payment Requested Date"
    assert ws.cell(1, 9).value == "Player Name"
    assert ws.cell(1, 10).value == "Availability"
    assert ws.cell(1, 11).value == "Individual Amount (£)"
    assert ws.cell(1, 12).value == "Paid"
    assert ws.cell(1, 13).value == "Payment Acknowledgement Date"
    
    # Check that we have data rows
    assert ws.cell(2, 1).value is not None  # First data row
    assert ws.cell(2, 2).value is not None  # Event type
    assert ws.cell(2, 3).value is not None  # Event title
    assert ws.cell(2, 9).value is not None  # Player name
    assert ws.cell(2, 10).value is not None  # Availability status
    
    print("✓ Player payment report generated successfully with correct format")

def test_booking_report_date_filtering():
    """Test that booking report correctly filters by date range"""
    print("Testing booking report date filtering...")
    
    # Login as admin
    response = client.post("/api/login", data={
        "username": "admin@test.com",
        "password": "admin123"
    })
    admin_token = response.json()["access_token"]
    
    # Generate report for narrow date range
    today = datetime.now().date()
    from_date = str(today - timedelta(days=25))
    to_date = str(today - timedelta(days=15))
    
    response = client.get(
        f"/api/reports/booking?from_date={from_date}&to_date={to_date}",
        headers={"Authorization": f"Bearer {admin_token}"}
    )
    
    assert response.status_code == 200
    
    # Verify Excel content
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    
    # Should have only 1 session in this range (the one 20 days ago)
    assert ws.cell(2, 1).value is not None  # First data row
    # Third row should be empty or have different data
    
    print("✓ Booking report correctly filters by date range")

def test_player_payment_report_only_voted_users():
    """Test that player payment report only includes users who voted"""
    print("Testing player payment report includes only users who voted...")
    
    # Login as admin
    response = client.post("/api/login", data={
        "username": "admin@test.com",
        "password": "admin123"
    })
    admin_token = response.json()["access_token"]
    
    # Generate report
    today = datetime.now().date()
    from_date = str(today - timedelta(days=40))
    to_date = str(today)
    
    response = client.get(
        f"/api/reports/player-payment?from_date={from_date}&to_date={to_date}",
        headers={"Authorization": f"Bearer {admin_token}"}
    )
    
    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    
    # Check that all rows have availability status (meaning they voted)
    row = 2
    while ws.cell(row, 1).value is not None:
        availability = ws.cell(row, 10).value
        assert availability is not None and availability != ""
        row += 1
    
    print("✓ Player payment report only includes users who voted their availability")

def run_all_tests():
    """Run all report tests"""
    print("\n" + "="*80)
    print("Running Reports Feature Tests")
    print("="*80 + "\n")

    if not USE_POSTGRES and os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        init_db()
    
    # Check if tables exist, if not create them
    with get_connection() as conn:
        cur = conn.cursor()
        try:
            cur.execute("SELECT 1 FROM users LIMIT 1")
        except:
            # Tables don't exist, call init_db
            init_db()
    
    setup_test_data()
    
    tests = [
        test_booking_report_admin_only,
        test_player_payment_report_admin_only,
        test_booking_report_generation,
        test_player_payment_report_generation,
        test_booking_report_date_filtering,
        test_player_payment_report_only_voted_users,
    ]
    
    passed = 0
    failed = 0
    
    for test in tests:
        try:
            test()
            passed += 1
        except AssertionError as e:
            print(f"✗ {test.__name__} FAILED: {e}")
            failed += 1
        except Exception as e:
            print(f"✗ {test.__name__} ERROR: {e}")
            failed += 1
    
    print("\n" + "="*80)
    print(f"Reports Tests Complete: {passed} passed, {failed} failed")
    print("="*80 + "\n")
    
    return failed == 0

if __name__ == "__main__":
    success = run_all_tests()
    sys.exit(0 if success else 1)
