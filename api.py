from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, BackgroundTasks
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, EmailStr
from datetime import date, datetime, timedelta
from html.parser import HTMLParser
from typing import List, Optional
import html
import re
import sqlite3
import hashlib
import uuid
import secrets
import json
import os
import shutil
from contextlib import contextmanager
import cloudinary
import cloudinary.uploader
from urllib.parse import urlparse
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from apscheduler.schedulers.background import BackgroundScheduler
from local_env import load_local_env

load_local_env()

from whatsapp_notifier import (
    find_group_chat_id,
    get_instance_state,
    keep_whatsapp_instance_alive,
    resolve_group_chat_id,
    send_group_message,
    whatsapp_is_configured,
)

# --- Database helpers (supports both SQLite and Postgres) ---
DATABASE_URL = os.environ.get("DATABASE_URL")  # Render provides this
SESSION_DURATION = timedelta(days=90)
USE_POSTGRES = DATABASE_URL is not None

# Test database configuration - use separate database for testing
TEST_MODE = os.environ.get("TEST_MODE", "false").lower() == "true"

if USE_POSTGRES:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    # Fix for Render's postgres:// URL (needs postgresql://)
    if DATABASE_URL.startswith("postgres://"):
        DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
else:
    # Use test database if in test mode, otherwise use production database
    DB_PATH = "test_football_club.db" if TEST_MODE else "football_club.db"
    UPLOAD_DIR = "uploads"
    os.makedirs(UPLOAD_DIR, exist_ok=True)

# SQL placeholder - PostgreSQL uses %s, SQLite uses ?
PLACEHOLDER = "%s" if USE_POSTGRES else "?"

# Cloudinary configuration
CLOUDINARY_URL = os.environ.get("CLOUDINARY_URL")
if CLOUDINARY_URL:
    cloudinary.config(
        cloud_name=os.environ.get("CLOUDINARY_CLOUD_NAME"),
        api_key=os.environ.get("CLOUDINARY_API_KEY"),
        api_secret=os.environ.get("CLOUDINARY_API_SECRET"),
    )

# ========== FORGOT PASSWORD FEATURE - Email Configuration ==========
# These environment variables configure the email service for password recovery
# Set these in your environment or .env file to enable email functionality
# See EMAIL_SETUP.md for detailed configuration instructions
# ====================================================================
SMTP_SERVER = os.environ.get("SMTP_SERVER")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USERNAME = os.environ.get("SMTP_USERNAME")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD")
FROM_EMAIL = os.environ.get("FROM_EMAIL") or SMTP_USERNAME or ""
FRONTEND_URL = os.environ.get("FRONTEND_URL", "http://localhost:5173")
WHATSAPP_NOTIFICATIONS_ENABLED = os.environ.get("WHATSAPP_NOTIFICATIONS_ENABLED", "false").lower() == "true"

whatsapp_scheduler = BackgroundScheduler()

NOTIFICATION_TARGET_OPTIONS = {"all_active_users", "admin_users", "available_players", "direct_user"}
THEME_PREFERENCES = {"nordic_neutral", "east_bengal", "mohun_bagan"}
EVENT_TYPE_OPTIONS = {"practice", "match", "social", "others"}
NOTIFICATION_TYPE_DEFAULTS = {
    "practice": {
        "display_name": "New Event Added",
        "description": "Sent when a new bookable event is created.",
        "app_enabled": True,
        "email_enabled": False,
        "whatsapp_enabled": True,
        "target_audience": "all_active_users",
        "app_template": "New {{event_name}} added on {{date}}{{time_suffix}}{{location_suffix}}. Please update your availability.",
        "email_subject": "New {{event_name}} on {{date}}",
        "email_template": "{{event_name}} has been added for {{date}}{{time_suffix}}{{location_suffix}}.\n\nPlease open the app and update your availability.",
        "whatsapp_template": "📅 *NEW EVENT ADDED*\n\n{{event_name}}\n📅 {{date}}\n{{time_line}}{{location_line}}\nPlease update your availability in the app.",
    },
    "forum_post": {
        "display_name": "New Forum Post Added",
        "description": "Sent when a new forum post is created.",
        "app_enabled": True,
        "email_enabled": False,
        "whatsapp_enabled": False,
        "target_audience": "all_active_users",
        "app_template": "New post added by {{author_name}}",
        "email_subject": "New forum post from {{author_name}}",
        "email_template": "A new forum post was added by {{author_name}}.\n\n{{content_preview}}",
        "whatsapp_template": "💬 *NEW FORUM POST*\n\nBy {{author_name}}\n\n{{content_preview}}\n\nOpen the app to join the conversation.",
    },
    "payment_request": {
        "display_name": "Event Payment Requested",
        "description": "Sent when payment is requested for a bookable event.",
        "app_enabled": True,
        "email_enabled": False,
        "whatsapp_enabled": True,
        "target_audience": "available_players",
        "app_template": "Payment requested by Admin for {{event_name}} on {{date}}{{time_suffix}}{{location_comma_suffix}}.",
        "email_subject": "Payment requested for {{event_name}} on {{date}}",
        "email_template": "Payment has been requested for {{event_name}} on {{date}}{{time_suffix}}{{location_comma_suffix}}.\n\nPlease confirm your payment in the app.",
        "whatsapp_template": "💷 *EVENT PAYMENT REQUEST*\n\n{{event_name}}\n📅 {{date}}\n{{time_line}}{{location_line}}\nAvailable players should confirm payment in the app.",
    },
    "payment_confirmed": {
        "display_name": "Event Payment Confirmed",
        "description": "Sent to admins in the app when a member confirms payment for a bookable event.",
        "app_enabled": True,
        "email_enabled": False,
        "whatsapp_enabled": False,
        "target_audience": "admin_users",
        "app_template": "{{member_name}} confirmed payment for {{event_name}} on {{date}}{{time_suffix}}{{location_comma_suffix}}.",
        "email_subject": "Payment confirmed for {{event_name}} on {{date}}",
        "email_template": "{{member_name}} confirmed payment for {{event_name}} on {{date}}{{time_suffix}}{{location_comma_suffix}}.",
        "whatsapp_template": "",
    },
    "pending_payment_reminder": {
        "display_name": "Pending Event Payment Reminder",
        "description": "Sent daily at 8 PM when at least 72 hours have passed since the last admin payment request and the latest requested event still has pending payments.",
        "app_enabled": True,
        "email_enabled": False,
        "whatsapp_enabled": True,
        "target_audience": "all_active_users",
        "app_template": "Gentle reminder: There are pending payments for {{event_name}}. Please check the app using {{payments_link}}.",
        "email_subject": "Pending payment reminder for {{event_name}}",
        "email_template": "Gentle reminder: There are pending payments for {{event_name}} on {{date}}{{time_suffix}}{{location_comma_suffix}}. Please check in the app using {{payments_link}}.",
        "whatsapp_template": "💷 *PENDING PAYMENT REMINDER*\n\n{{event_name}}\n📅 {{date}}\n{{time_line}}{{location_line}}Please check the app: {{payments_link}}",
    },
    "session_capacity_reached": {
        "display_name": "Event Capacity Reached",
        "description": "Sent when the final available slot is taken for a bookable event.",
        "app_enabled": True,
        "email_enabled": False,
        "whatsapp_enabled": True,
        "target_audience": "all_active_users",
        "app_template": "{{event_name}} has reached maximum capacity for {{date}}{{time_suffix}}{{location_comma_suffix}}. Maximum capacity is {{maximum_capacity}} players, so no more Available selections are allowed right now.",
        "email_subject": "{{event_name}} capacity reached for {{date}}",
        "email_template": "{{event_name}} on {{date}}{{time_suffix}}{{location_comma_suffix}} has reached its maximum capacity of {{maximum_capacity}} players. No more Available selections are allowed right now. We will notify players if slots become available before the event.",
        "whatsapp_template": "⛔ *EVENT FULL*\n\n{{event_name}}\n📅 {{date}}\n{{time_line}}{{location_line}}👥 Maximum capacity reached: {{maximum_capacity}}\nNo more *Available* selections are allowed right now. We will notify everyone if slots open up before the event.",
    },
    "practice_slot_available": {
        "display_name": "Event Slot Available",
        "description": "Sent when upcoming event slots are still available within 72 hours of the event.",
        "app_enabled": True,
        "email_enabled": False,
        "whatsapp_enabled": True,
        "target_audience": "all_active_users",
        "app_template": "There are {{remaining_slots}} slots available for {{event_name}} on {{date}}{{time_suffix}}{{location_comma_suffix}}. Maximum capacity: {{maximum_capacity}}.",
        "email_subject": "Slots available for {{event_name}} on {{date}}",
        "email_template": "There are {{remaining_slots}} slots available for {{event_name}} on {{date}}{{time_suffix}}{{location_comma_suffix}}. Maximum capacity: {{maximum_capacity}}.",
        "whatsapp_template": "✅ *EVENT SLOTS AVAILABLE*\n\n{{event_name}}\n📅 {{date}}\n{{time_line}}{{location_line}}👥 Slots available: {{remaining_slots}} of {{maximum_capacity}}\nBook your place in the app if you want to join.",
    },
    "welcome_signup": {
        "display_name": "Welcome Email On Signup",
        "description": "Sent only to the user who has just registered.",
        "app_enabled": False,
        "email_enabled": True,
        "whatsapp_enabled": False,
        "target_audience": "direct_user",
        "app_template": "",
        "email_subject": "Welcome to Glasgow Bengali FC",
        "email_template": "Hi {{full_name}},\n\nWelcome to Glasgow Bengali FC. Thanks for joining our ever-growing club, where fun meets football.\n\nBest wishes,\nGlasgow Bengali FC",
        "whatsapp_template": "",
    },
}

@contextmanager
def get_connection():
    if USE_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
        try:
            yield conn
        finally:
            conn.close()
    else:
        conn = sqlite3.connect(DB_PATH, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
        finally:
            conn.close()

def dict_factory(cursor, row):
    """Helper for SQLite to return dict-like rows"""
    if USE_POSTGRES:
        return row
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d

def init_db():
    # Safety check: Prevent tests from running on production database
    if not TEST_MODE and not USE_POSTGRES and os.path.exists(DB_PATH):
        try:
            with sqlite3.connect(DB_PATH, timeout=1) as check_conn:
                check_cursor = check_conn.cursor()
                check_cursor.execute("SELECT COUNT(*) FROM users")
                user_count = check_cursor.fetchone()[0]
                if user_count > 0:
                    # Check if this is a test run
                    import sys
                    is_testing = any("test" in arg.lower() for arg in sys.argv)
                    if is_testing:
                        print("=" * 80)
                        print("🚨 CRITICAL WARNING: TESTS RUNNING ON PRODUCTION DATABASE! 🚨")
                        print("=" * 80)
                        print(f"❌ Production database detected with {user_count} users")
                        print("❌ Tests are NOT ALLOWED to run on production database")
                        print("❌ This may delete or corrupt production data")
                        print("")
                        print("✅ SOLUTION: Set TEST_MODE=true environment variable")
                        print("✅ Example: export TEST_MODE=true")
                        print("✅ Or run tests with: TEST_MODE=true python test_script.py")
                        print("")
                        print("🛑 TEST EXECUTION FAILED - DATABASE SAFETY VIOLATION")
                        print("=" * 80)
                        raise RuntimeError("TESTS NOT ALLOWED ON PRODUCTION DATABASE. Set TEST_MODE=true to use test database.")
                    else:
                        # Normal production startup - no warning needed
                        pass
        except RuntimeError:
            # Re-raise the database safety violation
            raise
        except:
            pass
    
    with get_connection() as conn:
        if USE_POSTGRES:
            cur = conn.cursor()
            # Postgres uses SERIAL instead of AUTOINCREMENT
            cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                email VARCHAR(255) UNIQUE NOT NULL,
                full_name VARCHAR(255) NOT NULL,
                password VARCHAR(255) NOT NULL,
                user_type VARCHAR(50) DEFAULT 'member',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_login TIMESTAMP,
                birthday DATE,
                bank_name VARCHAR(255),
                sort_code VARCHAR(20),
                account_number VARCHAR(20),
                theme_preference VARCHAR(50) DEFAULT 'nordic_neutral',
                is_deleted BOOLEAN DEFAULT FALSE,
                deleted_at TIMESTAMP,
                deleted_by VARCHAR(255)
            )
            """)
            # Add user_type column if it doesn't exist (for existing databases)
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='users' AND column_name='user_type'
                        ) THEN
                            ALTER TABLE users ADD COLUMN user_type VARCHAR(50) DEFAULT 'member';
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add user_type column: {e}")
                conn.rollback()
            
            # Add created_at column if it doesn't exist (for existing databases)
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='users' AND column_name='created_at'
                        ) THEN
                            ALTER TABLE users ADD COLUMN created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP;
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add created_at column: {e}")
                conn.rollback()
            
            # Add is_deleted column if it doesn't exist (for existing databases)
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='users' AND column_name='is_deleted'
                        ) THEN
                            ALTER TABLE users ADD COLUMN is_deleted BOOLEAN DEFAULT FALSE;
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add is_deleted column: {e}")
                conn.rollback()
            
            # Add deleted_at column if it doesn't exist (for existing databases)
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='users' AND column_name='deleted_at'
                        ) THEN
                            ALTER TABLE users ADD COLUMN deleted_at TIMESTAMP;
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add deleted_at column: {e}")
                conn.rollback()
            
            # Add deleted_by column if it doesn't exist (for existing databases)
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='users' AND column_name='deleted_by'
                        ) THEN
                            ALTER TABLE users ADD COLUMN deleted_by VARCHAR(255);
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add deleted_by column: {e}")
                conn.rollback()
            
            # Add last_login column if it doesn't exist (for existing databases)
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='users' AND column_name='last_login'
                        ) THEN
                            ALTER TABLE users ADD COLUMN last_login TIMESTAMP;
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add last_login column: {e}")
                conn.rollback()
            
            # Add birthday column if it doesn't exist (for existing databases)
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='users' AND column_name='birthday'
                        ) THEN
                            ALTER TABLE users ADD COLUMN birthday DATE;
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add birthday column: {e}")
                conn.rollback()

            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='users' AND column_name='bank_name'
                        ) THEN
                            ALTER TABLE users ADD COLUMN bank_name VARCHAR(255);
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add bank_name column: {e}")
                conn.rollback()

            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='users' AND column_name='sort_code'
                        ) THEN
                            ALTER TABLE users ADD COLUMN sort_code VARCHAR(20);
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add sort_code column: {e}")
                conn.rollback()

            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='users' AND column_name='account_number'
                        ) THEN
                            ALTER TABLE users ADD COLUMN account_number VARCHAR(20);
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add account_number column: {e}")
                conn.rollback()

            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='users' AND column_name='theme_preference'
                        ) THEN
                            ALTER TABLE users ADD COLUMN theme_preference VARCHAR(50) DEFAULT 'nordic_neutral';
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add theme_preference column: {e}")
                conn.rollback()
            
            # Add session_cost column to practice_sessions if it doesn't exist
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_sessions' AND column_name='session_cost'
                        ) THEN
                            ALTER TABLE practice_sessions ADD COLUMN session_cost DECIMAL(10, 2);
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add session_cost column: {e}")
                conn.rollback()

            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_sessions' AND column_name='event_type'
                        ) THEN
                            ALTER TABLE practice_sessions ADD COLUMN event_type VARCHAR(50) DEFAULT 'practice';
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add event_type column: {e}")
                conn.rollback()

            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_sessions' AND column_name='event_title'
                        ) THEN
                            ALTER TABLE practice_sessions ADD COLUMN event_title VARCHAR(50);
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add event_title column: {e}")
                conn.rollback()
            
            # Add paid_by column to practice_sessions if it doesn't exist
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_sessions' AND column_name='paid_by'
                        ) THEN
                            ALTER TABLE practice_sessions ADD COLUMN paid_by VARCHAR(255);
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add paid_by column: {e}")
                conn.rollback()
            
            # Add payment_requested column to practice_sessions if it doesn't exist
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_sessions' AND column_name='payment_requested'
                        ) THEN
                            ALTER TABLE practice_sessions ADD COLUMN payment_requested BOOLEAN DEFAULT FALSE;
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add payment_requested column: {e}")
                conn.rollback()

            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_sessions' AND column_name='payment_requested_at'
                        ) THEN
                            ALTER TABLE practice_sessions ADD COLUMN payment_requested_at TIMESTAMP;
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add payment_requested_at column: {e}")
                conn.rollback()

            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_sessions' AND column_name='maximum_capacity'
                        ) THEN
                            ALTER TABLE practice_sessions ADD COLUMN maximum_capacity INTEGER DEFAULT 100;
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add maximum_capacity column: {e}")
                conn.rollback()

            try:
                cur.execute("UPDATE practice_sessions SET maximum_capacity = 100 WHERE maximum_capacity IS NULL")
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not backfill maximum_capacity values: {e}")
                conn.rollback()

            try:
                cur.execute("UPDATE practice_sessions SET event_type = 'practice' WHERE event_type IS NULL OR TRIM(event_type) = ''")
                cur.execute("UPDATE practice_sessions SET event_title = 'Session' WHERE event_title IS NULL OR TRIM(event_title) = ''")
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not backfill practice event metadata: {e}")
                conn.rollback()

            try:
                backfill_practice_times(cur, conn)
            except Exception as e:
                print(f"Warning: Could not normalize practice time values: {e}")
                conn.rollback()
            
            # Add user_full_name to forum_posts if it doesn't exist
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='forum_posts' AND column_name='user_full_name'
                        ) THEN
                            ALTER TABLE forum_posts ADD COLUMN user_full_name VARCHAR(255);
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add user_full_name to forum_posts: {e}")
                conn.rollback()
            
            # Add user_full_name to forum_comments if it doesn't exist
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='forum_comments' AND column_name='user_full_name'
                        ) THEN
                            ALTER TABLE forum_comments ADD COLUMN user_full_name VARCHAR(255);
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add user_full_name to forum_comments: {e}")
                conn.rollback()

            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='event_comments' AND column_name='user_full_name'
                        ) THEN
                            ALTER TABLE event_comments ADD COLUMN user_full_name VARCHAR(255);
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add user_full_name to event_comments: {e}")
                conn.rollback()

            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_availability' AND column_name='user_full_name'
                        ) THEN
                            ALTER TABLE practice_availability ADD COLUMN user_full_name VARCHAR(255);
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add user_full_name to practice_availability: {e}")
                conn.rollback()
            
            # Add related_date to notifications if it doesn't exist
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='notifications' AND column_name='related_date'
                        ) THEN
                            ALTER TABLE notifications ADD COLUMN related_date DATE;
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add related_date to notifications: {e}")
                conn.rollback()
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='notifications' AND column_name='practice_session_id'
                        ) THEN
                            ALTER TABLE notifications ADD COLUMN practice_session_id BIGINT;
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add practice_session_id to notifications: {e}")
                conn.rollback()
            cur.execute("""
            CREATE TABLE IF NOT EXISTS events (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) NOT NULL,
                date VARCHAR(50) NOT NULL,
                time VARCHAR(50),
                location TEXT,
                description TEXT,
                image_url TEXT,
                youtube_url TEXT
            )
            """)
            try:
                cur.execute("""
                    DO $$
                    BEGIN
                        IF EXISTS (
                            SELECT 1 FROM information_schema.columns
                            WHERE table_name='events' AND column_name='type'
                        ) THEN
                            ALTER TABLE events DROP COLUMN type;
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not drop legacy events.type column: {e}")
                conn.rollback()
            cur.execute("""
            CREATE TABLE IF NOT EXISTS event_media (
                id SERIAL PRIMARY KEY,
                event_id INTEGER REFERENCES events(id),
                media_url TEXT
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS event_likes (
                id SERIAL PRIMARY KEY,
                event_id INTEGER REFERENCES events(id),
                user_email VARCHAR(255)
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS event_comments (
                id SERIAL PRIMARY KEY,
                event_id INTEGER REFERENCES events(id),
                user_email VARCHAR(255),
                user_full_name VARCHAR(255),
                comment TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS practice_availability (
                id SERIAL PRIMARY KEY,
                practice_session_id BIGINT,
                date VARCHAR(50) NOT NULL,
                user_email VARCHAR(255),
                user_full_name VARCHAR(255),
                status VARCHAR(50) NOT NULL CHECK(status IN ('available', 'tentative', 'not_available')),
                option_choice VARCHAR(50),
                UNIQUE(practice_session_id, user_email)
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS practice_sessions (
                id SERIAL PRIMARY KEY,
                date VARCHAR(50) NOT NULL,
                time VARCHAR(50),
                location TEXT,
                event_type VARCHAR(50) DEFAULT 'practice',
                event_title VARCHAR(50),
                description TEXT,
                image_url TEXT,
                youtube_url TEXT,
                option_a_text TEXT,
                option_b_text TEXT,
                session_cost DECIMAL(10, 2),
                paid_by VARCHAR(255),
                payment_requested BOOLEAN DEFAULT FALSE,
                payment_requested_at TIMESTAMP,
                maximum_capacity INTEGER DEFAULT 100
            )
            """)
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_sessions' AND column_name='description'
                        ) THEN
                            ALTER TABLE practice_sessions ADD COLUMN description TEXT;
                        END IF;
                    END $$;
                """)
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_sessions' AND column_name='image_url'
                        ) THEN
                            ALTER TABLE practice_sessions ADD COLUMN image_url TEXT;
                        END IF;
                    END $$;
                """)
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_sessions' AND column_name='youtube_url'
                        ) THEN
                            ALTER TABLE practice_sessions ADD COLUMN youtube_url TEXT;
                        END IF;
                    END $$;
                """)
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_sessions' AND column_name='option_a_text'
                        ) THEN
                            ALTER TABLE practice_sessions ADD COLUMN option_a_text TEXT;
                        END IF;
                    END $$;
                """)
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_sessions' AND column_name='option_b_text'
                        ) THEN
                            ALTER TABLE practice_sessions ADD COLUMN option_b_text TEXT;
                        END IF;
                    END $$;
                """)
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='practice_availability' AND column_name='option_choice'
                        ) THEN
                            ALTER TABLE practice_availability ADD COLUMN option_choice VARCHAR(50);
                        END IF;
                    END $$;
                """)
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='events' AND column_name='practice_session_date'
                        ) THEN
                            ALTER TABLE events ADD COLUMN practice_session_date TEXT;
                        END IF;
                    END $$;
                """)
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add extended practice session columns: {e}")
                conn.rollback()
            cur.execute("""
            CREATE TABLE IF NOT EXISTS practice_payments (
                id SERIAL PRIMARY KEY,
                practice_session_id BIGINT,
                date VARCHAR(50) NOT NULL,
                user_email VARCHAR(255) NOT NULL,
                paid BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(practice_session_id, user_email)
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS expenses (
                id SERIAL PRIMARY KEY,
                title VARCHAR(255) NOT NULL,
                amount DECIMAL(10, 2) NOT NULL,
                paid_by VARCHAR(255),
                expense_date VARCHAR(50) NOT NULL,
                category VARCHAR(100),
                payment_method VARCHAR(100),
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS forum_posts (
                id SERIAL PRIMARY KEY,
                user_email VARCHAR(255),
                user_full_name VARCHAR(255),
                content TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS forum_likes (
                id SERIAL PRIMARY KEY,
                post_id INTEGER REFERENCES forum_posts(id),
                user_email VARCHAR(255),
                UNIQUE(post_id, user_email)
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS forum_comments (
                id SERIAL PRIMARY KEY,
                post_id INTEGER REFERENCES forum_posts(id),
                user_email VARCHAR(255),
                user_full_name VARCHAR(255),
                comment TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS notifications (
                id SERIAL PRIMARY KEY,
                user_email VARCHAR(255),
                type VARCHAR(50) NOT NULL,
                message TEXT NOT NULL,
                read BOOLEAN DEFAULT FALSE,
                related_date DATE,
                practice_session_id BIGINT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS password_reset_tokens (
                id SERIAL PRIMARY KEY,
                user_email VARCHAR(255) NOT NULL,
                token VARCHAR(255) UNIQUE NOT NULL,
                expires_at TIMESTAMP NOT NULL,
                used BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS auth_sessions (
                id SERIAL PRIMARY KEY,
                token VARCHAR(255) UNIQUE NOT NULL,
                user_email VARCHAR(255) NOT NULL,
                user_full_name VARCHAR(255) NOT NULL,
                user_id INTEGER NOT NULL,
                expires_at TIMESTAMP NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS notification_channel_history (
                notif_type VARCHAR(50) NOT NULL,
                channel VARCHAR(50) NOT NULL,
                last_sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (notif_type, channel)
            )
            """)
            try:
                cur.execute("""
                    DO $$ 
                    BEGIN 
                        IF NOT EXISTS (
                            SELECT 1 FROM information_schema.columns 
                            WHERE table_name='auth_sessions' AND column_name='expires_at'
                        ) THEN
                            ALTER TABLE auth_sessions ADD COLUMN expires_at TIMESTAMP;
                        END IF;
                    END $$;
                """)
                cur.execute("UPDATE auth_sessions SET expires_at = COALESCE(expires_at, created_at + INTERVAL '90 days', CURRENT_TIMESTAMP + INTERVAL '90 days')")
                cur.execute("ALTER TABLE auth_sessions ALTER COLUMN expires_at SET NOT NULL")
                conn.commit()
            except Exception as e:
                print(f"Warning: Could not add auth session expiry column: {e}")
                conn.rollback()
            conn.commit()

            ensure_practice_session_ids(cur, conn)

        else:
            # SQLite version (for local development)
            cur = conn.cursor()
            try:
                cur.execute("ALTER TABLE users DROP COLUMN password_hash")
            except sqlite3.OperationalError:
                pass
            try:
                cur.execute("ALTER TABLE users ADD COLUMN theme_preference TEXT DEFAULT 'nordic_neutral'")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add theme_preference column: {e}")
            cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT UNIQUE NOT NULL,
                full_name TEXT NOT NULL,
                password TEXT NOT NULL,
                user_type TEXT DEFAULT 'member',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_login TIMESTAMP,
                birthday DATE,
                bank_name TEXT,
                sort_code TEXT,
                account_number TEXT,
                theme_preference TEXT DEFAULT 'nordic_neutral',
                is_deleted BOOLEAN DEFAULT 0,
                deleted_at TIMESTAMP,
                deleted_by TEXT
            )
            """)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS auth_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                token TEXT UNIQUE NOT NULL,
                user_email TEXT NOT NULL,
                user_full_name TEXT NOT NULL,
                user_id INTEGER NOT NULL,
                expires_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """)
            try:
                cur.execute("ALTER TABLE auth_sessions ADD COLUMN expires_at TIMESTAMP")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add auth session expiry column: {e}")
            try:
                cur.execute("UPDATE auth_sessions SET expires_at = datetime(COALESCE(created_at, CURRENT_TIMESTAMP), '+90 days') WHERE expires_at IS NULL")
            except sqlite3.OperationalError as e:
                print(f"Warning: Could not backfill auth session expiry values: {e}")
            try:
                cur.execute("ALTER TABLE users ADD COLUMN password TEXT")
            except sqlite3.OperationalError:
                pass
            # Add user_type column if it doesn't exist
            try:
                cur.execute("ALTER TABLE users ADD COLUMN user_type TEXT DEFAULT 'member'")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add user_type column: {e}")
            # Add created_at column if it doesn't exist (without default due to SQLite limitation)
            try:
                cur.execute("ALTER TABLE users ADD COLUMN created_at TIMESTAMP")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add created_at column: {e}")
            # Update NULL created_at values for existing users
            try:
                cur.execute("UPDATE users SET created_at = CURRENT_TIMESTAMP WHERE created_at IS NULL")
            except sqlite3.OperationalError:
                pass
            # Add is_deleted column if it doesn't exist
            try:
                cur.execute("ALTER TABLE users ADD COLUMN is_deleted BOOLEAN DEFAULT 0")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add is_deleted column: {e}")
            # Add deleted_at column if it doesn't exist
            try:
                cur.execute("ALTER TABLE users ADD COLUMN deleted_at TIMESTAMP")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add deleted_at column: {e}")
            # Add deleted_by column if it doesn't exist
            try:
                cur.execute("ALTER TABLE users ADD COLUMN deleted_by TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add deleted_by column: {e}")
            # Add last_login column if it doesn't exist
            try:
                cur.execute("ALTER TABLE users ADD COLUMN last_login TIMESTAMP")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add last_login column: {e}")
            # Add birthday column if it doesn't exist
            try:
                cur.execute("ALTER TABLE users ADD COLUMN birthday DATE")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add birthday column: {e}")
            try:
                cur.execute("ALTER TABLE users ADD COLUMN bank_name TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add bank_name column: {e}")
            try:
                cur.execute("ALTER TABLE users ADD COLUMN sort_code TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add sort_code column: {e}")
            try:
                cur.execute("ALTER TABLE users ADD COLUMN account_number TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add account_number column: {e}")
            # Add session_cost column to practice_sessions if it doesn't exist
            try:
                cur.execute("ALTER TABLE practice_sessions ADD COLUMN session_cost REAL")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add session_cost column: {e}")
            try:
                cur.execute("ALTER TABLE practice_sessions ADD COLUMN event_type TEXT DEFAULT 'practice'")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add event_type column: {e}")
            try:
                cur.execute("ALTER TABLE practice_sessions ADD COLUMN event_title TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add event_title column: {e}")
            try:
                cur.execute("ALTER TABLE practice_sessions ADD COLUMN description TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add description column: {e}")
            try:
                cur.execute("ALTER TABLE practice_sessions ADD COLUMN image_url TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add image_url column: {e}")
            try:
                cur.execute("ALTER TABLE practice_sessions ADD COLUMN youtube_url TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add youtube_url column: {e}")
            try:
                cur.execute("ALTER TABLE practice_sessions ADD COLUMN option_a_text TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add option_a_text column: {e}")
            try:
                cur.execute("ALTER TABLE practice_sessions ADD COLUMN option_b_text TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add option_b_text column: {e}")
            # Add paid_by column to practice_sessions if it doesn't exist
            try:
                cur.execute("ALTER TABLE practice_sessions ADD COLUMN paid_by TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add paid_by column: {e}")
            # Add payment_requested column to practice_sessions if it doesn't exist
            try:
                cur.execute("ALTER TABLE practice_sessions ADD COLUMN payment_requested INTEGER DEFAULT 0")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add payment_requested column: {e}")
            try:
                cur.execute("ALTER TABLE practice_sessions ADD COLUMN payment_requested_at TIMESTAMP")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add payment_requested_at column: {e}")
            try:
                cur.execute("ALTER TABLE practice_sessions ADD COLUMN maximum_capacity INTEGER DEFAULT 100")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add maximum_capacity column: {e}")
            try:
                cur.execute("UPDATE practice_sessions SET event_type = 'practice' WHERE event_type IS NULL OR TRIM(event_type) = ''")
                cur.execute("UPDATE practice_sessions SET event_title = 'Session' WHERE event_title IS NULL OR TRIM(event_title) = ''")
            except sqlite3.OperationalError as e:
                print(f"Warning: Could not backfill practice event metadata: {e}")
            try:
                cur.execute("UPDATE practice_sessions SET maximum_capacity = 100 WHERE maximum_capacity IS NULL")
            except sqlite3.OperationalError as e:
                print(f"Warning: Could not backfill maximum_capacity values: {e}")
            # Add user_full_name to forum_posts if it doesn't exist
            try:
                cur.execute("ALTER TABLE forum_posts ADD COLUMN user_full_name TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add user_full_name to forum_posts: {e}")
            # Add user_full_name to forum_comments if it doesn't exist
            try:
                cur.execute("ALTER TABLE forum_comments ADD COLUMN user_full_name TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add user_full_name to forum_comments: {e}")
            # Add user_full_name to event_comments if it doesn't exist
            try:
                cur.execute("ALTER TABLE event_comments ADD COLUMN user_full_name TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add user_full_name to event_comments: {e}")
            # Add user_full_name to practice_availability if it doesn't exist
            try:
                cur.execute("ALTER TABLE practice_availability ADD COLUMN user_full_name TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add user_full_name to practice_availability: {e}")
            try:
                cur.execute("ALTER TABLE practice_availability ADD COLUMN option_choice TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add option_choice to practice_availability: {e}")
            try:
                cur.execute("ALTER TABLE events ADD COLUMN practice_session_date TEXT")
            except sqlite3.OperationalError as e:
                if "duplicate column name" not in str(e).lower():
                    print(f"Warning: Could not add practice_session_date to events: {e}")
            cur.executescript("""
            CREATE TABLE IF NOT EXISTS events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                date TEXT NOT NULL,
                time TEXT,
                location TEXT,
                description TEXT,
                image_url TEXT,
                youtube_url TEXT,
                practice_session_date TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS event_media (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_id INTEGER,
                media_url TEXT,
                FOREIGN KEY(event_id) REFERENCES events(id)
            );
            CREATE TABLE IF NOT EXISTS event_likes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_id INTEGER,
                user_email TEXT,
                FOREIGN KEY(event_id) REFERENCES events(id)
            );
            CREATE TABLE IF NOT EXISTS event_comments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_id INTEGER,
                user_email TEXT,
                user_full_name TEXT,
                comment TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(event_id) REFERENCES events(id)
            );
            CREATE TABLE IF NOT EXISTS practice_availability (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                practice_session_id INTEGER,
                date TEXT NOT NULL,
                user_email TEXT,
                user_full_name TEXT,
                status TEXT NOT NULL CHECK(status IN ('available', 'tentative', 'not_available')),
                option_choice TEXT,
                UNIQUE(practice_session_id, user_email)
            );
            CREATE TABLE IF NOT EXISTS practice_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                time TEXT,
                location TEXT,
                event_type TEXT DEFAULT 'practice',
                event_title TEXT,
                description TEXT,
                image_url TEXT,
                youtube_url TEXT,
                option_a_text TEXT,
                option_b_text TEXT,
                session_cost REAL,
                paid_by TEXT,
                payment_requested INTEGER DEFAULT 0,
                payment_requested_at TIMESTAMP,
                maximum_capacity INTEGER DEFAULT 100
            );
            CREATE TABLE IF NOT EXISTS practice_payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                practice_session_id INTEGER,
                date TEXT NOT NULL,
                user_email TEXT NOT NULL,
                paid INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(practice_session_id, user_email)
            );
            CREATE TABLE IF NOT EXISTS expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                amount REAL NOT NULL,
                paid_by TEXT,
                expense_date TEXT NOT NULL,
                category TEXT,
                payment_method TEXT,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS forum_posts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_email TEXT,
                user_full_name TEXT,
                content TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS forum_likes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                post_id INTEGER,
                user_email TEXT,
                UNIQUE(post_id, user_email),
                FOREIGN KEY(post_id) REFERENCES forum_posts(id)
            );
            CREATE TABLE IF NOT EXISTS forum_comments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                post_id INTEGER,
                user_email TEXT,
                user_full_name TEXT,
                comment TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(post_id) REFERENCES forum_posts(id)
            );
            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_email TEXT,
                type TEXT NOT NULL,
                message TEXT NOT NULL,
                read INTEGER DEFAULT 0,
                related_date DATE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS password_reset_tokens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_email TEXT NOT NULL,
                token TEXT UNIQUE NOT NULL,
                expires_at TIMESTAMP NOT NULL,
                used INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS notification_settings (
                notif_type TEXT PRIMARY KEY,
                display_name TEXT NOT NULL,
                description TEXT,
                app_enabled BOOLEAN DEFAULT TRUE,
                email_enabled INTEGER DEFAULT 0,
                whatsapp_enabled INTEGER DEFAULT 0,
                target_audience TEXT NOT NULL DEFAULT 'all_active_users',
                app_template TEXT NOT NULL,
                email_subject TEXT NOT NULL,
                email_template TEXT NOT NULL,
                whatsapp_template TEXT NOT NULL,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS notification_channel_history (
                notif_type TEXT NOT NULL,
                channel TEXT NOT NULL,
                last_sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (notif_type, channel)
            );
            """)
            try:
                cur.execute("PRAGMA table_info(events)")
                event_columns = [row[1] for row in cur.fetchall()]
                if "type" in event_columns:
                    cur.executescript("""
                    ALTER TABLE events RENAME TO events_legacy;
                    CREATE TABLE events (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL,
                        date TEXT NOT NULL,
                        time TEXT,
                        location TEXT,
                        description TEXT,
                        image_url TEXT,
                        youtube_url TEXT
                    );
                    INSERT INTO events (id, name, date, time, location, description, image_url, youtube_url)
                    SELECT id, name, date, time, location, description, image_url, youtube_url FROM events_legacy;
                    DROP TABLE events_legacy;
                    """)
            except sqlite3.OperationalError as e:
                print(f"Warning: Could not drop legacy events.type column: {e}")
            try:
                cur.execute("ALTER TABLE events ADD COLUMN image_url TEXT")
            except sqlite3.OperationalError:
                pass
            try:
                cur.execute("ALTER TABLE events ADD COLUMN youtube_url TEXT")
            except sqlite3.OperationalError:
                pass
            try:
                cur.execute("ALTER TABLE notifications ADD COLUMN related_date DATE")
            except sqlite3.OperationalError:
                pass
            try:
                cur.execute("ALTER TABLE notifications ADD COLUMN practice_session_id INTEGER")
            except sqlite3.OperationalError:
                pass
            ensure_practice_session_ids(cur, conn)
            conn.commit()

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password: str, hashed: str) -> bool:
    return hash_password(password) == hashed

def generate_password_reset_token() -> str:
    return secrets.token_urlsafe(32)

def build_password_reset_link(token: str) -> str:
    return f"{FRONTEND_URL.rstrip('/')}/reset-password?token={token}"

# ========== FORGOT PASSWORD FEATURE - Email Sending Function ==========
# This function sends emails using SMTP (Simple Mail Transfer Protocol)
# Used by the forgot password endpoint to send recovery emails to users
# Returns True if email sent successfully, False otherwise
# =======================================================================
def send_email(to_email: str, subject: str, body: str) -> bool:
    """Send email using SMTP"""
    if not SMTP_USERNAME or not SMTP_PASSWORD:
        print("Email not configured. Set SMTP_USERNAME and SMTP_PASSWORD environment variables.")
        return False
    
    try:
        msg = MIMEMultipart()
        msg['From'] = FROM_EMAIL
        msg['To'] = to_email
        msg['Subject'] = subject
        
        msg.attach(MIMEText(body, 'plain'))
        
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.send_message(msg)
        server.quit()
        
        return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False

def send_whatsapp_notification(message: str) -> bool:
    if not WHATSAPP_NOTIFICATIONS_ENABLED:
        return False
    result = send_group_message(message)
    if not result.get("success"):
        print(f"Failed to send WhatsApp message: {result.get('error', 'Unknown error')}")
        return False
    return True

def render_notification_template(template: str, context: dict) -> str:
    rendered = template or ""
    for key, value in context.items():
        rendered = rendered.replace(f"{{{{{key}}}}}", "" if value is None else str(value))
    return rendered

def format_notification_date(date_value: str) -> str:
    if not date_value:
        return ""
    try:
        parsed_date = datetime.strptime(date_value, "%Y-%m-%d")
    except ValueError:
        return date_value

    day = parsed_date.day
    if 10 <= day % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")

    return parsed_date.strftime(f"%A, {day}{suffix} %B %Y")

def build_notification_context(payload: dict) -> dict:
    date_value = payload.get("date") or ""
    formatted_date_value = format_notification_date(date_value)
    time_value = payload.get("time") or ""
    location_value = payload.get("location") or ""
    session_id_value = payload.get("session_id") or payload.get("practice_session_id") or ""
    content_value = (payload.get("content") or "").strip()
    content_preview = content_value[:180] + ("..." if len(content_value) > 180 else "")
    event_type_value = normalize_event_type(payload.get("event_type")) if payload.get("event_type") else "practice"
    event_title_value = normalize_event_title(payload.get("event_title"), event_type_value)
    event_type_label_value = default_event_type_label(event_type_value)
    event_type_template_value = event_type_value.upper()
    event_name_value = payload.get("event_name") or payload.get("name") or f"{event_type_label_value} - {event_title_value}"
    return {
        "date": formatted_date_value,
        "date_iso": date_value,
        "time": time_value,
        "location": location_value,
        "session_id": session_id_value,
        "maximum_capacity": payload.get("maximum_capacity") if payload.get("maximum_capacity") is not None else "",
        "available_count": payload.get("available_count") if payload.get("available_count") is not None else "",
        "remaining_slots": payload.get("remaining_slots") if payload.get("remaining_slots") is not None else "",
        "event_name": event_name_value,
        "event_type": event_type_template_value,
        "event_type_label": event_type_label_value,
        "event_title": event_title_value,
        "author_name": payload.get("author_name") or "",
        "full_name": payload.get("full_name") or "",
        "member_name": payload.get("member_name") or payload.get("full_name") or "",
        "club_name": payload.get("club_name") or "Glasgow Bengali FC",
        "payments_link": payload.get("payments_link") or "https://glasgow-bengali-fc.vercel.app/user-actions/payments",
        "content": content_value,
        "content_preview": content_preview,
        "time_suffix": f" at {time_value}" if time_value else "",
        "location_suffix": f" at {location_value}" if location_value else "",
        "location_comma_suffix": f", {location_value}" if location_value else "",
        "time_line": f"🕐 {time_value}\n" if time_value else "",
        "location_line": f"📍 {location_value}\n" if location_value else "",
    }

def seed_notification_settings():
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"DELETE FROM notification_settings WHERE notif_type = {PLACEHOLDER}", ("match",))
        for notif_type, defaults in NOTIFICATION_TYPE_DEFAULTS.items():
            if USE_POSTGRES:
                cur.execute(
                    f"""
                    INSERT INTO notification_settings (
                        notif_type, display_name, description, app_enabled, email_enabled, whatsapp_enabled,
                        target_audience, app_template, email_subject, email_template, whatsapp_template
                    ) VALUES (
                        {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER},
                        {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}
                    )
                    ON CONFLICT (notif_type) DO NOTHING
                    """,
                    (
                        notif_type,
                        defaults["display_name"],
                        defaults["description"],
                        defaults["app_enabled"],
                        defaults["email_enabled"],
                        defaults["whatsapp_enabled"],
                        defaults["target_audience"],
                        defaults["app_template"],
                        defaults["email_subject"],
                        defaults["email_template"],
                        defaults["whatsapp_template"],
                    ),
                )
            else:
                cur.execute(
                    """
                    INSERT OR IGNORE INTO notification_settings (
                        notif_type, display_name, description, app_enabled, email_enabled, whatsapp_enabled,
                        target_audience, app_template, email_subject, email_template, whatsapp_template
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        notif_type,
                        defaults["display_name"],
                        defaults["description"],
                        1 if defaults["app_enabled"] else 0,
                        1 if defaults["email_enabled"] else 0,
                        1 if defaults["whatsapp_enabled"] else 0,
                        defaults["target_audience"],
                        defaults["app_template"],
                        defaults["email_subject"],
                        defaults["email_template"],
                        defaults["whatsapp_template"],
                    ),
                )
        conn.commit()

        for notif_type, defaults in NOTIFICATION_TYPE_DEFAULTS.items():
            if USE_POSTGRES:
                cur.execute(
                    f"""
                    UPDATE notification_settings
                    SET display_name = {PLACEHOLDER},
                        description = {PLACEHOLDER},
                        app_template = CASE WHEN app_template IS NULL OR TRIM(app_template) = '' OR app_template = {PLACEHOLDER} THEN {PLACEHOLDER} ELSE app_template END,
                        email_subject = CASE WHEN email_subject IS NULL OR TRIM(email_subject) = '' OR email_subject = {PLACEHOLDER} THEN {PLACEHOLDER} ELSE email_subject END,
                        email_template = CASE WHEN email_template IS NULL OR TRIM(email_template) = '' OR email_template = {PLACEHOLDER} THEN {PLACEHOLDER} ELSE email_template END,
                        whatsapp_template = CASE WHEN whatsapp_template IS NULL OR TRIM(whatsapp_template) = '' OR whatsapp_template = {PLACEHOLDER} THEN {PLACEHOLDER} ELSE whatsapp_template END
                    WHERE notif_type = {PLACEHOLDER}
                    """,
                    (
                        defaults["display_name"],
                        defaults["description"],
                        defaults["app_template"],
                        defaults["app_template"],
                        defaults["email_subject"],
                        defaults["email_subject"],
                        defaults["email_template"],
                        defaults["email_template"],
                        defaults["whatsapp_template"],
                        defaults["whatsapp_template"],
                        notif_type,
                    ),
                )
            else:
                cur.execute(
                    """
                    UPDATE notification_settings
                    SET display_name = ?,
                        description = ?,
                        app_template = CASE WHEN app_template IS NULL OR TRIM(app_template) = '' OR app_template = ? THEN ? ELSE app_template END,
                        email_subject = CASE WHEN email_subject IS NULL OR TRIM(email_subject) = '' OR email_subject = ? THEN ? ELSE email_subject END,
                        email_template = CASE WHEN email_template IS NULL OR TRIM(email_template) = '' OR email_template = ? THEN ? ELSE email_template END,
                        whatsapp_template = CASE WHEN whatsapp_template IS NULL OR TRIM(whatsapp_template) = '' OR whatsapp_template = ? THEN ? ELSE whatsapp_template END
                    WHERE notif_type = ?
                    """,
                    (
                        defaults["display_name"],
                        defaults["description"],
                        defaults["app_template"],
                        defaults["app_template"],
                        defaults["email_subject"],
                        defaults["email_subject"],
                        defaults["email_template"],
                        defaults["email_template"],
                        defaults["whatsapp_template"],
                        defaults["whatsapp_template"],
                        notif_type,
                    ),
                )
        conn.commit()

def get_notification_settings_map() -> dict:
    seed_notification_settings()
    settings_map = {}
    for notif_type, defaults in NOTIFICATION_TYPE_DEFAULTS.items():
        settings_map[notif_type] = {
            "notif_type": notif_type,
            "display_name": defaults["display_name"],
            "description": defaults["description"],
            "app_enabled": defaults["app_enabled"],
            "email_enabled": defaults["email_enabled"],
            "whatsapp_enabled": defaults["whatsapp_enabled"],
            "target_audience": defaults["target_audience"],
            "app_template": defaults["app_template"],
            "email_subject": defaults["email_subject"],
            "email_template": defaults["email_template"],
            "whatsapp_template": defaults["whatsapp_template"],
            "updated_at": None,
        }
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM notification_settings ORDER BY display_name ASC")
        rows = cur.fetchall()
        for row in rows:
            row_dict = dict(row)
            settings_map[row_dict["notif_type"]] = {
                **settings_map.get(row_dict["notif_type"], {}),
                **row_dict,
            }
        return settings_map

def get_notification_setting(notif_type: str) -> dict:
    setting = get_notification_settings_map().get(notif_type)
    if not setting:
        raise HTTPException(status_code=404, detail="Notification type not found")
    return setting

def resolve_notification_recipients(target_audience: str, payload: dict, notif_type: str = None) -> list:
    with get_connection() as conn:
        cur = conn.cursor()
        if target_audience == "direct_user":
            recipient_email = (payload.get("email") or "").strip().lower()
            if not recipient_email:
                return []
            cur.execute(
                f"SELECT email, full_name FROM users WHERE email = {PLACEHOLDER} AND (is_deleted = FALSE OR is_deleted IS NULL)",
                (recipient_email,),
            )
        elif target_audience == "admin_users":
            cur.execute(
                f"SELECT email, full_name FROM users WHERE (is_deleted = FALSE OR is_deleted IS NULL) AND user_type = {PLACEHOLDER}",
                ("admin",),
            )
        elif target_audience == "available_players":
            if notif_type == "payment_request":
                cur.execute(
                    f"""
                    SELECT DISTINCT u.email, u.full_name
                    FROM practice_availability pa
                    JOIN users u ON pa.user_email = u.email
                    WHERE pa.practice_session_id = {PLACEHOLDER}
                      AND pa.status = {PLACEHOLDER}
                      AND (u.is_deleted = FALSE OR u.is_deleted IS NULL)
                    """,
                    (payload["session_id"], "available"),
                )
            elif not payload.get("date"):
                return []
            else:
                cur.execute(
                    f"""
                    SELECT u.email, u.full_name
                    FROM practice_availability pa
                    JOIN users u ON pa.user_email = u.email
                    WHERE pa.date = {PLACEHOLDER}
                      AND pa.status = {PLACEHOLDER}
                      AND (u.is_deleted = FALSE OR u.is_deleted IS NULL)
                    """,
                    (payload["date"], "available"),
                )
        elif notif_type == "pending_payment_reminder":
            cur.execute(
                f"""
                SELECT DISTINCT u.email, u.full_name
                FROM practice_sessions ps
                JOIN practice_availability pa ON ps.date = pa.date
                JOIN users u ON pa.user_email = u.email
                LEFT JOIN practice_payments pp ON pa.date = pp.date AND pa.user_email = pp.user_email
                WHERE ps.payment_requested = {PLACEHOLDER}
                  AND ps.payment_requested_at IS NOT NULL
                  AND pa.status = {PLACEHOLDER}
                  AND (pp.paid IS NULL OR pp.paid = {PLACEHOLDER})
                  AND (u.is_deleted = FALSE OR u.is_deleted IS NULL)
                """,
                (True if USE_POSTGRES else 1, "available", False if USE_POSTGRES else 0),
            )
            recipients = []
            for row in cur.fetchall():
                row_dict = dict(row)
                cur.execute(
                    f"SELECT date, time FROM practice_sessions WHERE payment_requested = {PLACEHOLDER} AND payment_requested_at IS NOT NULL AND date IN (SELECT date FROM practice_availability WHERE user_email = {PLACEHOLDER} AND status = {PLACEHOLDER}) ORDER BY date DESC",
                    (True if USE_POSTGRES else 1, row_dict["email"], "available"),
                )
                user_sessions = [dict(session_row) for session_row in cur.fetchall()]
                if any(is_practice_datetime_in_past(session["date"], session.get("time")) for session in user_sessions):
                    recipients.append(row_dict)
            return recipients
        else:
            cur.execute("SELECT email, full_name FROM users WHERE (is_deleted = FALSE OR is_deleted IS NULL)")
        return [dict(row) for row in cur.fetchall()]

def send_direct_notification_email(notif_type: str, payload: dict, recipient_email: str):
    setting = get_notification_setting(notif_type)
    if not setting["email_enabled"]:
        return
    context = build_notification_context(payload)
    subject = render_notification_template(setting["email_subject"], context)
    email_body = render_notification_template(setting["email_template"], context)
    send_email(recipient_email, subject, email_body)

def send_direct_notification_email_safe(notif_type: str, payload: dict, recipient_email: str):
    try:
        send_direct_notification_email(notif_type, payload, recipient_email)
    except Exception as exc:
        print(f"Background email send failed for {notif_type} to {recipient_email}: {exc}")

def get_notification_channel_last_sent_at(cur, notif_type: str, channel: str):
    cur.execute(
        f"SELECT last_sent_at FROM notification_channel_history WHERE notif_type = {PLACEHOLDER} AND channel = {PLACEHOLDER}",
        (notif_type, channel),
    )
    row = cur.fetchone()
    if not row:
        return None
    row_dict = dict(row)
    last_sent_at = row_dict.get("last_sent_at")
    if isinstance(last_sent_at, str):
        try:
            return datetime.fromisoformat(last_sent_at.replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            return None
    return last_sent_at

def record_notification_channel_sent(cur, notif_type: str, channel: str):
    if USE_POSTGRES:
        cur.execute(
            f"""
            INSERT INTO notification_channel_history (notif_type, channel, last_sent_at)
            VALUES ({PLACEHOLDER}, {PLACEHOLDER}, CURRENT_TIMESTAMP)
            ON CONFLICT (notif_type, channel)
            DO UPDATE SET last_sent_at = CURRENT_TIMESTAMP
            """,
            (notif_type, channel),
        )
    else:
        cur.execute(
            """
            INSERT INTO notification_channel_history (notif_type, channel, last_sent_at)
            VALUES (?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(notif_type, channel) DO UPDATE SET last_sent_at = CURRENT_TIMESTAMP
            """,
            (notif_type, channel),
        )

def should_send_whatsapp_notification(cur, notif_type: str) -> bool:
    if notif_type != "pending_payment_reminder":
        return True
    last_sent_at = get_notification_channel_last_sent_at(cur, notif_type, "whatsapp")
    if not last_sent_at:
        return True
    return datetime.now() >= last_sent_at + timedelta(days=5)

def should_send_app_notification(cur, notif_type: str) -> bool:
    if notif_type != "pending_payment_reminder":
        return True
    last_sent_at = get_notification_channel_last_sent_at(cur, notif_type, "app")
    if not last_sent_at:
        return True
    return datetime.now() >= last_sent_at + timedelta(days=5)

def deliver_notification(notif_type: str, payload: dict, related_date: str = None, exclude_email: str = None):
    guarded_notif_types = {"practice", "match", "practice_slot_available", "session_capacity_reached"}
    effective_date = related_date or payload.get("date")
    if notif_type in guarded_notif_types and effective_date:
        try:
            effective_time = payload.get("time") if notif_type in {"practice", "practice_slot_available", "session_capacity_reached"} else None
            if notif_type in {"practice", "practice_slot_available", "session_capacity_reached"}:
                if is_practice_datetime_in_past(effective_date, effective_time):
                    return
            else:
                notification_date = datetime.strptime(effective_date, "%Y-%m-%d").date()
                if notification_date < datetime.now().date():
                    return
        except HTTPException:
            return
        except ValueError:
            pass

    setting = get_notification_setting(notif_type)
    context = build_notification_context(payload)
    recipients = resolve_notification_recipients(setting["target_audience"], payload, notif_type)
    if exclude_email:
        recipients = [recipient for recipient in recipients if recipient["email"] != exclude_email]

    with get_connection() as conn:
        cur = conn.cursor()
        practice_session_id = payload.get("session_id") or payload.get("practice_session_id")

        if setting["app_enabled"]:
            if should_send_app_notification(cur, notif_type):
                app_message = render_notification_template(setting["app_template"], context)
                for recipient in recipients:
                    create_notification(recipient["email"], notif_type, app_message, related_date, practice_session_id)
                record_notification_channel_sent(cur, notif_type, "app")

        if setting["email_enabled"]:
            subject = render_notification_template(setting["email_subject"], context)
            email_body = render_notification_template(setting["email_template"], context)
            for recipient in recipients:
                send_email(recipient["email"], subject, email_body)

        if setting["whatsapp_enabled"]:
            if should_send_whatsapp_notification(cur, notif_type):
                whatsapp_message = render_notification_template(setting["whatsapp_template"], context)
                if send_whatsapp_notification(whatsapp_message):
                    record_notification_channel_sent(cur, notif_type, "whatsapp")

        conn.commit()

def normalize_maximum_capacity(value) -> int:
    if value is None or value == "":
        return 100
    try:
        normalized = int(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Maximum capacity must be a whole number")
    if normalized <= 0:
        raise HTTPException(status_code=400, detail="Maximum capacity must be greater than 0")
    return normalized

def normalize_practice_time(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    raw_value = str(value).strip()
    if not raw_value:
        return None

    supported_formats = [
        "%H:%M",
        "%H:%M:%S",
        "%I:%M %p",
        "%I:%M%p",
        "%I %p",
        "%I%p",
    ]

    for time_format in supported_formats:
        try:
            parsed_time = datetime.strptime(raw_value.upper(), time_format)
            return parsed_time.strftime("%H:%M")
        except ValueError:
            continue

    raise HTTPException(status_code=400, detail="Practice time must be a valid time")

def backfill_practice_times(cur, conn):
    cur.execute(f"SELECT date, time FROM practice_sessions WHERE time IS NOT NULL AND TRIM(time) <> ''")
    rows = cur.fetchall()
    for row in rows:
        row_dict = dict(row)
        try:
            normalized_time = normalize_practice_time(row_dict.get("time"))
        except HTTPException:
            normalized_time = "21:00"
        if normalized_time != row_dict.get("time"):
            cur.execute(
                f"UPDATE practice_sessions SET time = {PLACEHOLDER} WHERE date = {PLACEHOLDER}",
                (normalized_time, row_dict["date"]),
            )
    conn.commit()

def ensure_postgres_practice_session_ids(cur, conn):
    try:
        cur.execute(
            """
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name='practice_sessions' AND column_name='id'
                ) THEN
                    ALTER TABLE practice_sessions ADD COLUMN id BIGINT;
                END IF;
            END $$;
            """
        )
        cur.execute("CREATE SEQUENCE IF NOT EXISTS practice_sessions_id_seq")
        cur.execute("ALTER SEQUENCE practice_sessions_id_seq OWNED BY practice_sessions.id")
        cur.execute("SELECT setval('practice_sessions_id_seq', COALESCE((SELECT MAX(id) FROM practice_sessions), 0) + 1, false)")
        cur.execute("UPDATE practice_sessions SET id = nextval('practice_sessions_id_seq') WHERE id IS NULL")
        cur.execute("ALTER TABLE practice_sessions ALTER COLUMN id SET NOT NULL")
        cur.execute("ALTER TABLE practice_sessions ALTER COLUMN id SET DEFAULT nextval('practice_sessions_id_seq')")
        cur.execute(
            """
            DO $$
            BEGIN
                IF EXISTS (
                    SELECT 1
                    FROM pg_constraint c
                    JOIN pg_class t ON t.oid = c.conrelid
                    WHERE t.relname = 'practice_sessions'
                      AND c.contype = 'p'
                      AND c.conname <> 'practice_sessions_pkey'
                ) THEN
                    EXECUTE (
                        SELECT 'ALTER TABLE practice_sessions DROP CONSTRAINT ' || quote_ident(c.conname)
                        FROM pg_constraint c
                        JOIN pg_class t ON t.oid = c.conrelid
                        WHERE t.relname = 'practice_sessions'
                          AND c.contype = 'p'
                          AND c.conname <> 'practice_sessions_pkey'
                        LIMIT 1
                    );
                END IF;

                IF EXISTS (
                    SELECT 1
                    FROM pg_constraint c
                    JOIN pg_class t ON t.oid = c.conrelid
                    JOIN pg_attribute a ON a.attrelid = t.oid AND a.attnum = ANY(c.conkey)
                    WHERE t.relname = 'practice_sessions'
                      AND c.contype = 'p'
                      AND a.attname <> 'id'
                ) THEN
                    EXECUTE (
                        SELECT 'ALTER TABLE practice_sessions DROP CONSTRAINT ' || quote_ident(c.conname)
                        FROM pg_constraint c
                        JOIN pg_class t ON t.oid = c.conrelid
                        JOIN pg_attribute a ON a.attrelid = t.oid AND a.attnum = ANY(c.conkey)
                        WHERE t.relname = 'practice_sessions'
                          AND c.contype = 'p'
                          AND a.attname <> 'id'
                        LIMIT 1
                    );
                END IF;

                IF NOT EXISTS (
                    SELECT 1
                    FROM pg_constraint c
                    JOIN pg_class t ON t.oid = c.conrelid
                    JOIN pg_attribute a ON a.attrelid = t.oid AND a.attnum = ANY(c.conkey)
                    WHERE t.relname = 'practice_sessions'
                      AND c.contype = 'p'
                      AND a.attname = 'id'
                ) THEN
                    ALTER TABLE practice_sessions ADD CONSTRAINT practice_sessions_pkey PRIMARY KEY (id);
                END IF;

                IF NOT EXISTS (
                    SELECT 1 FROM pg_constraint
                    WHERE conname = 'practice_sessions_id_key'
                ) THEN
                    ALTER TABLE practice_sessions ADD CONSTRAINT practice_sessions_id_key UNIQUE (id);
                END IF;
            END $$;
            """
        )
        cur.execute("CREATE INDEX IF NOT EXISTS idx_practice_sessions_id ON practice_sessions(id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_practice_sessions_date ON practice_sessions(date)")
        try:
            cur.execute("ALTER TABLE practice_sessions DROP CONSTRAINT IF EXISTS practice_sessions_date_key")
            conn.commit()
        except Exception:
            conn.rollback()
        cur.execute(
            """
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name='practice_availability' AND column_name='practice_session_id'
                ) THEN
                    ALTER TABLE practice_availability ADD COLUMN practice_session_id BIGINT;
                END IF;
            END $$;
            """
        )
        cur.execute(
            """
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name='practice_payments' AND column_name='practice_session_id'
                ) THEN
                    ALTER TABLE practice_payments ADD COLUMN practice_session_id BIGINT;
                END IF;
            END $$;
            """
        )
        cur.execute(
            """
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name='events' AND column_name='practice_session_id'
                ) THEN
                    ALTER TABLE events ADD COLUMN practice_session_id BIGINT;
                END IF;
            END $$;
            """
        )
        cur.execute(
            f"""
            UPDATE practice_availability pa
            SET practice_session_id = ps.id
            FROM practice_sessions ps
            WHERE pa.practice_session_id IS NULL
              AND pa.date = ps.date
            """
        )
        cur.execute(
            f"""
            UPDATE practice_payments pp
            SET practice_session_id = ps.id
            FROM practice_sessions ps
            WHERE pp.practice_session_id IS NULL
              AND pp.date = ps.date
            """
        )
        cur.execute(
            f"""
            UPDATE events e
            SET practice_session_id = ps.id
            FROM practice_sessions ps
            WHERE e.practice_session_id IS NULL
              AND e.practice_session_date = ps.date
            """
        )
        cur.execute("CREATE INDEX IF NOT EXISTS idx_practice_availability_session_id ON practice_availability(practice_session_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_practice_payments_session_id ON practice_payments(practice_session_id)")
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_events_practice_session_id_unique ON events(practice_session_id) WHERE practice_session_id IS NOT NULL")
        try:
            cur.execute("ALTER TABLE practice_payments DROP CONSTRAINT IF EXISTS practice_payments_date_user_email_key")
            conn.commit()
        except Exception:
            conn.rollback()
        cur.execute(
            """
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM pg_constraint
                    WHERE conname = 'practice_payments_session_user_key'
                ) THEN
                    ALTER TABLE practice_payments ADD CONSTRAINT practice_payments_session_user_key UNIQUE (practice_session_id, user_email);
                END IF;
            END $$;
            """
        )
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_practice_payments_session_user_unique ON practice_payments(practice_session_id, user_email) WHERE practice_session_id IS NOT NULL AND user_email IS NOT NULL")
        try:
            cur.execute("ALTER TABLE practice_availability DROP CONSTRAINT IF EXISTS practice_availability_date_user_email_key")
            conn.commit()
        except Exception:
            conn.rollback()
        cur.execute(
            """
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM pg_constraint
                    WHERE conname = 'practice_availability_session_user_key'
                ) THEN
                    ALTER TABLE practice_availability ADD CONSTRAINT practice_availability_session_user_key UNIQUE (practice_session_id, user_email);
                END IF;
            END $$;
            """
        )
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_practice_availability_session_user_unique ON practice_availability(practice_session_id, user_email) WHERE practice_session_id IS NOT NULL AND user_email IS NOT NULL")
        try:
            cur.execute("ALTER TABLE events DROP CONSTRAINT IF EXISTS events_practice_session_date_key")
            conn.commit()
        except Exception:
            conn.rollback()
        conn.commit()
    except Exception as e:
        print(f"Warning: Could not migrate practice session IDs for PostgreSQL: {e}")
        conn.rollback()

def ensure_sqlite_practice_session_ids(cur, conn):
    try:
        cur.execute("PRAGMA table_info(practice_availability)")
        availability_columns = [row[1] if not isinstance(row, sqlite3.Row) else row["name"] for row in cur.fetchall()]
        cur.execute("PRAGMA index_list(practice_availability)")
        availability_indexes = cur.fetchall()
        has_unique_availability_date_index = False
        for index_row in availability_indexes:
            if isinstance(index_row, sqlite3.Row):
                if index_row[2]:
                    cur.execute(f"PRAGMA index_info({index_row['name']})")
                    indexed_columns = [info[2] if not isinstance(info, sqlite3.Row) else info["name"] for info in cur.fetchall()]
                    if indexed_columns == ["date", "user_email"]:
                        has_unique_availability_date_index = True
                        break
            elif index_row[2]:
                cur.execute(f"PRAGMA index_info({index_row[1]})")
                indexed_columns = [info[2] if not isinstance(info, sqlite3.Row) else info["name"] for info in cur.fetchall()]
                if indexed_columns == ["date", "user_email"]:
                    has_unique_availability_date_index = True
                    break
        if has_unique_availability_date_index:
            cur.executescript("""
            CREATE TABLE IF NOT EXISTS practice_availability_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                practice_session_id INTEGER,
                date TEXT NOT NULL,
                user_email TEXT,
                user_full_name TEXT,
                status TEXT NOT NULL CHECK(status IN ('available', 'tentative', 'not_available')),
                option_choice TEXT,
                UNIQUE(practice_session_id, user_email)
            );
            INSERT OR IGNORE INTO practice_availability_new (id, date, user_email, user_full_name, status, option_choice, practice_session_id)
            SELECT id, date, user_email, user_full_name, status, option_choice, practice_session_id
            FROM practice_availability;
            DROP TABLE practice_availability;
            ALTER TABLE practice_availability_new RENAME TO practice_availability;
            """)
            conn.commit()
            cur.execute("PRAGMA table_info(practice_availability)")
            availability_columns = [row[1] if not isinstance(row, sqlite3.Row) else row["name"] for row in cur.fetchall()]

        cur.execute("PRAGMA table_info(practice_payments)")
        payment_columns = [row[1] if not isinstance(row, sqlite3.Row) else row["name"] for row in cur.fetchall()]
        cur.execute("PRAGMA index_list(practice_payments)")
        payment_indexes = cur.fetchall()
        has_unique_payment_date_index = False
        for index_row in payment_indexes:
            if isinstance(index_row, sqlite3.Row):
                if index_row[2]:
                    cur.execute(f"PRAGMA index_info({index_row['name']})")
                    indexed_columns = [info[2] if not isinstance(info, sqlite3.Row) else info["name"] for info in cur.fetchall()]
                    if indexed_columns == ["date", "user_email"]:
                        has_unique_payment_date_index = True
                        break
            elif index_row[2]:
                cur.execute(f"PRAGMA index_info({index_row[1]})")
                indexed_columns = [info[2] if not isinstance(info, sqlite3.Row) else info["name"] for info in cur.fetchall()]
                if indexed_columns == ["date", "user_email"]:
                    has_unique_payment_date_index = True
                    break
        if "practice_session_id" not in payment_columns or has_unique_payment_date_index:
            cur.executescript("""
            CREATE TABLE IF NOT EXISTS practice_payments_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                practice_session_id INTEGER,
                date TEXT NOT NULL,
                user_email TEXT NOT NULL,
                paid INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(practice_session_id, user_email)
            );
            INSERT OR REPLACE INTO practice_payments_new (id, practice_session_id, date, user_email, paid, created_at)
            SELECT id, practice_session_id, date, user_email, paid, created_at
            FROM practice_payments;
            DROP TABLE practice_payments;
            ALTER TABLE practice_payments_new RENAME TO practice_payments;
            """)
            conn.commit()

        cur.execute("PRAGMA table_info(practice_sessions)")
        practice_session_columns = [row[1] if not isinstance(row, sqlite3.Row) else row["name"] for row in cur.fetchall()]
        cur.execute("PRAGMA index_list(practice_sessions)")
        practice_session_indexes = cur.fetchall()
        has_unique_date_index = False
        for index_row in practice_session_indexes:
            if isinstance(index_row, sqlite3.Row):
                if index_row[2]:
                    cur.execute(f"PRAGMA index_info({index_row['name']})")
                    indexed_columns = [info[2] if not isinstance(info, sqlite3.Row) else info["name"] for info in cur.fetchall()]
                    if indexed_columns == ["date"]:
                        has_unique_date_index = True
                        break
            elif index_row[2]:
                cur.execute(f"PRAGMA index_info({index_row[1]})")
                indexed_columns = [info[2] if not isinstance(info, sqlite3.Row) else info["name"] for info in cur.fetchall()]
                if indexed_columns == ["date"]:
                    has_unique_date_index = True
                    break
        if "id" not in practice_session_columns or has_unique_date_index:
            cur.executescript("""
            CREATE TABLE IF NOT EXISTS practice_sessions_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                time TEXT,
                location TEXT,
                event_type TEXT DEFAULT 'practice',
                event_title TEXT,
                description TEXT,
                image_url TEXT,
                youtube_url TEXT,
                option_a_text TEXT,
                option_b_text TEXT,
                session_cost REAL,
                paid_by TEXT,
                payment_requested INTEGER DEFAULT 0,
                payment_requested_at TIMESTAMP,
                maximum_capacity INTEGER DEFAULT 100
            );
            INSERT INTO practice_sessions_new (
                date, time, location, event_type, event_title, description, image_url, youtube_url,
                option_a_text, option_b_text, session_cost, paid_by, payment_requested,
                payment_requested_at, maximum_capacity
            )
            SELECT
                date, time, location, event_type, event_title, description, image_url, youtube_url,
                option_a_text, option_b_text, session_cost, paid_by, payment_requested,
                payment_requested_at, maximum_capacity
            FROM practice_sessions
            ORDER BY date, COALESCE(time, ''), rowid;
            DROP TABLE practice_sessions;
            ALTER TABLE practice_sessions_new RENAME TO practice_sessions;
            CREATE INDEX IF NOT EXISTS idx_practice_sessions_date ON practice_sessions(date);
            """)
            conn.commit()
        else:
            cur.execute("CREATE INDEX IF NOT EXISTS idx_practice_sessions_date ON practice_sessions(date)")
            conn.commit()

        if "practice_session_id" not in availability_columns:
            cur.execute("ALTER TABLE practice_availability ADD COLUMN practice_session_id INTEGER")
        if "practice_session_id" not in payment_columns:
            cur.execute("ALTER TABLE practice_payments ADD COLUMN practice_session_id INTEGER")
        cur.execute("PRAGMA table_info(events)")
        event_columns = [row[1] if not isinstance(row, sqlite3.Row) else row["name"] for row in cur.fetchall()]
        if "practice_session_id" not in event_columns:
            cur.execute("ALTER TABLE events ADD COLUMN practice_session_id INTEGER")

        cur.execute("PRAGMA index_list(events)")
        event_indexes = cur.fetchall()
        has_unique_practice_session_date_index = False
        for index_row in event_indexes:
            if isinstance(index_row, sqlite3.Row):
                if index_row[2]:
                    cur.execute(f"PRAGMA index_info({index_row['name']})")
                    indexed_columns = [info[2] if not isinstance(info, sqlite3.Row) else info["name"] for info in cur.fetchall()]
                    if indexed_columns == ["practice_session_date"]:
                        has_unique_practice_session_date_index = True
                        break
            elif index_row[2]:
                cur.execute(f"PRAGMA index_info({index_row[1]})")
                indexed_columns = [info[2] if not isinstance(info, sqlite3.Row) else info["name"] for info in cur.fetchall()]
                if indexed_columns == ["practice_session_date"]:
                    has_unique_practice_session_date_index = True
                    break
        if has_unique_practice_session_date_index:
            cur.executescript("""
            CREATE TABLE IF NOT EXISTS events_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                date TEXT NOT NULL,
                time TEXT,
                location TEXT,
                description TEXT,
                image_url TEXT,
                youtube_url TEXT,
                practice_session_date TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                practice_session_id INTEGER
            );
            INSERT INTO events_new (id, name, date, time, location, description, image_url, youtube_url, practice_session_date, created_at, practice_session_id)
            SELECT id, name, date, time, location, description, image_url, youtube_url, practice_session_date, created_at, practice_session_id
            FROM events;
            DROP TABLE events;
            ALTER TABLE events_new RENAME TO events;
            """)
            conn.commit()

        cur.execute("SELECT id, date FROM practice_sessions ORDER BY id ASC")
        for row in cur.fetchall():
            row_dict = dict(row)
            cur.execute(
                "UPDATE practice_availability SET practice_session_id = ? WHERE practice_session_id IS NULL AND date = ?",
                (row_dict["id"], row_dict["date"]),
            )
            cur.execute(
                "UPDATE practice_payments SET practice_session_id = ? WHERE practice_session_id IS NULL AND date = ?",
                (row_dict["id"], row_dict["date"]),
            )
            cur.execute(
                "UPDATE events SET practice_session_id = ? WHERE practice_session_id IS NULL AND practice_session_date = ?",
                (row_dict["id"], row_dict["date"]),
            )

        cur.execute("CREATE INDEX IF NOT EXISTS idx_practice_availability_session_id ON practice_availability(practice_session_id)")
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_practice_availability_session_user_unique ON practice_availability(practice_session_id, user_email)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_practice_payments_session_id ON practice_payments(practice_session_id)")
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_practice_payments_session_user_unique ON practice_payments(practice_session_id, user_email)")
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_events_practice_session_id_unique ON events(practice_session_id) WHERE practice_session_id IS NOT NULL")
        conn.commit()
    except Exception as e:
        print(f"Warning: Could not migrate practice session IDs for SQLite: {e}")
        conn.rollback()

def ensure_practice_session_ids(cur, conn):
    if USE_POSTGRES:
        ensure_postgres_practice_session_ids(cur, conn)
    else:
        ensure_sqlite_practice_session_ids(cur, conn)

def get_practice_effective_time(time_value: Optional[str]) -> str:
    if not time_value:
        return "21:00"
    try:
        normalized_time = normalize_practice_time(time_value)
        return normalized_time or "21:00"
    except HTTPException:
        return "21:00"

def get_practice_datetime(date_value: str, time_value: Optional[str]) -> datetime:
    effective_time = get_practice_effective_time(time_value)
    return datetime.strptime(f"{date_value} {effective_time}", "%Y-%m-%d %H:%M")

def is_practice_datetime_in_past(date_value: str, time_value: Optional[str]) -> bool:
    try:
        practice_datetime = get_practice_datetime(date_value, time_value)
    except ValueError:
        return False
    return practice_datetime < datetime.now()

def get_practice_session_basic(cur, date_str: str) -> Optional[dict]:
    session_id = get_practice_session_id_by_date(cur, date_str)
    if session_id is None:
        return None
    return get_practice_session_basic_by_id(cur, session_id)

def get_practice_session_id_by_date(cur, date_str: str) -> Optional[int]:
    cur.execute(
        f"SELECT id FROM practice_sessions WHERE date = {PLACEHOLDER} ORDER BY id ASC LIMIT 1",
        (date_str,),
    )
    row = cur.fetchone()
    return dict(row).get("id") if row else None

def get_available_count_for_session(cur, date_str: str) -> int:
    session_id = get_practice_session_id_by_date(cur, date_str)
    if session_id is None:
        return 0
    return get_available_count_for_session_id(cur, session_id)

def get_practice_session_with_capacity(cur, date_str: str) -> Optional[dict]:
    session_id = get_practice_session_id_by_date(cur, date_str)
    if session_id is None:
        return None
    return get_practice_session_with_capacity_by_id(cur, session_id)

def get_practice_session_basic_by_id(cur, session_id: int) -> Optional[dict]:
    cur.execute(
        f"SELECT id, date, time, location, event_type, event_title, description, image_url, youtube_url, option_a_text, option_b_text, payment_requested, payment_requested_at, COALESCE(maximum_capacity, 100) as maximum_capacity FROM practice_sessions WHERE id = {PLACEHOLDER}",
        (session_id,),
    )
    row = cur.fetchone()
    if not row:
        return None
    session = dict(row)
    session["time"] = get_practice_effective_time(session.get("time")) if session.get("time") else "21:00"
    session["event_type"] = normalize_event_type(session.get("event_type")) if session.get("event_type") else "practice"
    session["event_title"] = normalize_event_title(session.get("event_title"), session["event_type"])
    session["option_a_text"] = (session.get("option_a_text") or "").strip() or None
    session["option_b_text"] = (session.get("option_b_text") or "").strip() or None
    return session

def get_practice_session_with_capacity_by_id(cur, session_id: int) -> Optional[dict]:
    cur.execute(
        f"""
        SELECT 
            ps.id,
            ps.date,
            ps.time,
            ps.location,
            ps.event_type,
            ps.event_title,
            ps.description,
            ps.image_url,
            ps.youtube_url,
            ps.option_a_text,
            ps.option_b_text,
            ps.session_cost,
            ps.paid_by,
            ps.payment_requested,
            ps.payment_requested_at,
            COALESCE(ps.maximum_capacity, 100) as maximum_capacity,
            u.full_name as paid_by_name,
            u.bank_name as paid_by_bank_name,
            u.sort_code as paid_by_sort_code,
            u.account_number as paid_by_account_number
        FROM practice_sessions ps
        LEFT JOIN users u ON ps.paid_by = u.email AND (u.is_deleted = FALSE OR u.is_deleted IS NULL)
        WHERE ps.id = {PLACEHOLDER}
        """,
        (session_id,),
    )
    row = cur.fetchone()
    if not row:
        return None
    session = dict(row)
    if session.get("time"):
        try:
            session["time"] = normalize_practice_time(session.get("time"))
        except HTTPException:
            session["time"] = "21:00"
    else:
        session["time"] = None
    session["event_type"] = normalize_event_type(session.get("event_type")) if session.get("event_type") else "practice"
    session["event_title"] = normalize_event_title(session.get("event_title"), session["event_type"])
    session["option_a_text"] = (session.get("option_a_text") or "").strip() or None
    session["option_b_text"] = (session.get("option_b_text") or "").strip() or None
    available_count = get_available_count_for_session_id(cur, session_id)
    maximum_capacity = normalize_maximum_capacity(session.get("maximum_capacity"))
    session["maximum_capacity"] = maximum_capacity
    session["available_count"] = available_count
    session["remaining_slots"] = max(maximum_capacity - available_count, 0)
    session["capacity_reached"] = available_count >= maximum_capacity
    return session

def get_available_count_for_session_id(cur, session_id: int) -> int:
    cur.execute(
        f"SELECT COUNT(*) as count FROM practice_availability WHERE practice_session_id = {PLACEHOLDER} AND status = {PLACEHOLDER}",
        (session_id, "available"),
    )
    row = cur.fetchone()
    if not row:
        return 0
    return dict(row).get("count", 0) if USE_POSTGRES else row[0]

def get_practice_session_date_by_id(cur, session_id: int) -> Optional[str]:
    cur.execute(
        f"SELECT date FROM practice_sessions WHERE id = {PLACEHOLDER}",
        (session_id,),
    )
    row = cur.fetchone()
    if not row:
        return None
    return dict(row).get("date")

def normalize_option_pair(option_a_text: Optional[str], option_b_text: Optional[str]) -> tuple[Optional[str], Optional[str]]:
    normalized_a = (option_a_text or "").strip() or None
    normalized_b = (option_b_text or "").strip() or None
    if bool(normalized_a) != bool(normalized_b):
        raise HTTPException(status_code=400, detail="Both Option A and Option B must be filled to enable member selection")
    return normalized_a, normalized_b

def sync_match_session_to_events(cur, session: dict):
    practice_date = session["date"]
    practice_session_id = session.get("id")
    event_type = normalize_event_type(session.get("event_type")) if session.get("event_type") else "practice"
    if event_type != "match":
        if practice_session_id is not None:
            cur.execute(f"DELETE FROM event_likes WHERE event_id IN (SELECT id FROM events WHERE practice_session_id = {PLACEHOLDER})", (practice_session_id,))
            cur.execute(f"DELETE FROM event_comments WHERE event_id IN (SELECT id FROM events WHERE practice_session_id = {PLACEHOLDER})", (practice_session_id,))
            cur.execute(f"DELETE FROM events WHERE practice_session_id = {PLACEHOLDER}", (practice_session_id,))
        else:
            cur.execute(f"DELETE FROM event_likes WHERE event_id IN (SELECT id FROM events WHERE practice_session_date = {PLACEHOLDER})", (practice_date,))
            cur.execute(f"DELETE FROM event_comments WHERE event_id IN (SELECT id FROM events WHERE practice_session_date = {PLACEHOLDER})", (practice_date,))
            cur.execute(f"DELETE FROM events WHERE practice_session_date = {PLACEHOLDER}", (practice_date,))
        return

    match_name = normalize_event_title(session.get("event_title"), "match")
    if practice_session_id is not None:
        cur.execute(
            f"SELECT id FROM events WHERE practice_session_date = {PLACEHOLDER} AND (practice_session_id IS NULL OR practice_session_id <> {PLACEHOLDER})",
            (practice_date, practice_session_id),
        )
        legacy_event_ids = [dict(row)["id"] for row in cur.fetchall()]
        if legacy_event_ids:
            for legacy_event_id in legacy_event_ids:
                cur.execute(f"DELETE FROM event_likes WHERE event_id = {PLACEHOLDER}", (legacy_event_id,))
                cur.execute(f"DELETE FROM event_comments WHERE event_id = {PLACEHOLDER}", (legacy_event_id,))
                cur.execute(f"DELETE FROM event_media WHERE event_id = {PLACEHOLDER}", (legacy_event_id,))
                cur.execute(f"DELETE FROM events WHERE id = {PLACEHOLDER}", (legacy_event_id,))
    if USE_POSTGRES:
        cur.execute(
            f"SELECT id FROM events WHERE practice_session_id = {PLACEHOLDER}",
            (practice_session_id,),
        )
        existing_event = cur.fetchone()
        if existing_event:
            cur.execute(
                f"UPDATE events SET name = {PLACEHOLDER}, date = {PLACEHOLDER}, time = {PLACEHOLDER}, location = {PLACEHOLDER}, description = {PLACEHOLDER}, image_url = {PLACEHOLDER}, youtube_url = {PLACEHOLDER}, practice_session_date = {PLACEHOLDER}, practice_session_id = {PLACEHOLDER} WHERE practice_session_id = {PLACEHOLDER}",
                (match_name, practice_date, session.get("time"), session.get("location"), session.get("description"), session.get("image_url"), session.get("youtube_url"), practice_date, practice_session_id, practice_session_id),
            )
        else:
            cur.execute(
                f"INSERT INTO events (name, date, time, location, description, image_url, youtube_url, practice_session_date, practice_session_id) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
                (match_name, practice_date, session.get("time"), session.get("location"), session.get("description"), session.get("image_url"), session.get("youtube_url"), practice_date, practice_session_id),
            )
    else:
        cur.execute(
            f"SELECT id FROM events WHERE practice_session_id = {PLACEHOLDER}",
            (practice_session_id,),
        )
        existing_event = cur.fetchone()
        if existing_event:
            cur.execute(
                f"UPDATE events SET name = {PLACEHOLDER}, date = {PLACEHOLDER}, time = {PLACEHOLDER}, location = {PLACEHOLDER}, description = {PLACEHOLDER}, image_url = {PLACEHOLDER}, youtube_url = {PLACEHOLDER}, practice_session_date = {PLACEHOLDER}, practice_session_id = {PLACEHOLDER} WHERE practice_session_id = {PLACEHOLDER}",
                (match_name, practice_date, session.get("time"), session.get("location"), session.get("description"), session.get("image_url"), session.get("youtube_url"), practice_date, practice_session_id, practice_session_id),
            )
        else:
            cur.execute(
                f"INSERT INTO events (name, date, time, location, description, image_url, youtube_url, practice_session_date, practice_session_id) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
                (match_name, practice_date, session.get("time"), session.get("location"), session.get("description"), session.get("image_url"), session.get("youtube_url"), practice_date, practice_session_id),
            )

def set_practice_availability_for_session_id(cur, conn, session_id: int, user_email: str, user_full_name: str, status_value: str, option_choice_value: Optional[str]):
    session = get_practice_session_with_capacity_by_id(cur, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Practice session not found")

    if is_practice_datetime_in_past(session["date"], session.get("time")):
        raise HTTPException(status_code=403, detail="Cannot modify availability after the practice session date and time has passed.")

    if session["payment_requested"]:
        raise HTTPException(status_code=403, detail="Cannot modify availability after payment request has been enabled.")

    cur.execute(
        f"SELECT status FROM practice_availability WHERE practice_session_id = {PLACEHOLDER} AND user_email = {PLACEHOLDER}",
        (session_id, user_email),
    )
    existing_row = cur.fetchone()
    previous_status = dict(existing_row).get("status") if existing_row else None

    option_choice = (option_choice_value or "").strip().upper() or None
    option_a_text, option_b_text = normalize_option_pair(session.get("option_a_text"), session.get("option_b_text"))
    if option_choice and status_value != "available":
        raise HTTPException(status_code=400, detail="Option selection is only allowed when availability is Available")
    if option_choice and option_choice not in {"A", "B"}:
        raise HTTPException(status_code=400, detail="Invalid option choice")
    if option_choice and not (option_a_text and option_b_text):
        raise HTTPException(status_code=400, detail="This event does not have member options enabled")

    is_new_available_vote = status_value == "available" and previous_status != "available"
    if is_new_available_vote and session["capacity_reached"]:
        raise HTTPException(status_code=403, detail="Maximum capacity has been reached for this session. No more Available selections are allowed right now.")

    if status_value == "none":
        cur.execute(
            f"DELETE FROM practice_availability WHERE practice_session_id = {PLACEHOLDER} AND user_email = {PLACEHOLDER}",
            (session_id, user_email),
        )
        conn.commit()
        return {"message": "Availability removed"}

    if USE_POSTGRES:
        cur.execute(
            f"INSERT INTO practice_availability (practice_session_id, date, user_email, user_full_name, status, option_choice) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}) ON CONFLICT (practice_session_id, user_email) DO UPDATE SET date = EXCLUDED.date, status = EXCLUDED.status, user_full_name = EXCLUDED.user_full_name, option_choice = EXCLUDED.option_choice",
            (session_id, session["date"], user_email, user_full_name, status_value, option_choice if status_value == "available" else None),
        )
    else:
        cur.execute(
            f"INSERT OR REPLACE INTO practice_availability (id, practice_session_id, date, user_email, user_full_name, status, option_choice) VALUES ((SELECT id FROM practice_availability WHERE practice_session_id = {PLACEHOLDER} AND user_email = {PLACEHOLDER}), {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
            (session_id, user_email, session_id, session["date"], user_email, user_full_name, status_value, option_choice if status_value == "available" else None),
        )
    conn.commit()

    updated_session = get_practice_session_with_capacity_by_id(cur, session_id)
    if is_new_available_vote and updated_session and updated_session["capacity_reached"]:
        deliver_notification(
            "session_capacity_reached",
            {
                "session_id": updated_session["id"],
                "date": updated_session["date"],
                "time": updated_session.get("time"),
                "location": updated_session.get("location"),
                "event_type": updated_session.get("event_type"),
                "event_title": updated_session.get("event_title"),
                "maximum_capacity": updated_session["maximum_capacity"],
                "available_count": updated_session["available_count"],
                "remaining_slots": updated_session["remaining_slots"],
            },
            related_date=session["date"],
        )
    return {"message": "Availability set"}

def admin_set_practice_availability_for_session_id(cur, conn, session_id: int, user_email: str, user_full_name: str, status_value: str, option_choice_value: Optional[str]):
    session = get_practice_session_with_capacity_by_id(cur, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Practice session not found")

    option_choice = (option_choice_value or "").strip().upper() or None
    option_a_text, option_b_text = normalize_option_pair(session.get("option_a_text"), session.get("option_b_text"))
    if option_choice and status_value != "available":
        raise HTTPException(status_code=400, detail="Option selection is only allowed when availability is Available")
    if option_choice and option_choice not in {"A", "B"}:
        raise HTTPException(status_code=400, detail="Invalid option choice")
    if option_choice and not (option_a_text and option_b_text):
        raise HTTPException(status_code=400, detail="This event does not have member options enabled")

    if status_value == "delete":
        cur.execute(
            f"DELETE FROM practice_availability WHERE practice_session_id = {PLACEHOLDER} AND user_email = {PLACEHOLDER}",
            (session_id, user_email),
        )
        conn.commit()
        return {"message": "Availability removed"}

    if USE_POSTGRES:
        cur.execute(
            f"INSERT INTO practice_availability (practice_session_id, date, user_email, user_full_name, status, option_choice) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}) ON CONFLICT (practice_session_id, user_email) DO UPDATE SET date = EXCLUDED.date, status = EXCLUDED.status, user_full_name = EXCLUDED.user_full_name, option_choice = EXCLUDED.option_choice",
            (session_id, session["date"], user_email, user_full_name, status_value, option_choice if status_value == "available" else None),
        )
    else:
        cur.execute(
            f"INSERT OR REPLACE INTO practice_availability (id, practice_session_id, date, user_email, user_full_name, status, option_choice) VALUES ((SELECT id FROM practice_availability WHERE practice_session_id = {PLACEHOLDER} AND user_email = {PLACEHOLDER}), {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
            (session_id, user_email, session_id, session["date"], user_email, user_full_name, status_value, option_choice if status_value == "available" else None),
        )
    conn.commit()
    return {"message": "Availability set"}

def get_practice_availability_summary_by_session(cur, session: dict):
    cur.execute(
        f"SELECT user_email, user_full_name, status, option_choice FROM practice_availability WHERE practice_session_id = {PLACEHOLDER}",
        (session["id"],),
    )
    rows = [dict(r) for r in cur.fetchall()]

    available = []
    tentative = []
    not_available = []
    option_a = []
    option_b = []
    option_a_text = session.get("option_a_text")
    option_b_text = session.get("option_b_text")

    for r in rows:
        name = r.get("user_full_name")
        if not name:
            name = "[OldData]"

        if r["status"] == "available":
            available.append(name)
            if option_a_text and option_b_text and r.get("option_choice") == "A":
                option_a.append(name)
            elif option_a_text and option_b_text and r.get("option_choice") == "B":
                option_b.append(name)
        elif r["status"] == "tentative":
            tentative.append(name)
        elif r["status"] == "not_available":
            not_available.append(name)

    maximum_capacity = session["maximum_capacity"] if session else 100
    available_count = len(available)
    remaining_slots = max(maximum_capacity - available_count, 0)

    return {
        "available": available,
        "tentative": tentative,
        "not_available": not_available,
        "option_a": option_a,
        "option_b": option_b,
        "option_a_text": option_a_text,
        "option_b_text": option_b_text,
        "user_emails": {r["user_full_name"] or r["user_email"]: r["user_email"] for r in rows},
        "maximum_capacity": maximum_capacity,
        "available_count": available_count,
        "remaining_slots": remaining_slots,
        "capacity_reached": available_count >= maximum_capacity,
    }

def notify_practice_slots_available():
    now = datetime.now()
    window_end = now + timedelta(hours=72)
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT date FROM practice_sessions ORDER BY date ASC")
        session = None
        for row in cur.fetchall():
            row_dict = dict(row)
            candidate_session = get_practice_session_with_capacity(cur, row_dict["date"])
            if not candidate_session:
                continue
            candidate_datetime = get_practice_datetime(candidate_session["date"], candidate_session.get("time"))
            if candidate_datetime < now or candidate_datetime > window_end:
                continue
            if candidate_session["remaining_slots"] <= 0:
                continue
            session = candidate_session
            break
        if not session:
            return

        deliver_notification(
            "practice_slot_available",
            {
                "session_id": session["id"],
                "date": session["date"],
                "time": session.get("time"),
                "location": session.get("location"),
                "event_type": session.get("event_type"),
                "event_title": session.get("event_title"),
                "maximum_capacity": session["maximum_capacity"],
                "available_count": session["available_count"],
                "remaining_slots": session["remaining_slots"],
            },
            related_date=session["date"],
        )

def get_pending_payment_count_for_session(cur, date_str: str) -> int:
    session_id = get_practice_session_id_by_date(cur, date_str)
    if session_id is None:
        return 0
    cur.execute(
        f"""
        SELECT COUNT(*) as count
        FROM practice_availability pa
        LEFT JOIN practice_payments pp
            ON pa.practice_session_id = pp.practice_session_id AND pa.user_email = pp.user_email
        WHERE pa.practice_session_id = {PLACEHOLDER}
          AND pa.status = {PLACEHOLDER}
          AND (pp.paid IS NULL OR pp.paid = {PLACEHOLDER})
        """,
        (session_id, "available", False if USE_POSTGRES else 0),
    )
    row = cur.fetchone()
    if not row:
        return 0
    row_dict = dict(row)
    return int(row_dict.get("count", 0))

def notify_pending_payment_reminders():
    now = datetime.now()
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT id, date, time, location, event_type, event_title, payment_requested_at
            FROM practice_sessions
            WHERE payment_requested = {PLACEHOLDER}
              AND payment_requested_at IS NOT NULL
            ORDER BY payment_requested_at DESC, date DESC
            """,
            (True if USE_POSTGRES else 1,),
        )
        sessions = [dict(row) for row in cur.fetchall()]
        if not sessions:
            return

        latest_request = sessions[0]
        payment_requested_at = latest_request.get("payment_requested_at")
        if not payment_requested_at:
            return
        if isinstance(payment_requested_at, str):
            try:
                payment_requested_at = datetime.fromisoformat(payment_requested_at.replace("Z", "+00:00")).replace(tzinfo=None)
            except ValueError:
                return

        if now < payment_requested_at + timedelta(hours=72):
            return

        has_pending_payments = False
        reminder_session = None
        for candidate_session in sessions:
            if not is_practice_datetime_in_past(candidate_session["date"], candidate_session.get("time")):
                continue
            pending_count = get_pending_payment_count_for_session(cur, candidate_session["id"])
            if pending_count > 0:
                has_pending_payments = True
                reminder_session = candidate_session
                break

        if not has_pending_payments or not reminder_session:
            return

        deliver_notification(
            "pending_payment_reminder",
            {
                "session_id": reminder_session["id"],
                "date": reminder_session["date"],
                "time": reminder_session.get("time"),
                "location": reminder_session.get("location"),
                "event_type": reminder_session.get("event_type"),
                "event_title": reminder_session.get("event_title"),
                "payments_link": "https://glasgow-bengali-fc.vercel.app/user-actions/payments",
            },
            related_date=reminder_session["date"],
        )

def serialize_notification_setting(row: dict) -> dict:
    return {
        "notif_type": row["notif_type"],
        "display_name": row["display_name"],
        "description": row.get("description"),
        "app_enabled": bool(row["app_enabled"]),
        "email_enabled": bool(row["email_enabled"]),
        "whatsapp_enabled": bool(row["whatsapp_enabled"]),
        "target_audience": row["target_audience"],
        "app_template": row["app_template"],
        "email_subject": row["email_subject"],
        "email_template": row["email_template"],
        "whatsapp_template": row["whatsapp_template"],
        "updated_at": row.get("updated_at").isoformat() if hasattr(row.get("updated_at"), "isoformat") else row.get("updated_at"),
    }

def serialize_expense(row: dict) -> dict:
    return {
        "id": row["id"],
        "title": row["title"],
        "amount": float(row["amount"]),
        "paid_by": row.get("paid_by"),
        "expense_date": row["expense_date"],
        "category": row.get("category"),
        "payment_method": row.get("payment_method"),
        "description": row.get("description"),
        "paid_by_name": row.get("paid_by_name"),
        "source": row.get("source") or "expense",
        "is_booking_expense": bool(row.get("is_booking_expense", False)),
        "practice_session_date": row.get("practice_session_date"),
        "linked_practice_time": row.get("linked_practice_time"),
        "linked_practice_location": row.get("linked_practice_location"),
        "can_edit": bool(row.get("can_edit", True)),
        "can_delete": bool(row.get("can_delete", True)),
        "created_at": row.get("created_at").isoformat() if hasattr(row.get("created_at"), "isoformat") else row.get("created_at"),
        "updated_at": row.get("updated_at").isoformat() if hasattr(row.get("updated_at"), "isoformat") else row.get("updated_at"),
    }

# --- Pydantic models ---
class UserCreate(BaseModel):
    email: EmailStr
    full_name: str
    password: str

class UserOut(BaseModel):
    id: int
    email: str
    full_name: str
    user_type: str = "member"
    created_at: Optional[str] = None
    last_login: Optional[str] = None
    birthday: Optional[str] = None
    bank_name: Optional[str] = None
    sort_code: Optional[str] = None
    account_number: Optional[str] = None
    theme_preference: str = "nordic_neutral"

class Token(BaseModel):
    access_token: str
    token_type: str

class ForgotPasswordRequest(BaseModel):
    email: EmailStr

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

class WhatsAppMessageRequest(BaseModel):
    message: str

class WhatsAppGroupLookupRequest(BaseModel):
    group_name: str

class NotificationSettingOut(BaseModel):
    notif_type: str
    display_name: str
    description: Optional[str] = None
    app_enabled: bool
    email_enabled: bool
    whatsapp_enabled: bool
    target_audience: str
    app_template: str
    email_subject: str
    email_template: str
    whatsapp_template: str
    updated_at: Optional[str] = None

class NotificationSettingUpdate(BaseModel):
    display_name: str
    description: Optional[str] = None
    app_enabled: bool
    email_enabled: bool
    whatsapp_enabled: bool
    target_audience: str
    app_template: str
    email_subject: str
    email_template: str
    whatsapp_template: str

class EventOut(BaseModel):
    id: int
    name: str
    date: str
    time: Optional[str]
    location: Optional[str]
    description: Optional[str]
    image_url: Optional[str] = None
    youtube_url: Optional[str] = None
    likes: Optional[List[dict]] = []
    comments: Optional[List[dict]] = []

class EventCreate(BaseModel):
    name: str
    date: str
    time: Optional[str]
    location: Optional[str]
    description: Optional[str]
    image_url: Optional[str] = None
    youtube_url: Optional[str] = None

class EventComment(BaseModel):
    id: int
    user_email: str
    comment: str
    created_at: str

class ForumCommentOut(BaseModel):
    id: int
    user_full_name: str
    comment: str
    created_at: str

class ForumPostOut(BaseModel):
    id: int
    user_full_name: str
    user_email: str
    content: str
    created_at: str
    likes_count: int
    likes: List[dict] = []
    comments: List[ForumCommentOut]

class ForumPostCreate(BaseModel):
    content: str

class ForumPostUpdate(BaseModel):
    content: str

class ForumComment(BaseModel):
    comment: str

class PracticeAvailability(BaseModel):
    date: str
    status: str
    option_choice: Optional[str] = None

class AdminPracticeAvailability(BaseModel):
    date: str
    user_email: str
    status: str  # 'available', 'tentative', 'not_available', or 'delete' to remove
    option_choice: Optional[str] = None

class PracticeSessionCreate(BaseModel):
    date: str
    time: Optional[str] = None
    location: Optional[str] = None
    event_type: str = "practice"
    event_title: Optional[str] = None
    description: Optional[str] = None
    image_url: Optional[str] = None
    youtube_url: Optional[str] = None
    option_a_text: Optional[str] = None
    option_b_text: Optional[str] = None
    session_cost: Optional[float] = None
    paid_by: Optional[str] = None
    maximum_capacity: int = 100

class PracticeSessionOut(BaseModel):
    id: int
    date: str
    time: Optional[str] = None
    location: Optional[str] = None
    event_type: str = "practice"
    event_title: Optional[str] = None
    description: Optional[str] = None
    image_url: Optional[str] = None
    youtube_url: Optional[str] = None
    option_a_text: Optional[str] = None
    option_b_text: Optional[str] = None
    session_cost: Optional[float] = None
    paid_by: Optional[str] = None
    maximum_capacity: int = 100
    available_count: Optional[int] = 0
    remaining_slots: Optional[int] = 100
    capacity_reached: Optional[bool] = False
    paid_by_name: Optional[str] = None
    paid_by_bank_name: Optional[str] = None
    paid_by_sort_code: Optional[str] = None
    paid_by_account_number: Optional[str] = None
    payment_requested: Optional[bool] = False

class ExpenseCreate(BaseModel):
    title: str
    amount: float
    paid_by: Optional[str] = None
    expense_date: str
    category: Optional[str] = None
    payment_method: Optional[str] = None
    description: Optional[str] = None

class ExpenseOut(BaseModel):
    id: int
    title: str
    amount: float
    paid_by: Optional[str] = None
    expense_date: str
    category: Optional[str] = None
    payment_method: Optional[str] = None
    description: Optional[str] = None
    paid_by_name: Optional[str] = None
    source: Optional[str] = "expense"
    is_booking_expense: Optional[bool] = False
    practice_session_date: Optional[str] = None
    linked_practice_time: Optional[str] = None
    linked_practice_location: Optional[str] = None
    can_edit: Optional[bool] = True
    can_delete: Optional[bool] = True
    created_at: Optional[str] = None
    updated_at: Optional[str] = None

# --- Helper Functions ---
def is_admin(current_user: dict) -> bool:
    """Check if user is admin based on user_type or super admin email"""
    # Fetch user_type from database
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT user_type, is_deleted FROM users WHERE email = {PLACEHOLDER}", (current_user["email"],))
        row = cur.fetchone()
        if row:
            row_dict = dict(row)
            # Check if user is deleted
            if row_dict.get('is_deleted'):
                return False
            user_type = row_dict.get("user_type") or "member"
        else:
            user_type = "member"
    
    # Check if user_type is 'admin' OR email is 'super@admin.com'
    return user_type == "admin" or current_user.get("email") == "super@admin.com"

def normalize_event_type(value: Optional[str]) -> str:
    normalized = (value or "practice").strip().lower()
    if normalized not in EVENT_TYPE_OPTIONS:
        raise HTTPException(status_code=400, detail="Event type must be one of: Practice, Match, Social, Others")
    return normalized

def sanitize_forum_comment_text(comment: str) -> str:
    return (comment or "").strip()

def sanitize_forum_post_html(content: str) -> str:
    allowed_tags = {"a", "br", "div", "iframe", "img"}
    allowed_attributes = {
        "a": {"href", "target", "rel", "style"},
        "div": {"style"},
        "iframe": {"width", "height", "style", "src", "frameborder", "allow", "allowfullscreen"},
        "img": {"src", "style"},
    }
    allowed_css_properties = {
        "margin",
        "margin-top",
        "max-width",
        "border-radius",
        "width",
        "height",
        "display",
        "text-decoration",
        "color",
    }

    def sanitize_style(style_value: str) -> str:
        sanitized_parts = []
        for part in (style_value or "").split(";"):
            if ":" not in part:
                continue
            name, value = part.split(":", 1)
            normalized_name = name.strip().lower()
            normalized_value = value.strip()
            if not normalized_name or not normalized_value:
                continue
            if normalized_name not in allowed_css_properties:
                continue
            lower_value = normalized_value.lower()
            if "expression" in lower_value or "javascript:" in lower_value or "url(" in lower_value:
                continue
            sanitized_parts.append(f"{normalized_name}: {normalized_value}")
        return "; ".join(sanitized_parts)

    def sanitize_url(raw_url: Optional[str], *, image_only: bool = False, iframe_only: bool = False) -> Optional[str]:
        candidate = (raw_url or "").strip()
        if not candidate:
            return None
        parsed = urlparse(candidate)
        if image_only:
            if candidate.startswith("/uploads/"):
                return candidate
            if parsed.scheme == "https" and parsed.netloc:
                return candidate
            if parsed.scheme == "http" and parsed.netloc in {"localhost:8000", "127.0.0.1:8000"}:
                return candidate
            return None
        if parsed.scheme != "https":
            return None
        if iframe_only:
            if parsed.netloc not in {"www.youtube.com", "youtube.com"}:
                return None
            if not parsed.path.startswith("/embed/"):
                return None
            return candidate
        return candidate

    class ForumContentSanitizer(HTMLParser):
        def __init__(self):
            super().__init__(convert_charrefs=True)
            self.parts = []

        def handle_starttag(self, tag, attrs):
            normalized_tag = tag.lower()
            if normalized_tag not in allowed_tags:
                return

            sanitized_attrs = []
            attr_map = dict(attrs)
            for attr_name, attr_value in attrs:
                normalized_attr = attr_name.lower()
                if normalized_attr not in allowed_attributes.get(normalized_tag, set()):
                    continue
                if normalized_attr == "style":
                    sanitized_style = sanitize_style(attr_value or "")
                    if sanitized_style:
                        sanitized_attrs.append((normalized_attr, sanitized_style))
                    continue
                if normalized_tag == "a" and normalized_attr == "href":
                    sanitized_url = sanitize_url(attr_value)
                    if sanitized_url:
                        sanitized_attrs.append((normalized_attr, sanitized_url))
                    continue
                if normalized_tag == "img" and normalized_attr == "src":
                    sanitized_url = sanitize_url(attr_value, image_only=True)
                    if sanitized_url:
                        sanitized_attrs.append((normalized_attr, sanitized_url))
                    continue
                if normalized_tag == "iframe" and normalized_attr == "src":
                    sanitized_url = sanitize_url(attr_value, iframe_only=True)
                    if sanitized_url:
                        sanitized_attrs.append((normalized_attr, sanitized_url))
                    continue
                if normalized_tag == "iframe" and normalized_attr == "allow":
                    lower_allow = (attr_value or "").lower()
                    if "javascript" in lower_allow:
                        continue
                if normalized_tag == "a" and normalized_attr == "target":
                    if attr_value != "_blank":
                        continue
                if normalized_tag == "a" and normalized_attr == "rel":
                    sanitized_attrs.append((normalized_attr, "noopener noreferrer"))
                    continue
                sanitized_attrs.append((normalized_attr, attr_value or ""))

            if normalized_tag == "a" and "href" not in dict(sanitized_attrs):
                return
            if normalized_tag == "iframe" and "src" not in dict(sanitized_attrs):
                return
            if normalized_tag == "img" and "src" not in dict(sanitized_attrs):
                return
            if normalized_tag == "a":
                attr_keys = {name for name, _ in sanitized_attrs}
                if "target" not in attr_keys:
                    sanitized_attrs.append(("target", "_blank"))
                if "rel" not in attr_keys:
                    sanitized_attrs.append(("rel", "noopener noreferrer"))

            rendered_attrs = "".join(
                f' {name}="{html.escape(value, quote=True)}"'
                for name, value in sanitized_attrs
            )
            self.parts.append(f"<{normalized_tag}{rendered_attrs}>")

        def handle_startendtag(self, tag, attrs):
            normalized_tag = tag.lower()
            if normalized_tag != "br":
                self.handle_starttag(tag, attrs)
                return
            self.parts.append("<br>")

        def handle_endtag(self, tag):
            normalized_tag = tag.lower()
            if normalized_tag in allowed_tags and normalized_tag != "img" and normalized_tag != "br":
                self.parts.append(f"</{normalized_tag}>")

        def handle_data(self, data):
            self.parts.append(html.escape(data))

        def handle_entityref(self, name):
            self.parts.append(f"&{name};")

        def handle_charref(self, name):
            self.parts.append(f"&#{name};")

    sanitizer = ForumContentSanitizer()
    sanitizer.feed(content or "")
    sanitizer.close()
    sanitized = "".join(sanitizer.parts)
    return sanitized.strip()

def default_event_title_for_type(event_type: str) -> str:
    return {
        "practice": "Session",
        "match": "Match",
        "social": "Social Event",
        "others": "Other Event",
    }.get(event_type, "Event")

def default_event_type_label(event_type: str) -> str:
    return {
        "practice": "Practice",
        "match": "Match",
        "social": "Social",
        "others": "Other",
    }.get(event_type, "Event")

def normalize_event_title(event_title: Optional[str], event_type: str) -> str:
    cleaned = (event_title or "").strip()
    if len(cleaned) > 30:
        raise HTTPException(status_code=400, detail="Event title must be 30 characters or less")
    return cleaned if cleaned else default_event_title_for_type(event_type)

# --- FastAPI app ---
app = FastAPI(title="Glasgow Bengali FC API", version="1.0")

# CORS configuration - MUST be before startup event
FRONTEND_ORIGIN = os.environ.get("FRONTEND_ORIGIN")
allowed_origins = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "http://localhost:5174",
    "http://127.0.0.1:5174",
    "https://glasgow-bengali-fc.vercel.app",  # Production frontend
    "https://gbfc.onrender.com",  # Backend domain (for potential same-origin requests)
    "https://www.glasgow-bengali-fc.vercel.app",  # WWW subdomain
]
if FRONTEND_ORIGIN and FRONTEND_ORIGIN not in allowed_origins:
    allowed_origins.append(FRONTEND_ORIGIN)

print(f"Configuring CORS with allowed origins: {allowed_origins}")

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize database on startup
@app.on_event("startup")
async def startup_event():
    init_db()
    seed_notification_settings()
    print("Database initialized successfully")
    if whatsapp_is_configured() and not whatsapp_scheduler.running:
        whatsapp_scheduler.add_job(keep_whatsapp_instance_alive, "interval", minutes=30, id="whatsapp_keepalive", replace_existing=True)
        whatsapp_scheduler.add_job(notify_practice_slots_available, "cron", hour=9, minute=0, id="practice_slot_available_daily", replace_existing=True)
        whatsapp_scheduler.add_job(notify_pending_payment_reminders, "cron", hour=20, minute=0, id="pending_payment_reminder_daily", replace_existing=True)
        whatsapp_scheduler.start()
        print("WhatsApp keep-alive scheduler started")
    elif not whatsapp_scheduler.running:
        whatsapp_scheduler.add_job(notify_practice_slots_available, "cron", hour=9, minute=0, id="practice_slot_available_daily", replace_existing=True)
        whatsapp_scheduler.add_job(notify_pending_payment_reminders, "cron", hour=20, minute=0, id="pending_payment_reminder_daily", replace_existing=True)
        whatsapp_scheduler.start()
        print("Notification scheduler started")

@app.on_event("shutdown")
async def shutdown_event():
    if whatsapp_scheduler.running:
        whatsapp_scheduler.shutdown(wait=False)

# Simple in-memory session token store (use JWT in production)
SESSIONS = {}

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

def persist_session_token(token: str, user: dict):
    expires_at = datetime.utcnow() + SESSION_DURATION
    SESSIONS[token] = {"email": user["email"], "full_name": user["full_name"], "id": user["id"], "expires_at": expires_at}
    with get_connection() as conn:
        cur = conn.cursor()
        if USE_POSTGRES:
            cur.execute(
                f"INSERT INTO auth_sessions (token, user_email, user_full_name, user_id, expires_at) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}) ON CONFLICT (token) DO UPDATE SET user_email = EXCLUDED.user_email, user_full_name = EXCLUDED.user_full_name, user_id = EXCLUDED.user_id, expires_at = EXCLUDED.expires_at",
                (token, user["email"], user["full_name"], user["id"], expires_at),
            )
        else:
            cur.execute(
                f"INSERT OR REPLACE INTO auth_sessions (token, user_email, user_full_name, user_id, expires_at) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
                (token, user["email"], user["full_name"], user["id"], expires_at),
            )
        conn.commit()

def delete_session_token(token: str):
    SESSIONS.pop(token, None)
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"DELETE FROM auth_sessions WHERE token = {PLACEHOLDER}", (token,))
        conn.commit()

def get_current_user(token: str = Depends(oauth2_scheme)):
    if token in SESSIONS:
        cached_session = SESSIONS[token]
        if cached_session.get("expires_at") and cached_session["expires_at"] < datetime.utcnow():
            delete_session_token(token)
            raise HTTPException(status_code=401, detail="Session expired")
        return {
            "email": cached_session["email"],
            "full_name": cached_session["full_name"],
            "id": cached_session["id"],
        }

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"DELETE FROM auth_sessions WHERE expires_at < {PLACEHOLDER}", (datetime.utcnow(),))
        cur.execute(
            f"SELECT token, user_email, user_full_name, user_id, expires_at FROM auth_sessions WHERE token = {PLACEHOLDER}",
            (token,),
        )
        row = cur.fetchone()
        if not row:
            conn.commit()
            raise HTTPException(status_code=401, detail="Invalid token")
        row_dict = dict(row)
        expires_at = row_dict.get("expires_at")
        if expires_at:
            if isinstance(expires_at, str):
                try:
                    expires_at = datetime.fromisoformat(expires_at.replace("Z", "+00:00")).replace(tzinfo=None)
                except ValueError:
                    try:
                        expires_at = datetime.strptime(expires_at, "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        expires_at = None
        if expires_at and expires_at < datetime.utcnow():
            cur.execute(f"DELETE FROM auth_sessions WHERE token = {PLACEHOLDER}", (token,))
            conn.commit()
            raise HTTPException(status_code=401, detail="Session expired")
        session_user = {
            "email": row_dict["user_email"],
            "full_name": row_dict["user_full_name"],
            "id": row_dict["user_id"],
        }
        SESSIONS[token] = {**session_user, "expires_at": expires_at}
        conn.commit()
        return session_user

# --- Auth endpoints ---
@app.post("/api/signup", response_model=UserOut)
def signup(user: UserCreate, background_tasks: BackgroundTasks):
    try:
        welcome_payload = {
            "email": user.email,
            "full_name": user.full_name,
            "club_name": "Glasgow Bengali FC",
        }
        with get_connection() as conn:
            cur = conn.cursor()
            
            # Check if email already exists
            cur.execute(
                f"SELECT id, is_deleted FROM users WHERE email = {PLACEHOLDER}",
                (user.email,)
            )
            existing = cur.fetchone()
            
            if existing:
                existing_dict = dict(existing)
                if existing_dict.get('is_deleted'):
                    # Reactivate deleted account
                    if USE_POSTGRES:
                        cur.execute(
                            f"UPDATE users SET is_deleted = FALSE, deleted_at = NULL, "
                            f"full_name = {PLACEHOLDER}, password = {PLACEHOLDER}, user_type = 'member', "
                            f"last_login = CURRENT_TIMESTAMP WHERE email = {PLACEHOLDER}",
                            (user.full_name, hash_password(user.password), user.email)
                        )
                    else:
                        cur.execute(
                            f"UPDATE users SET is_deleted = 0, deleted_at = NULL, "
                            f"full_name = {PLACEHOLDER}, password = {PLACEHOLDER}, user_type = 'member', "
                            f"last_login = CURRENT_TIMESTAMP WHERE email = {PLACEHOLDER}",
                            (user.full_name, hash_password(user.password), user.email)
                        )
                    conn.commit()
                    background_tasks.add_task(send_direct_notification_email_safe, "welcome_signup", welcome_payload, user.email)
                    return UserOut(id=existing_dict['id'], email=user.email, full_name=user.full_name, user_type='member')
                else:
                    # Active user already exists
                    raise HTTPException(status_code=400, detail="Email already registered")
            
            # Create new user with initial last_login timestamp
            if USE_POSTGRES:
                cur.execute(
                    f"INSERT INTO users (email, full_name, password, last_login) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, CURRENT_TIMESTAMP)",
                    (user.email, user.full_name, hash_password(user.password)),
                )
            else:
                cur.execute(
                    f"INSERT INTO users (email, full_name, password, last_login) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, CURRENT_TIMESTAMP)",
                    (user.email, user.full_name, hash_password(user.password)),
                )
            conn.commit()
            if USE_POSTGRES:
                # PostgreSQL doesn't support lastrowid, need to fetch the inserted row
                cur.execute(f"SELECT id FROM users WHERE email = {PLACEHOLDER}", (user.email,))
                user_id = cur.fetchone()['id']
            else:
                user_id = cur.lastrowid
            background_tasks.add_task(send_direct_notification_email_safe, "welcome_signup", welcome_payload, user.email)
            return UserOut(id=user_id, email=user.email, full_name=user.full_name, user_type='member')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/token", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends()):
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM users WHERE email = {PLACEHOLDER} AND (is_deleted = FALSE OR is_deleted IS NULL)", (form_data.username,))
        user = cur.fetchone()
        if not user:
            print(f"Login failed: no user found for email {form_data.username}")
            raise HTTPException(status_code=401, detail="Invalid email or password")
        # Debug: compare stored hash with computed hash
        input_hash = hash_password(form_data.password)
        stored_hash = user["password"]
        print(f"Login attempt for {form_data.username}: input_hash={input_hash}, stored_hash={stored_hash}")
        if not verify_password(form_data.password, user["password"]):
            print("Login failed: password mismatch")
            raise HTTPException(status_code=401, detail="Invalid email or password")
        
        # Update last_login timestamp
        if USE_POSTGRES:
            cur.execute(
                f"UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE email = {PLACEHOLDER}",
                (form_data.username,)
            )
        else:
            cur.execute(
                f"UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE email = {PLACEHOLDER}",
                (form_data.username,)
            )
        conn.commit()
        
        token = str(uuid.uuid4())
        persist_session_token(token, user)
        return {"access_token": token, "token_type": "bearer"}

@app.post("/api/login", response_model=Token)
def login_alias(form_data: OAuth2PasswordRequestForm = Depends()):
    return login(form_data)

# ========== FORGOT PASSWORD FEATURE - API Endpoint ==========
# This endpoint handles password recovery requests
# Frontend: Login.jsx (button currently hidden - see comments there)
# To enable: Configure email settings (see EMAIL_SETUP.md)
# PostgreSQL Compatible: Uses PLACEHOLDER for queries
# ============================================================
@app.post("/api/forgot-password")
def forgot_password(data: ForgotPasswordRequest):
    email = data.email.strip().lower()
    generic_message = {"message": "If this email is registered, a password reset link has been sent."}

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"SELECT email, full_name FROM users WHERE email = {PLACEHOLDER} AND (is_deleted = FALSE OR is_deleted IS NULL)",
            (email,)
        )
        user = cur.fetchone()

        if not user:
            return generic_message

        user_dict = dict(user)
        token = generate_password_reset_token()
        expires_at = datetime.utcnow() + timedelta(hours=1)

        cur.execute(
            f"DELETE FROM password_reset_tokens WHERE user_email = {PLACEHOLDER} OR expires_at < CURRENT_TIMESTAMP OR used = {PLACEHOLDER}",
            (email, True if USE_POSTGRES else 1)
        )
        cur.execute(
            f"INSERT INTO password_reset_tokens (user_email, token, expires_at, used) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
            (email, token, expires_at, False if USE_POSTGRES else 0)
        )
        conn.commit()

        reset_link = build_password_reset_link(token)
        subject = "Glasgow Bengali FC - Reset Your Password"
        body = f"""Hello {user_dict['full_name']},

We received a request to reset your Glasgow Bengali FC account password.

Use the secure link below to set a new password:
{reset_link}

This link will expire in 1 hour.

If you did not request this, please ignore this email.

Best regards,
Glasgow Bengali FC Team"""

        email_sent = send_email(email, subject, body)

        if not email_sent:
            raise HTTPException(
                status_code=503,
                detail="Email service is not configured. Please contact the administrator."
            )

        return generic_message

@app.post("/api/reset-password")
def reset_password(data: ResetPasswordRequest):
    token = data.token.strip()
    new_password = data.new_password

    if not token:
        raise HTTPException(status_code=400, detail="Reset token is required")
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    if len(new_password) > 128:
        raise HTTPException(status_code=400, detail="Password must be less than 128 characters")
    if not any(char.isalpha() for char in new_password):
        raise HTTPException(status_code=400, detail="Password must contain at least one letter")
    if not any(char.isdigit() for char in new_password):
        raise HTTPException(status_code=400, detail="Password must contain at least one number")

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"SELECT user_email, expires_at, used FROM password_reset_tokens WHERE token = {PLACEHOLDER}",
            (token,)
        )
        token_row = cur.fetchone()

        if not token_row:
            raise HTTPException(status_code=400, detail="Invalid or expired reset link")

        token_dict = dict(token_row)
        expires_at = token_dict["expires_at"]
        if isinstance(expires_at, str):
            expires_at = datetime.fromisoformat(expires_at.replace("Z", "+00:00")).replace(tzinfo=None)

        if token_dict.get("used") or expires_at < datetime.utcnow():
            raise HTTPException(status_code=400, detail="Invalid or expired reset link")

        cur.execute(
            f"UPDATE users SET password = {PLACEHOLDER} WHERE email = {PLACEHOLDER}",
            (hash_password(new_password), token_dict["user_email"])
        )
        cur.execute(
            f"UPDATE password_reset_tokens SET used = {PLACEHOLDER} WHERE token = {PLACEHOLDER}",
            (True if USE_POSTGRES else 1, token)
        )
        cur.execute(
            f"DELETE FROM password_reset_tokens WHERE user_email = {PLACEHOLDER} AND token != {PLACEHOLDER}",
            (token_dict["user_email"], token)
        )
        conn.commit()

        return {"message": "Password reset successful"}

@app.get("/api/me", response_model=UserOut)
def me(current_user: dict = Depends(get_current_user)):
    # Fetch user details from database including created_at and last_login
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"SELECT full_name, user_type, is_deleted, created_at, last_login, birthday, bank_name, sort_code, account_number, theme_preference FROM users WHERE email = {PLACEHOLDER}", 
            (current_user["email"],)
        )
        row = cur.fetchone()
        if row:
            row_dict = dict(row)
            # Check if user is deleted
            if row_dict.get("is_deleted"):
                raise HTTPException(status_code=401, detail="Account has been deleted")
            
            full_name = row_dict.get("full_name") or current_user["full_name"]
            user_type = row_dict.get("user_type") or "member"
            
            # Convert datetime fields to ISO string
            created_at = None
            if row_dict.get("created_at"):
                if hasattr(row_dict["created_at"], 'isoformat'):
                    created_at = row_dict["created_at"].isoformat()
                else:
                    created_at = str(row_dict["created_at"])
            
            last_login = None
            if row_dict.get("last_login"):
                if hasattr(row_dict["last_login"], 'isoformat'):
                    last_login = row_dict["last_login"].isoformat()
                else:
                    last_login = str(row_dict["last_login"])
            
            # Convert birthday date to ISO string
            birthday = None
            if row_dict.get("birthday"):
                if hasattr(row_dict["birthday"], 'isoformat'):
                    birthday = row_dict["birthday"].isoformat()
                else:
                    birthday = str(row_dict["birthday"])
            bank_name = row_dict.get("bank_name")
            sort_code = row_dict.get("sort_code")
            account_number = row_dict.get("account_number")
            theme_preference = row_dict.get("theme_preference") or "nordic_neutral"
        else:
            full_name = current_user["full_name"]
            user_type = "member"
            created_at = None
            last_login = None
            birthday = None
            bank_name = None
            sort_code = None
            account_number = None
            theme_preference = "nordic_neutral"
    
    return UserOut(
        id=current_user["id"], 
        email=current_user["email"], 
        full_name=full_name, 
        user_type=user_type,
        created_at=created_at,
        last_login=last_login,
        birthday=birthday,
        bank_name=bank_name,
        sort_code=sort_code,
        account_number=account_number,
        theme_preference=theme_preference,
    )

@app.post("/api/logout")
def logout(current_user: dict = Depends(get_current_user), token: str = Depends(oauth2_scheme)):
    delete_session_token(token)
    return {"message": "Logged out"}

@app.get("/api/users", response_model=List[UserOut])
def get_all_users(current_user: dict = Depends(get_current_user)):
    # Admin only
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT id, email, full_name, user_type, created_at, last_login, birthday, bank_name, sort_code, account_number, theme_preference FROM users WHERE (is_deleted = FALSE OR is_deleted IS NULL) ORDER BY id DESC")
        users = []
        for row in cur.fetchall():
            user_dict = dict(row)
            # Convert datetime to ISO string if needed, or set to None
            if user_dict.get("created_at"):
                if hasattr(user_dict["created_at"], 'isoformat'):
                    user_dict["created_at"] = user_dict["created_at"].isoformat()
                else:
                    user_dict["created_at"] = str(user_dict["created_at"])
            
            # Convert last_login datetime to ISO string if needed, or set to None
            if user_dict.get("last_login"):
                if hasattr(user_dict["last_login"], 'isoformat'):
                    user_dict["last_login"] = user_dict["last_login"].isoformat()
                else:
                    user_dict["last_login"] = str(user_dict["last_login"])
            
            # Convert birthday date to ISO string if needed, or set to None
            if user_dict.get("birthday"):
                if hasattr(user_dict["birthday"], 'isoformat'):
                    user_dict["birthday"] = user_dict["birthday"].isoformat()
                else:
                    user_dict["birthday"] = user_dict.get("birthday") or None
            user_dict["bank_name"] = user_dict.get("bank_name") or None
            user_dict["sort_code"] = user_dict.get("sort_code") or None
            user_dict["account_number"] = user_dict.get("account_number") or None
            user_dict["theme_preference"] = user_dict.get("theme_preference") or "nordic_neutral"
            # Ensure user_type has a default
            if not user_dict.get("user_type"):
                user_dict["user_type"] = "member"
            users.append(UserOut(**user_dict))
        return users

@app.patch("/api/users/{email}/type")
def update_user_type(email: str, data: dict, current_user: dict = Depends(get_current_user)):
    # Admin only
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    
    user_type = data.get("user_type")
    if user_type not in ["member", "admin"]:
        raise HTTPException(status_code=400, detail="Invalid user type. Must be 'member' or 'admin'")
    
    with get_connection() as conn:
        cur = conn.cursor()
        
        # Check if user exists
        cur.execute(f"SELECT id FROM users WHERE email = {PLACEHOLDER}", (email,))
        user = cur.fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Update user type
        cur.execute(f"UPDATE users SET user_type = {PLACEHOLDER} WHERE email = {PLACEHOLDER}", (user_type, email))
        conn.commit()
        
        return {"message": f"User type updated to {user_type}"}

@app.put("/api/users/{email}/name")
def update_user_name(email: str, data: dict, current_user: dict = Depends(get_current_user)):
    # Admin only
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    
    # Prevent editing super admin account
    if email == "super@admin.com":
        raise HTTPException(status_code=403, detail="Cannot edit super admin account")
    
    # Validate input
    full_name = data.get("full_name", "").strip()
    if not full_name:
        raise HTTPException(status_code=400, detail="Full name cannot be empty")
    
    if len(full_name) > 100:
        raise HTTPException(status_code=400, detail="Full name must be 100 characters or less")
    
    with get_connection() as conn:
        cur = conn.cursor()
        
        # Check if user exists
        cur.execute(f"SELECT id FROM users WHERE email = {PLACEHOLDER}", (email,))
        user = cur.fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Update user's full name
        cur.execute(
            f"UPDATE users SET full_name = {PLACEHOLDER} WHERE email = {PLACEHOLDER}",
            (full_name, email)
        )
        conn.commit()
        
        return {"message": "User name updated successfully"}

@app.put("/api/profile/name")
def update_own_name(data: dict, current_user: dict = Depends(get_current_user)):
    """Allow user to update their own name"""
    # Validate input
    full_name = data.get("full_name", "").strip()
    if not full_name:
        raise HTTPException(status_code=400, detail="Full name cannot be empty")
    
    if len(full_name) > 100:
        raise HTTPException(status_code=400, detail="Full name must be 100 characters or less")
    
    with get_connection() as conn:
        cur = conn.cursor()
        
        # Update user's own full name
        cur.execute(
            f"UPDATE users SET full_name = {PLACEHOLDER} WHERE email = {PLACEHOLDER}",
            (full_name, current_user["email"])
        )
        conn.commit()
        
        # Return updated user info
        return {"full_name": full_name, "email": current_user["email"]}

@app.put("/api/profile/password")
def update_own_password(data: dict, current_user: dict = Depends(get_current_user)):
    """Allow user to update their own password"""
    current_password = data.get("current_password", "")
    new_password = data.get("new_password", "")
    
    if not current_password or not new_password:
        raise HTTPException(status_code=400, detail="Current and new passwords are required")
    
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    
    with get_connection() as conn:
        cur = conn.cursor()
        
        # Verify current password
        cur.execute(
            f"SELECT password FROM users WHERE email = {PLACEHOLDER}",
            (current_user["email"],)
        )
        user = cur.fetchone()
        
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        if not verify_password(current_password, user["password"]):
            raise HTTPException(status_code=401, detail="Current password is incorrect")
        
        # Update password
        cur.execute(
            f"UPDATE users SET password = {PLACEHOLDER} WHERE email = {PLACEHOLDER}",
            (hash_password(new_password), current_user["email"])
        )
        conn.commit()
        
        return {"message": "Password updated successfully"}

@app.put("/api/profile/birthday")
def update_own_birthday(data: dict, current_user: dict = Depends(get_current_user)):
    """Allow user to update their own birthday"""
    birthday = data.get("birthday", "").strip()
    
    # Birthday is optional, allow empty string to clear it
    if birthday and birthday != "":
        # Validate date format (YYYY-MM-DD)
        try:
            from datetime import datetime
            datetime.strptime(birthday, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    with get_connection() as conn:
        cur = conn.cursor()
        
        # Update user's birthday (or set to NULL if empty)
        if birthday and birthday != "":
            cur.execute(
                f"UPDATE users SET birthday = {PLACEHOLDER} WHERE email = {PLACEHOLDER}",
                (birthday, current_user["email"])
            )
        else:
            cur.execute(
                f"UPDATE users SET birthday = NULL WHERE email = {PLACEHOLDER}",
                (current_user["email"],)
            )
        conn.commit()
        
        return {"message": "Birthday updated successfully", "birthday": birthday if birthday else None}

@app.put("/api/profile/bank-details")
def update_own_bank_details(data: dict, current_user: dict = Depends(get_current_user)):
    bank_name = (data.get("bank_name") or "").strip()
    sort_code = (data.get("sort_code") or "").strip()
    account_number = (data.get("account_number") or "").strip()

    if bank_name and len(bank_name) > 100:
        raise HTTPException(status_code=400, detail="Bank name must be 100 characters or less")

    if sort_code:
        normalized_sort_code = sort_code.replace("-", "").replace(" ", "")
        if not normalized_sort_code.isdigit() or len(normalized_sort_code) != 6:
            raise HTTPException(status_code=400, detail="Sort code must be 6 digits")
        sort_code = f"{normalized_sort_code[0:2]}-{normalized_sort_code[2:4]}-{normalized_sort_code[4:6]}"
    else:
        sort_code = None

    if account_number:
        normalized_account_number = account_number.replace(" ", "")
        if not normalized_account_number.isdigit() or len(normalized_account_number) < 6 or len(normalized_account_number) > 8:
            raise HTTPException(status_code=400, detail="Account number must be 6 to 8 digits")
        account_number = normalized_account_number
    else:
        account_number = None

    if any([bank_name, sort_code, account_number]) and not all([bank_name, sort_code, account_number]):
        raise HTTPException(status_code=400, detail="Bank name, sort code, and account number must all be provided together")

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"UPDATE users SET bank_name = {PLACEHOLDER}, sort_code = {PLACEHOLDER}, account_number = {PLACEHOLDER} WHERE email = {PLACEHOLDER}",
            (bank_name or None, sort_code, account_number, current_user["email"])
        )
        conn.commit()

    return {
        "message": "Bank details updated successfully",
        "bank_name": bank_name or None,
        "sort_code": sort_code,
        "account_number": account_number,
    }

@app.put("/api/profile/theme")
def update_own_theme(data: dict, current_user: dict = Depends(get_current_user)):
    theme_preference = (data.get("theme_preference") or "").strip()
    if theme_preference not in THEME_PREFERENCES:
        raise HTTPException(status_code=400, detail="Invalid theme preference")

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"UPDATE users SET theme_preference = {PLACEHOLDER} WHERE email = {PLACEHOLDER}",
            (theme_preference, current_user["email"])
        )
        conn.commit()

    return {"theme_preference": theme_preference}

@app.delete("/api/users/{email}")
def delete_user(email: str, current_user: dict = Depends(get_current_user)):
    # Admin only
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    
    # Prevent deleting super admin account
    if email == "super@admin.com":
        raise HTTPException(status_code=403, detail="Cannot delete super admin account")
    
    with get_connection() as conn:
        cur = conn.cursor()
        
        # Check if user exists and is not already deleted
        cur.execute(f"SELECT id, is_deleted FROM users WHERE email = {PLACEHOLDER}", (email,))
        user = cur.fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        user_dict = dict(user)
        if user_dict.get('is_deleted'):
            raise HTTPException(status_code=400, detail="User already deleted")
        
        # Soft delete: Mark user as deleted and record who deleted them
        # Also remove user's likes to prevent confusion when account is reactivated
        
        # Delete forum post likes
        cur.execute(f"DELETE FROM forum_likes WHERE user_email = {PLACEHOLDER}", (email,))
        
        # Delete event likes
        cur.execute(f"DELETE FROM event_likes WHERE user_email = {PLACEHOLDER}", (email,))
        
        # Delete only FUTURE practice availability (preserve historical records)
        if USE_POSTGRES:
            cur.execute(
                f"DELETE FROM practice_availability WHERE user_email = {PLACEHOLDER} AND date::date > CURRENT_DATE",
                (email,)
            )
        else:
            cur.execute(
                f"DELETE FROM practice_availability WHERE user_email = {PLACEHOLDER} AND date > date('now')",
                (email,)
            )
        
        # Mark user as deleted
        if USE_POSTGRES:
            cur.execute(
                f"UPDATE users SET is_deleted = TRUE, deleted_at = CURRENT_TIMESTAMP, "
                f"deleted_by = {PLACEHOLDER} WHERE email = {PLACEHOLDER}",
                (current_user["email"], email)
            )
        else:
            cur.execute(
                f"UPDATE users SET is_deleted = 1, deleted_at = CURRENT_TIMESTAMP, "
                f"deleted_by = {PLACEHOLDER} WHERE email = {PLACEHOLDER}",
                (current_user["email"], email)
            )
        
        conn.commit()
        return {"message": "User deleted successfully. Historical posts/comments preserved, likes removed."}

# --- Events endpoints ---
@app.get("/api/events", response_model=List[EventOut])
def get_events():
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"SELECT id, date, time, location, event_type, event_title, description, image_url, youtube_url FROM practice_sessions WHERE event_type = {PLACEHOLDER}",
            ("match",),
        )
        for session_row in cur.fetchall():
            sync_match_session_to_events(cur, dict(session_row))
        conn.commit()
        cur.execute("SELECT * FROM events ORDER BY date ASC")
        events = []
        for row in cur.fetchall():
            event = dict(row)
            event.pop("type", None)
            # likes with full names
            cur.execute(
                f"SELECT el.user_email, u.full_name FROM event_likes el JOIN users u ON el.user_email = u.email WHERE el.event_id = {PLACEHOLDER}",
                (event["id"],)
            )
            event["likes"] = [dict(r) for r in cur.fetchall()]
            # comments
            cur.execute(
                f"SELECT * FROM event_comments WHERE event_id = {PLACEHOLDER} ORDER BY created_at ASC",
                (event["id"],),
            )
            comments = []
            for comment_row in cur.fetchall():
                comment_dict = dict(comment_row)
                # Use stored user_full_name, fallback to users table if not set (for old comments)
                full_name = comment_dict.get("user_full_name")
                if not full_name:
                    cur.execute(f"SELECT full_name FROM users WHERE email = {PLACEHOLDER}", (comment_dict["user_email"],))
                    user_row = cur.fetchone()
                    full_name = user_row["full_name"] if user_row else comment_dict["user_email"]
                comment_dict["full_name"] = full_name
                comments.append(comment_dict)
            event["comments"] = comments
            events.append(EventOut(**event))
        return events

@app.get("/api/events/{event_id}", response_model=EventOut)
def get_event(event_id: int):
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM events WHERE id = {PLACEHOLDER}", (event_id,))
        event = cur.fetchone()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        event_dict = dict(event)
        event_dict.pop("type", None)
        return EventOut(**event_dict)

@app.post("/api/events", response_model=EventOut)
def create_event(event: EventCreate, current_user: dict = Depends(get_current_user)):
    # Admin only
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"INSERT INTO events (name, date, time, location, description, image_url, youtube_url) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
            (event.name, event.date, event.time, event.location, event.description, event.image_url, event.youtube_url),
        )
        conn.commit()
        
        # Get the inserted event ID (PostgreSQL compatible)
        if USE_POSTGRES:
            cur.execute(
                f"SELECT id FROM events WHERE name = {PLACEHOLDER} AND date = {PLACEHOLDER} ORDER BY id DESC LIMIT 1",
                (event.name, event.date)
            )
            event_id = cur.fetchone()['id']
        else:
            event_id = cur.lastrowid
        
        deliver_notification(
            "practice",
            {
                "date": event.date,
                "time": event.time,
                "location": event.location,
                "event_name": event.name,
            },
            related_date=event.date
        )

        return EventOut(id=event_id, **event.model_dump())

@app.put("/api/events/{event_id}", response_model=EventOut)
def update_event(event_id: int, event: EventCreate, current_user: dict = Depends(get_current_user)):
    # Admin only
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"UPDATE events SET name={PLACEHOLDER}, date={PLACEHOLDER}, time={PLACEHOLDER}, location={PLACEHOLDER}, description={PLACEHOLDER}, image_url={PLACEHOLDER}, youtube_url={PLACEHOLDER} WHERE id={PLACEHOLDER}",
            (event.name, event.date, event.time, event.location, event.description, event.image_url, event.youtube_url, event_id),
        )
        conn.commit()
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Event not found")
        return EventOut(id=event_id, **event.model_dump())

@app.delete("/api/events/{event_id}")
def delete_event(event_id: int, current_user: dict = Depends(get_current_user)):
    # Admin only
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    with get_connection() as conn:
        cur = conn.cursor()
        # Delete related records first to avoid foreign key constraint violations
        cur.execute(f"DELETE FROM event_likes WHERE event_id={PLACEHOLDER}", (event_id,))
        cur.execute(f"DELETE FROM event_comments WHERE event_id={PLACEHOLDER}", (event_id,))
        cur.execute(f"DELETE FROM events WHERE id={PLACEHOLDER}", (event_id,))
        conn.commit()
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Event not found")
        return {"message": "Event deleted"}

@app.post("/api/upload-image")
async def upload_image(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    # Admin only
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    # Validate file type
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="File must be an image")
    
    if CLOUDINARY_URL:
        # Upload to Cloudinary (production)
        file_content = await file.read()
        result = cloudinary.uploader.upload(file_content, folder="football_club/events")
        return {"image_url": result["secure_url"]}
    else:
        # Save locally (development)
        file_id = str(uuid.uuid4())
        ext = os.path.splitext(file.filename)[1]
        filename = f"{file_id}{ext}"
        path = os.path.join(UPLOAD_DIR, filename)
        with open(path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        return {"image_url": f"http://localhost:8000/{UPLOAD_DIR}/{filename}"}

@app.post("/api/forum/upload-image")
async def upload_forum_image(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="File must be an image")
    
    if CLOUDINARY_URL:
        # Upload to Cloudinary (production)
        file_content = await file.read()
        result = cloudinary.uploader.upload(file_content, folder="football_club/forum")
        return {"image_url": result["secure_url"]}
    else:
        # Save locally (development)
        file_id = str(uuid.uuid4())
        ext = os.path.splitext(file.filename)[1]
        filename = f"{file_id}{ext}"
        path = os.path.join(UPLOAD_DIR, filename)
        with open(path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        return {"image_url": f"http://localhost:8000/{UPLOAD_DIR}/{filename}"}

@app.get("/uploads/{filename}")
def get_uploaded_file(filename: str):
    path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path)

# --- Practice Sessions (admin-managed) ---
@app.get("/api/practice/sessions", response_model=List[PracticeSessionOut])
def list_practice_sessions():
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT id FROM practice_sessions ORDER BY date ASC, COALESCE(time, ''), id ASC")
        sessions = []
        for row in cur.fetchall():
            row_dict = dict(row)
            session = get_practice_session_with_capacity_by_id(cur, row_dict["id"])
            if session:
                sessions.append(PracticeSessionOut(**session))
        return sessions

@app.get("/api/practice/sessions/id/{session_id}", response_model=PracticeSessionOut)
def get_practice_session_by_id(session_id: int):
    with get_connection() as conn:
        cur = conn.cursor()
        session = get_practice_session_with_capacity_by_id(cur, session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Practice session not found")
        return PracticeSessionOut(**session)

@app.post("/api/practice/sessions", response_model=PracticeSessionOut)
def create_practice_session(session: PracticeSessionCreate, current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    maximum_capacity = normalize_maximum_capacity(session.maximum_capacity)
    normalized_time = normalize_practice_time(session.time)
    normalized_event_type = normalize_event_type(session.event_type)
    normalized_event_title = normalize_event_title(session.event_title, normalized_event_type)
    normalized_option_a, normalized_option_b = normalize_option_pair(session.option_a_text, session.option_b_text)
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"INSERT INTO practice_sessions (date, time, location, event_type, event_title, description, image_url, youtube_url, option_a_text, option_b_text, session_cost, paid_by, maximum_capacity) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
            (session.date, normalized_time, session.location, normalized_event_type, normalized_event_title, session.description, session.image_url, session.youtube_url, normalized_option_a, normalized_option_b, session.session_cost, session.paid_by, maximum_capacity),
        )
        created_session_id = cur.lastrowid if not USE_POSTGRES else None
        if USE_POSTGRES:
            cur.execute(
                f"SELECT id FROM practice_sessions WHERE date = {PLACEHOLDER} AND COALESCE(time, '') = {PLACEHOLDER} AND COALESCE(location, '') = COALESCE({PLACEHOLDER}, '') ORDER BY id DESC LIMIT 1",
                (session.date, normalized_time or '', session.location),
            )
            created_row = cur.fetchone()
            created_session_id = created_row["id"] if created_row else None
        sync_match_session_to_events(cur, {
            "id": created_session_id,
            "date": session.date,
            "time": normalized_time,
            "location": session.location,
            "event_type": normalized_event_type,
            "event_title": normalized_event_title,
            "description": session.description,
            "image_url": session.image_url,
            "youtube_url": session.youtube_url,
        })
        conn.commit()
        
        deliver_notification(
            "practice",
            {
                "session_id": created_session_id,
                "date": session.date,
                "time": normalized_time,
                "location": session.location,
                "event_type": normalized_event_type,
                "event_title": normalized_event_title,
                "event_name": f"{default_event_type_label(normalized_event_type)} - {normalized_event_title}",
                "maximum_capacity": maximum_capacity,
            },
            related_date=session.date
        )

        created_session = get_practice_session_with_capacity_by_id(cur, created_session_id)
        return PracticeSessionOut(**created_session)

@app.put("/api/practice/sessions/id/{session_id}", response_model=PracticeSessionOut)
def update_practice_session_by_id(session_id: int, session: PracticeSessionCreate, current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    maximum_capacity = normalize_maximum_capacity(session.maximum_capacity)
    normalized_time = normalize_practice_time(session.time)
    normalized_event_type = normalize_event_type(session.event_type)
    normalized_event_title = normalize_event_title(session.event_title, normalized_event_type)
    normalized_option_a, normalized_option_b = normalize_option_pair(session.option_a_text, session.option_b_text)
    with get_connection() as conn:
        cur = conn.cursor()
        existing_session = get_practice_session_basic_by_id(cur, session_id)
        if not existing_session:
            raise HTTPException(status_code=404, detail="Practice session not found")
        if existing_session["payment_requested"]:
            raise HTTPException(status_code=400, detail="Practice session cannot be edited after payment has been requested")
        cur.execute(
            f"UPDATE practice_sessions SET date = {PLACEHOLDER}, time = {PLACEHOLDER}, location = {PLACEHOLDER}, event_type = {PLACEHOLDER}, event_title = {PLACEHOLDER}, description = {PLACEHOLDER}, image_url = {PLACEHOLDER}, youtube_url = {PLACEHOLDER}, option_a_text = {PLACEHOLDER}, option_b_text = {PLACEHOLDER}, session_cost = {PLACEHOLDER}, paid_by = {PLACEHOLDER}, maximum_capacity = {PLACEHOLDER} WHERE id = {PLACEHOLDER}",
            (session.date, normalized_time, session.location, normalized_event_type, normalized_event_title, session.description, session.image_url, session.youtube_url, normalized_option_a, normalized_option_b, session.session_cost, session.paid_by, maximum_capacity, session_id),
        )
        sync_match_session_to_events(cur, {
            "id": session_id,
            "date": session.date,
            "time": normalized_time,
            "location": session.location,
            "event_type": normalized_event_type,
            "event_title": normalized_event_title,
            "description": session.description,
            "image_url": session.image_url,
            "youtube_url": session.youtube_url,
        })
        conn.commit()
        updated_session = get_practice_session_with_capacity_by_id(cur, session_id)
        return PracticeSessionOut(**updated_session)

@app.put("/api/practice/sessions/{date_str}", response_model=PracticeSessionOut)
def update_practice_session(date_str: str, session: PracticeSessionCreate, current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    maximum_capacity = normalize_maximum_capacity(session.maximum_capacity)
    normalized_time = normalize_practice_time(session.time)
    normalized_event_type = normalize_event_type(session.event_type)
    normalized_event_title = normalize_event_title(session.event_title, normalized_event_type)
    normalized_option_a, normalized_option_b = normalize_option_pair(session.option_a_text, session.option_b_text)
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"SELECT payment_requested FROM practice_sessions WHERE date = {PLACEHOLDER}",
            (date_str,),
        )
        existing_session = cur.fetchone()
        if not existing_session:
            raise HTTPException(status_code=404, detail="Practice session not found")
        if existing_session["payment_requested"]:
            raise HTTPException(status_code=400, detail="Practice session cannot be edited after payment has been requested")
        cur.execute(
            f"UPDATE practice_sessions SET time = {PLACEHOLDER}, location = {PLACEHOLDER}, event_type = {PLACEHOLDER}, event_title = {PLACEHOLDER}, description = {PLACEHOLDER}, image_url = {PLACEHOLDER}, youtube_url = {PLACEHOLDER}, option_a_text = {PLACEHOLDER}, option_b_text = {PLACEHOLDER}, session_cost = {PLACEHOLDER}, paid_by = {PLACEHOLDER}, maximum_capacity = {PLACEHOLDER} WHERE date = {PLACEHOLDER}",
            (normalized_time, session.location, normalized_event_type, normalized_event_title, session.description, session.image_url, session.youtube_url, normalized_option_a, normalized_option_b, session.session_cost, session.paid_by, maximum_capacity, date_str),
        )
        updated_rows = cur.rowcount
        updated_session_id = get_practice_session_id_by_date(cur, date_str)
        sync_match_session_to_events(cur, {
            "id": updated_session_id,
            "date": date_str,
            "time": normalized_time,
            "location": session.location,
            "event_type": normalized_event_type,
            "event_title": normalized_event_title,
            "description": session.description,
            "image_url": session.image_url,
            "youtube_url": session.youtube_url,
        })
        conn.commit()
        if updated_rows == 0:
            raise HTTPException(status_code=404, detail="Practice session not found")
        updated_session = get_practice_session_with_capacity(cur, date_str)
        return PracticeSessionOut(**updated_session)

@app.post("/api/practice/sessions/id/{session_id}/request-payment")
def request_payment_by_id(session_id: int, current_user: dict = Depends(get_current_user)):
    """Admin endpoint to enable payment request for a specific practice session"""
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")

    with get_connection() as conn:
        cur = conn.cursor()
        session = get_practice_session_with_capacity_by_id(cur, session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Practice session not found")

        if not is_practice_datetime_in_past(session["date"], session["time"]):
            raise HTTPException(status_code=400, detail="Payment request can only be enabled after the practice session date and time has passed")

        if session["payment_requested"]:
            raise HTTPException(status_code=400, detail="Payment request has already been enabled for this session")

        cur.execute(
            f"UPDATE practice_sessions SET payment_requested = {PLACEHOLDER}, payment_requested_at = CURRENT_TIMESTAMP WHERE id = {PLACEHOLDER}",
            (True if USE_POSTGRES else 1, session_id),
        )
        conn.commit()

        cur.execute(
            f"SELECT user_email FROM practice_availability WHERE practice_session_id = {PLACEHOLDER} AND status = {PLACEHOLDER}",
            (session_id, "available"),
        )
        available_users = cur.fetchall()

        deliver_notification(
            "payment_request",
            {
                "session_id": session["id"],
                "date": session["date"],
                "time": session["time"],
                "location": session["location"],
                "event_type": session["event_type"],
                "event_title": session["event_title"],
            },
            related_date=session["date"]
        )

        return {"message": "Payment requested successfully"}

@app.post("/api/practice/sessions/{date_str}/request-payment")
def request_payment(date_str: str, current_user: dict = Depends(get_current_user)):
    """Admin endpoint to enable payment request for a practice session"""
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")

    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"SELECT id FROM practice_sessions WHERE date = {PLACEHOLDER} ORDER BY id ASC LIMIT 1",
            (date_str,),
        )
        session = cur.fetchone()
    if not session:
        raise HTTPException(status_code=404, detail="Practice session not found")
    return request_payment_by_id(session["id"], current_user)

@app.get("/api/practice/sessions/id/{session_id}/payments")
def get_session_payments_by_id(session_id: int, current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        session = get_practice_session_with_capacity_by_id(cur, session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Practice session not found")
        cur.execute(
            f"SELECT user_email, paid FROM practice_payments WHERE practice_session_id = {PLACEHOLDER}",
            (session_id,),
        )
        payments = {}
        for row in cur.fetchall():
            row_dict = dict(row)
            payments[row_dict["user_email"]] = row_dict["paid"]
        return payments

@app.get("/api/practice/sessions/{date_str}/payments")
def get_session_payments(date_str: str, current_user: dict = Depends(get_current_user)):
    """Get payment status for all users in a practice session"""
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"SELECT user_email, paid FROM practice_payments WHERE date = {PLACEHOLDER}",
            (date_str,),
        )
        payments = {}
        for row in cur.fetchall():
            row_dict = dict(row)
            payments[row_dict["user_email"]] = row_dict["paid"]
        return payments

@app.post("/api/practice/sessions/id/{session_id}/payment")
def confirm_payment_by_id(session_id: int, data: dict, current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        session = get_practice_session_with_capacity_by_id(cur, session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Practice session not found")

        paid = data.get("paid", False)

        cur.execute(
            f"SELECT id, date, payment_requested, time, location, event_type, event_title FROM practice_sessions WHERE id = {PLACEHOLDER}",
            (session_id,)
        )
        session = cur.fetchone()
        if not session:
            raise HTTPException(status_code=404, detail="Practice session not found")

        if not session["payment_requested"]:
            raise HTTPException(status_code=400, detail="Payment has not been requested for this session")

        cur.execute(
            f"SELECT status FROM practice_availability WHERE practice_session_id = {PLACEHOLDER} AND user_email = {PLACEHOLDER}",
            (session_id, current_user["email"]),
        )
        availability = cur.fetchone()
        if not availability or availability["status"] != "available":
            raise HTTPException(status_code=400, detail="You must be marked as available for this session to confirm payment")

        if USE_POSTGRES:
            cur.execute(
                f"INSERT INTO practice_payments (practice_session_id, date, user_email, paid) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}) ON CONFLICT (practice_session_id, user_email) DO UPDATE SET date = EXCLUDED.date, paid = EXCLUDED.paid",
                (session_id, session["date"], current_user["email"], paid),
            )
        else:
            cur.execute(
                f"INSERT OR REPLACE INTO practice_payments (id, practice_session_id, date, user_email, paid) VALUES ((SELECT id FROM practice_payments WHERE practice_session_id = {PLACEHOLDER} AND user_email = {PLACEHOLDER}), {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
                (session_id, current_user["email"], session_id, session["date"], current_user["email"], 1 if paid else 0),
            )

        conn.commit()

        if paid:
            deliver_notification(
                "payment_confirmed",
                {
                    "session_id": session["id"],
                    "date": session["date"],
                    "time": session["time"],
                    "location": session["location"],
                    "event_type": session["event_type"],
                    "event_title": session["event_title"],
                    "member_name": current_user.get("full_name", current_user["email"]),
                    "full_name": current_user.get("full_name", current_user["email"]),
                },
                related_date=session["date"],
            )
        return {"message": "Payment confirmation updated", "paid": paid}

@app.post("/api/practice/{date}/payment")
def confirm_payment_by_date(date: str, data: dict, current_user: dict = Depends(get_current_user)):
    """User endpoint to confirm or unconfirm payment for a practice session (used by User Actions page)"""
    paid = data.get("paid", False)
    
    with get_connection() as conn:
        cur = conn.cursor()
        session_id = get_practice_session_id_by_date(cur, date)
        if session_id is None:
            raise HTTPException(status_code=404, detail="Practice session not found")
        
        # Check if payment is requested for this session and get session details
        cur.execute(
            f"SELECT id, payment_requested, time, location, event_type, event_title FROM practice_sessions WHERE date = {PLACEHOLDER}",
            (date,)
        )
        session = cur.fetchone()
        if not session:
            raise HTTPException(status_code=404, detail="Practice session not found")
        
        if not session["payment_requested"]:
            raise HTTPException(status_code=400, detail="Payment request has not been enabled for this session")

        session_time = session["time"] or "TBD"
        session_location = session["location"] or "TBD"
        
        # Check if user is available for this session
        cur.execute(
            f"SELECT status FROM practice_availability WHERE date = {PLACEHOLDER} AND user_email = {PLACEHOLDER}",
            (date, current_user["email"]),
        )
        availability = cur.fetchone()
        if not availability or availability["status"] != "available":
            raise HTTPException(status_code=400, detail="You must be marked as available for this session to confirm payment")
        
        # Insert or update payment confirmation
        if USE_POSTGRES:
            cur.execute(
                f"INSERT INTO practice_payments (practice_session_id, date, user_email, paid) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}) ON CONFLICT (date, user_email) DO UPDATE SET practice_session_id = EXCLUDED.practice_session_id, paid = EXCLUDED.paid",
                (session_id, date, current_user["email"], paid),
            )
        else:
            cur.execute(
                f"INSERT OR REPLACE INTO practice_payments (practice_session_id, date, user_email, paid) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
                (session_id, date, current_user["email"], 1 if paid else 0),
            )
        
        conn.commit()
        
        if paid:
            deliver_notification(
                "payment_confirmed",
                {
                    "session_id": session["id"],
                    "date": date,
                    "time": session["time"],
                    "location": session["location"],
                    "event_type": session["event_type"],
                    "event_title": session["event_title"],
                    "member_name": current_user.get("full_name", current_user["email"]),
                    "full_name": current_user.get("full_name", current_user["email"]),
                },
                related_date=date,
            )
        return {"message": "Payment confirmation updated", "paid": paid}

@app.post("/api/practice/sessions/{date_str}/payment")
def confirm_payment(date_str: str, data: dict, current_user: dict = Depends(get_current_user)):
    return confirm_payment_by_date(date_str, data, current_user)

@app.post("/api/practice/sessions/id/{session_id}/availability")
def set_practice_availability_by_session_id(session_id: int, status: dict, current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        return set_practice_availability_for_session_id(cur, conn, session_id, current_user["email"], current_user["full_name"], status.get("status", ""), status.get("option_choice"))

@app.post("/api/admin/practice/sessions/id/{session_id}/availability")
def admin_set_practice_availability_by_session_id(session_id: int, payload: dict, current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        session = get_practice_session_with_capacity_by_id(cur, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Practice session not found")
    avail = AdminPracticeAvailability(
        date=session["date"],
        user_email=payload.get("user_email", ""),
        status=payload.get("status", ""),
        option_choice=payload.get("option_choice"),
    )
    with get_connection() as conn:
        cur = conn.cursor()
        if not is_admin(current_user):
            raise HTTPException(status_code=403, detail="Admins only")
        cur.execute(f"SELECT full_name FROM users WHERE email = {PLACEHOLDER}", (avail.user_email,))
        user_row = cur.fetchone()
        if not user_row:
            raise HTTPException(status_code=404, detail="User not found")
        return admin_set_practice_availability_for_session_id(cur, conn, session_id, avail.user_email, user_row["full_name"], avail.status, avail.option_choice)

@app.get("/api/practice/sessions/id/{session_id}/availability")
def get_practice_availability_summary_by_session_id(session_id: int):
    with get_connection() as conn:
        cur = conn.cursor()
        session = get_practice_session_with_capacity_by_id(cur, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Practice session not found")
    with get_connection() as conn:
        cur = conn.cursor()
        session = get_practice_session_with_capacity_by_id(cur, session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Practice session not found")
        return get_practice_availability_summary_by_session(cur, session)

@app.delete("/api/practice/sessions/id/{session_id}")
def delete_practice_by_id(session_id: int, current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        session = get_practice_session_basic_by_id(cur, session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Practice session not found")
        if not is_admin(current_user):
            raise HTTPException(status_code=403, detail="Admins only")
        if session["payment_requested"]:
            raise HTTPException(status_code=400, detail="Practice session cannot be deleted after payment has been requested")
        cur.execute(f"DELETE FROM event_media WHERE event_id IN (SELECT id FROM events WHERE practice_session_id = {PLACEHOLDER})", (session_id,))
        cur.execute(f"DELETE FROM event_likes WHERE event_id IN (SELECT id FROM events WHERE practice_session_id = {PLACEHOLDER})", (session_id,))
        cur.execute(f"DELETE FROM event_comments WHERE event_id IN (SELECT id FROM events WHERE practice_session_id = {PLACEHOLDER})", (session_id,))
        cur.execute(f"DELETE FROM events WHERE practice_session_id = {PLACEHOLDER}", (session_id,))
        cur.execute(f"DELETE FROM practice_payments WHERE practice_session_id = {PLACEHOLDER}", (session_id,))
        cur.execute(f"DELETE FROM practice_availability WHERE practice_session_id = {PLACEHOLDER}", (session_id,))
        cur.execute(f"DELETE FROM practice_sessions WHERE id = {PLACEHOLDER}", (session_id,))
        conn.commit()
        return {"message": "Practice session deleted"}

# --- Likes/Comments for Events ---
@app.post("/api/events/{event_id}/like")
def like_event(event_id: int, current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"INSERT INTO event_likes (event_id, user_email) VALUES ({PLACEHOLDER}, {PLACEHOLDER}) ON CONFLICT DO NOTHING",
            (event_id, current_user["email"]),
        )
        conn.commit()
        return {"message": "Liked"}

@app.delete("/api/events/{event_id}/like")
def unlike_event(event_id: int, current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"DELETE FROM event_likes WHERE event_id = {PLACEHOLDER} AND user_email = {PLACEHOLDER}",
            (event_id, current_user["email"]),
        )
        conn.commit()
        return {"message": "Unliked"}

@app.get("/api/expenses", response_model=List[ExpenseOut])
def list_expenses(current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT e.id, e.title, e.amount, e.paid_by, e.expense_date, e.category, e.payment_method,
                   e.description, e.created_at, e.updated_at, u.full_name as paid_by_name
            FROM expenses e
            LEFT JOIN users u ON e.paid_by = u.email AND (u.is_deleted = FALSE OR u.is_deleted IS NULL)
            """
        )
        expense_rows = []
        for row in cur.fetchall():
            row_dict = dict(row)
            row_dict.update(
                {
                    "source": "expense",
                    "is_booking_expense": False,
                    "practice_session_date": None,
                    "linked_practice_time": None,
                    "linked_practice_location": None,
                    "can_edit": True,
                    "can_delete": True,
                }
            )
            expense_rows.append(serialize_expense(row_dict))

        cur.execute(
            f"""
            SELECT ps.id, ps.date, ps.event_type, ps.event_title, ps.time, ps.location, ps.session_cost, ps.paid_by, ps.payment_requested_at,
                   u.full_name as paid_by_name
            FROM practice_sessions ps
            LEFT JOIN users u ON ps.paid_by = u.email AND (u.is_deleted = FALSE OR u.is_deleted IS NULL)
            WHERE ps.session_cost IS NOT NULL
            """
        )
        booking_rows = []
        for row in cur.fetchall():
            row_dict = dict(row)
            session_date = row_dict.get("date")
            event_type = normalize_event_type(row_dict.get("event_type")) if row_dict.get("event_type") else "practice"
            event_title = normalize_event_title(row_dict.get("event_title"), event_type)
            event_name = f"{default_event_type_label(event_type)} - {event_title}"
            time_value = row_dict.get("time")
            location_value = row_dict.get("location")
            if time_value and location_value:
                description = f"{event_name} booking at {time_value} - {location_value}"
            elif time_value:
                description = f"{event_name} booking at {time_value}"
            elif location_value:
                description = f"{event_name} booking - {location_value}"
            else:
                description = f"{event_name} booking cost"
            booking_rows.append(
                serialize_expense(
                    {
                        "id": -int(row_dict.get("id")),  # Use actual session ID as negative
                        "title": event_name,
                        "amount": row_dict.get("session_cost") or 0,
                        "paid_by": row_dict.get("paid_by"),
                        "expense_date": session_date,
                        "category": "Event Booking",
                        "payment_method": None,
                        "description": description,
                        "paid_by_name": row_dict.get("paid_by_name"),
                        "source": "practice_booking",
                        "is_booking_expense": True,
                        "practice_session_date": session_date,
                        "linked_practice_time": row_dict.get("time"),
                        "linked_practice_location": row_dict.get("location"),
                        "can_edit": False,
                        "can_delete": False,
                        "created_at": row_dict.get("payment_requested_at"),
                        "updated_at": row_dict.get("payment_requested_at"),
                    }
                )
            )

        merged_rows = expense_rows + booking_rows
        merged_rows.sort(
            key=lambda item: (
                item.get("expense_date") or "",
                item.get("created_at") or "",
                item.get("id") or 0,
            ),
            reverse=True,
        )
        return [ExpenseOut(**row) for row in merged_rows]

@app.post("/api/expenses", response_model=ExpenseOut)
def create_expense(expense: ExpenseCreate, current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    paid_by = (expense.paid_by or "").strip() or None
    with get_connection() as conn:
        cur = conn.cursor()
        if USE_POSTGRES:
            cur.execute(
                f"""
                INSERT INTO expenses (title, amount, paid_by, expense_date, category, payment_method, description, updated_at)
                VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, CURRENT_TIMESTAMP)
                RETURNING id
                """,
                (expense.title, expense.amount, paid_by, expense.expense_date, expense.category, expense.payment_method, expense.description),
            )
            expense_id = dict(cur.fetchone())["id"]
        else:
            cur.execute(
                f"""
                INSERT INTO expenses (title, amount, paid_by, expense_date, category, payment_method, description, updated_at)
                VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, CURRENT_TIMESTAMP)
                """,
                (expense.title, expense.amount, paid_by, expense.expense_date, expense.category, expense.payment_method, expense.description),
            )
            expense_id = cur.lastrowid
        conn.commit()
        cur.execute(
            f"""
            SELECT e.id, e.title, e.amount, e.paid_by, e.expense_date, e.category, e.payment_method,
                   e.description, e.created_at, e.updated_at, u.full_name as paid_by_name
            FROM expenses e
            LEFT JOIN users u ON e.paid_by = u.email AND (u.is_deleted = FALSE OR u.is_deleted IS NULL)
            WHERE e.id = {PLACEHOLDER}
            """,
            (expense_id,),
        )
        row = cur.fetchone()
        return ExpenseOut(**serialize_expense(dict(row)))

@app.put("/api/expenses/{expense_id}", response_model=ExpenseOut)
def update_expense(expense_id: int, expense: ExpenseCreate, current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    paid_by = (expense.paid_by or "").strip() or None
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            UPDATE expenses
            SET title = {PLACEHOLDER}, amount = {PLACEHOLDER}, paid_by = {PLACEHOLDER}, expense_date = {PLACEHOLDER},
                category = {PLACEHOLDER}, payment_method = {PLACEHOLDER}, description = {PLACEHOLDER}, updated_at = CURRENT_TIMESTAMP
            WHERE id = {PLACEHOLDER}
            """,
            (expense.title, expense.amount, paid_by, expense.expense_date, expense.category, expense.payment_method, expense.description, expense_id),
        )
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Expense not found")
        conn.commit()
        cur.execute(
            f"""
            SELECT e.id, e.title, e.amount, e.paid_by, e.expense_date, e.category, e.payment_method,
                   e.description, e.created_at, e.updated_at, u.full_name as paid_by_name
            FROM expenses e
            LEFT JOIN users u ON e.paid_by = u.email AND (u.is_deleted = FALSE OR u.is_deleted IS NULL)
            WHERE e.id = {PLACEHOLDER}
            """,
            (expense_id,),
        )
        row = cur.fetchone()
        return ExpenseOut(**serialize_expense(dict(row)))

@app.delete("/api/expenses/{expense_id}")
def delete_expense(expense_id: int, current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"DELETE FROM expenses WHERE id = {PLACEHOLDER}", (expense_id,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Expense not found")
        conn.commit()
        return {"message": "Expense deleted"}

@app.get("/api/events/likes/me")
def get_my_event_likes(current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"SELECT event_id FROM event_likes WHERE user_email = {PLACEHOLDER}",
            (current_user["email"],),
        )
        return [row["event_id"] for row in cur.fetchall()]

@app.get("/api/events/{event_id}/comments", response_model=List[EventComment])
def get_event_comments(event_id: int):
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"SELECT * FROM event_comments WHERE event_id = {PLACEHOLDER} ORDER BY created_at ASC",
            (event_id,),
        )
        comments = []
        for row in cur.fetchall():
            comment_dict = dict(row)
            # Convert datetime to ISO string if needed
            if hasattr(comment_dict.get("created_at"), 'isoformat'):
                comment_dict["created_at"] = comment_dict["created_at"].isoformat()
            comments.append(EventComment(**comment_dict))
        return comments

@app.post("/api/events/{event_id}/comments", response_model=EventComment)
def create_event_comment(event_id: int, comment: dict, current_user: dict = Depends(get_current_user)):
    # Validate comment length (max 100 characters for security)
    if len(comment["comment"]) > 100:
        raise HTTPException(status_code=400, detail="Comment must be 100 characters or less")
    
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"INSERT INTO event_comments (event_id, user_email, user_full_name, comment) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
            (event_id, current_user["email"], current_user["full_name"], comment["comment"]),
        )
        conn.commit()
        return EventComment(id=cur.lastrowid, user_email=current_user["email"], comment=comment["comment"], created_at=datetime.utcnow().isoformat())

# --- Forum endpoints ---
@app.get("/api/forum", response_model=List[ForumPostOut])
def get_forum_posts():
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM forum_posts ORDER BY created_at DESC")
        posts = []
        for row in cur.fetchall():
            post = dict(row)
            post["content"] = sanitize_forum_post_html(post.get("content") or "")
            # Use stored user_full_name, fallback to email if not set (for old posts)
            post_full_name = post.get("user_full_name")
            if not post_full_name:
                cur.execute(f"SELECT full_name FROM users WHERE email = {PLACEHOLDER}", (post["user_email"],))
                user_row = cur.fetchone()
                post_full_name = user_row["full_name"] if user_row else post["user_email"]
            # likes with full names
            cur.execute(
                f"SELECT fl.user_email, u.full_name FROM forum_likes fl JOIN users u ON fl.user_email = u.email WHERE fl.post_id = {PLACEHOLDER}",
                (post["id"],)
            )
            likes_list = [dict(r) for r in cur.fetchall()]
            likes_count = len(likes_list)
            # comments
            cur.execute(
                f"SELECT * FROM forum_comments WHERE post_id = {PLACEHOLDER} ORDER BY created_at ASC",
                (post["id"],),
            )
            comments = []
            for c in cur.fetchall():
                cd = dict(c)
                # Use stored user_full_name, fallback to users table if not set (for old comments)
                c_full_name = cd.get("user_full_name")
                if not c_full_name:
                    cur.execute(f"SELECT full_name FROM users WHERE email = {PLACEHOLDER}", (cd["user_email"],))
                    cu = cur.fetchone()
                    c_full_name = cu["full_name"] if cu else cd["user_email"]
                # Convert datetime to ISO string if needed
                created_at_str = cd["created_at"].isoformat() if hasattr(cd["created_at"], 'isoformat') else cd["created_at"]
                comments.append(
                    ForumCommentOut(
                        id=cd["id"],
                        user_full_name=c_full_name,
                        comment=cd["comment"],
                        created_at=created_at_str,
                    )
                )
            # Convert datetime to ISO string if needed
            post_created_at = post["created_at"].isoformat() if hasattr(post["created_at"], 'isoformat') else post["created_at"]
            posts.append(
                ForumPostOut(
                    id=post["id"],
                    user_full_name=post_full_name,
                    user_email=post["user_email"],
                    content=post["content"],
                    created_at=post_created_at,
                    likes_count=likes_count,
                    likes=likes_list,
                    comments=comments,
                )
            )
        return posts

@app.post("/api/forum", response_model=ForumPostOut)
def create_forum_post(post: ForumPostCreate, current_user: dict = Depends(get_current_user)):
    sanitized_content = sanitize_forum_post_html(post.content)
    # Validate visible text length instead of generated HTML length so embedded media markup does not fail valid posts
    plain_text_content = re.sub(r"<[^>]+>", "", sanitized_content or "")
    if len(html.unescape(plain_text_content).strip()) > 500:
        raise HTTPException(status_code=400, detail="Post content must be 500 characters or less")
    if not sanitized_content:
        raise HTTPException(status_code=400, detail="Post content cannot be empty")
    
    with get_connection() as conn:
        cur = conn.cursor()
        if USE_POSTGRES:
            cur.execute(
                f"INSERT INTO forum_posts (user_email, user_full_name, content) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}) RETURNING id, created_at",
                (current_user["email"], current_user["full_name"], sanitized_content),
            )
            inserted_post = cur.fetchone()
            post_id = inserted_post["id"]
            created_at = inserted_post["created_at"].isoformat() if hasattr(inserted_post["created_at"], 'isoformat') else inserted_post["created_at"]
        else:
            cur.execute(
                f"INSERT INTO forum_posts (user_email, user_full_name, content) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
                (current_user["email"], current_user["full_name"], sanitized_content),
            )
            post_id = cur.lastrowid
            created_at = datetime.utcnow().isoformat()
        conn.commit()
        
        deliver_notification(
            "forum_post",
            {
                "author_name": current_user["full_name"],
                "content": sanitized_content,
            }
        )
        
        return ForumPostOut(
            id=post_id,
            user_full_name=current_user["full_name"],
            user_email=current_user["email"],
            content=sanitized_content,
            created_at=created_at,
            likes_count=0,
            comments=[],
        )

@app.put("/api/forum/{post_id}", response_model=ForumPostOut)
def update_forum_post(
    post_id: int,
    payload: ForumPostUpdate,
    current_user: dict = Depends(get_current_user),
):
    sanitized_content = sanitize_forum_post_html(payload.content)
    plain_text_content = re.sub(r"<[^>]+>", "", sanitized_content or "")
    if len(html.unescape(plain_text_content).strip()) > 500:
        raise HTTPException(status_code=400, detail="Post content must be 500 characters or less")
    if not sanitized_content:
        raise HTTPException(status_code=400, detail="Post content cannot be empty")

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT * FROM forum_posts WHERE id = {PLACEHOLDER}", (post_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Post not found")

        if row["user_email"] != current_user["email"] and not is_admin(current_user):
            raise HTTPException(status_code=403, detail="You can only edit your own posts")

        cur.execute(f"UPDATE forum_posts SET content = {PLACEHOLDER} WHERE id = {PLACEHOLDER}", (sanitized_content, post_id))
        conn.commit()

        post = dict(row)
        post["content"] = sanitized_content

        # Use stored user_full_name, fallback to users table if not set (for old posts)
        post_full_name = post.get("user_full_name")
        if not post_full_name:
            cur.execute(f"SELECT full_name FROM users WHERE email = {PLACEHOLDER}", (post["user_email"],))
            user_row = cur.fetchone()
            post_full_name = user_row["full_name"] if user_row else post["user_email"]

        cur.execute(f"SELECT COUNT(*) as cnt FROM forum_likes WHERE post_id = {PLACEHOLDER}", (post_id,))
        likes = cur.fetchone()["cnt"]

        cur.execute(
            f"SELECT * FROM forum_comments WHERE post_id = {PLACEHOLDER} ORDER BY created_at ASC",
            (post_id,),
        )
        comments = []
        for c in cur.fetchall():
            cd = dict(c)
            # Use stored user_full_name, fallback to users table if not set (for old comments)
            c_full_name = cd.get("user_full_name")
            if not c_full_name:
                cur.execute(f"SELECT full_name FROM users WHERE email = {PLACEHOLDER}", (cd["user_email"],))
                cu = cur.fetchone()
                c_full_name = cu["full_name"] if cu else cd["user_email"]
            # Convert datetime to ISO string if needed
            created_at_str = cd["created_at"].isoformat() if hasattr(cd["created_at"], 'isoformat') else cd["created_at"]
            comments.append(
                ForumCommentOut(
                    id=cd["id"],
                    user_full_name=c_full_name,
                    comment=cd["comment"],
                    created_at=created_at_str,
                )
            )

        # Convert datetime to ISO string if needed
        post_created_at = post["created_at"].isoformat() if hasattr(post["created_at"], 'isoformat') else post["created_at"]
        return ForumPostOut(
            id=post_id,
            user_full_name=post_full_name,
            user_email=post["user_email"],
            content=sanitized_content,
            created_at=post_created_at,
            likes_count=likes,
            comments=comments,
        )


@app.delete("/api/forum/{post_id}")
def delete_forum_post(post_id: int, current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT user_email FROM forum_posts WHERE id = {PLACEHOLDER}", (post_id,))
        post = cur.fetchone()
        if not post:
            raise HTTPException(status_code=404, detail="Post not found")
        
        # Allow deletion if user owns the post OR is admin
        if post["user_email"] != current_user["email"] and not is_admin(current_user):
            raise HTTPException(status_code=403, detail="You can only delete your own posts")

        cur.execute(f"DELETE FROM forum_likes WHERE post_id = {PLACEHOLDER}", (post_id,))
        cur.execute(f"DELETE FROM forum_comments WHERE post_id = {PLACEHOLDER}", (post_id,))
        cur.execute(f"DELETE FROM forum_posts WHERE id = {PLACEHOLDER}", (post_id,))
        conn.commit()
        return {"message": "Post deleted"}

@app.post("/api/forum/{post_id}/like")
def like_forum_post(post_id: int, current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"INSERT INTO forum_likes (post_id, user_email) VALUES ({PLACEHOLDER}, {PLACEHOLDER}) ON CONFLICT DO NOTHING",
            (post_id, current_user["email"]),
        )
        conn.commit()
        return {"message": "Liked"}

@app.delete("/api/forum/{post_id}/like")
def unlike_forum_post(post_id: int, current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"DELETE FROM forum_likes WHERE post_id = {PLACEHOLDER} AND user_email = {PLACEHOLDER}",
            (post_id, current_user["email"]),
        )
        conn.commit()
        return {"message": "Unliked"}

@app.get("/api/forum/likes/me")
def get_my_forum_likes(current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"SELECT post_id FROM forum_likes WHERE user_email = {PLACEHOLDER}",
            (current_user["email"],),
        )
        return [row["post_id"] for row in cur.fetchall()]

@app.post("/api/forum/{post_id}/comments", response_model=ForumCommentOut)
def create_forum_comment(post_id: int, comment: ForumComment, current_user: dict = Depends(get_current_user)):
    sanitized_comment = sanitize_forum_comment_text(comment.comment)
    # Validate comment length (max 100 characters for security)
    if len(sanitized_comment) > 100:
        raise HTTPException(status_code=400, detail="Comment must be 100 characters or less")
    if not sanitized_comment:
        raise HTTPException(status_code=400, detail="Comment cannot be empty")

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"INSERT INTO forum_comments (post_id, user_email, user_full_name, comment) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
            (post_id, current_user["email"], current_user["full_name"], sanitized_comment),
        )
        conn.commit()
        return ForumCommentOut(id=cur.lastrowid, user_full_name=current_user["full_name"], comment=sanitized_comment, created_at=datetime.utcnow().isoformat())

@app.delete("/api/practice/{date_str}")
def delete_practice(date_str: str, current_user: dict = Depends(get_current_user)):
    # Admin only
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"SELECT payment_requested FROM practice_sessions WHERE date = {PLACEHOLDER}",
            (date_str,),
        )
        existing_session = cur.fetchone()
        if not existing_session:
            raise HTTPException(status_code=404, detail="Practice session not found")
        if existing_session["payment_requested"]:
            raise HTTPException(status_code=400, detail="Practice session cannot be deleted after payment has been requested")
        cur.execute(f"DELETE FROM event_media WHERE event_id IN (SELECT id FROM events WHERE practice_session_date = {PLACEHOLDER})", (date_str,))
        cur.execute(f"DELETE FROM event_likes WHERE event_id IN (SELECT id FROM events WHERE practice_session_date = {PLACEHOLDER})", (date_str,))
        cur.execute(f"DELETE FROM event_comments WHERE event_id IN (SELECT id FROM events WHERE practice_session_date = {PLACEHOLDER})", (date_str,))
        cur.execute(f"DELETE FROM events WHERE practice_session_date = {PLACEHOLDER}", (date_str,))
        cur.execute(f"DELETE FROM practice_payments WHERE date = {PLACEHOLDER}", (date_str,))
        cur.execute(f"DELETE FROM practice_availability WHERE date = {PLACEHOLDER}", (date_str,))
        cur.execute(
            f"DELETE FROM notifications WHERE related_date = {PLACEHOLDER} AND type IN ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
            (date_str, "practice", "match", "practice_slot_available", "session_capacity_reached", "payment_request"),
        )
        cur.execute(f"DELETE FROM practice_sessions WHERE date = {PLACEHOLDER}", (date_str,))
        conn.commit()
        return {"message": "Practice session deleted"}

# --- Practice endpoints ---
@app.get("/api/practice/availability")
def get_my_practice_availability(current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        # Only return availability records for current user
        # This prevents reactivated users from seeing old user's selections
        cur.execute(
            f"SELECT practice_session_id, date, status, user_full_name FROM practice_availability WHERE user_email = {PLACEHOLDER}",
            (current_user["email"],),
        )
        result = {}
        for row in cur.fetchall():
            row_dict = dict(row)
            # Only include if user_full_name matches current user's name
            # This ensures reactivated users don't see previous user's selections
            stored_name = row_dict.get("user_full_name")
            # IMPORTANT: Only show if stored name exactly matches current user's name
            # Do NOT show if stored_name is None (old records before migration)
            if stored_name and stored_name == current_user["full_name"] and row_dict.get("practice_session_id") is not None:
                result[str(row_dict["practice_session_id"])] = row_dict["status"]
        return result

@app.post("/api/practice/{date}/availability")
def set_practice_availability_by_date(date: str, status: dict, current_user: dict = Depends(get_current_user)):
    avail = PracticeAvailability(
        date=date,
        status=status.get("status", ""),
        option_choice=status.get("option_choice"),
    )
    return set_my_practice_availability(avail, current_user)

@app.post("/api/practice/availability")
def set_my_practice_availability(avail: PracticeAvailability, current_user: dict = Depends(get_current_user)):
    try:
        datetime.strptime(avail.date, '%Y-%m-%d')
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    with get_connection() as conn:
        cur = conn.cursor()
        session = get_practice_session_basic(cur, avail.date)
        if not session:
            raise HTTPException(status_code=404, detail="Practice session not found")
        return set_practice_availability_for_session_id(cur, conn, session["id"], current_user["email"], current_user["full_name"], avail.status, avail.option_choice)

@app.post("/api/admin/practice/availability")
def admin_set_practice_availability(avail: AdminPracticeAvailability, current_user: dict = Depends(get_current_user)):
    # Only admins can use this endpoint
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    
    with get_connection() as conn:
        cur = conn.cursor()
        
        # If status is 'delete', remove the availability record
        if avail.status == 'delete':
            cur.execute(
                f"DELETE FROM practice_availability WHERE date = {PLACEHOLDER} AND user_email = {PLACEHOLDER}",
                (avail.date, avail.user_email)
            )
            conn.commit()
            return {"message": "Availability removed"}
        
        # Get user's full name
        cur.execute(f"SELECT full_name FROM users WHERE email = {PLACEHOLDER}", (avail.user_email,))
        user_row = cur.fetchone()
        if not user_row:
            raise HTTPException(status_code=404, detail="User not found")
        session = get_practice_session_with_capacity(cur, avail.date)
        if not session:
            raise HTTPException(status_code=404, detail="Practice session not found")
        return set_practice_availability_for_session_id(cur, conn, session["id"], avail.user_email, user_row["full_name"], avail.status, avail.option_choice)

@app.get("/api/practice/availability/{date_str}")
def get_practice_availability_summary(date_str: str):
    with get_connection() as conn:
        cur = conn.cursor()
        session = get_practice_session_with_capacity(cur, date_str)
        if not session:
            raise HTTPException(status_code=404, detail="Practice session not found")
        return get_practice_availability_summary_by_session(cur, session)

# --- Notifications ---
def create_notification(user_email: str, notif_type: str, message: str, related_date: str = None, practice_session_id: int = None):
    """Helper function to create a notification for a user"""
    with get_connection() as conn:
        cur = conn.cursor()
        if related_date and practice_session_id is not None:
            cur.execute(
                f"INSERT INTO notifications (user_email, type, message, related_date, practice_session_id) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
                (user_email, notif_type, message, related_date, practice_session_id)
            )
        elif related_date:
            cur.execute(
                f"INSERT INTO notifications (user_email, type, message, related_date) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
                (user_email, notif_type, message, related_date)
            )
        elif practice_session_id is not None:
            cur.execute(
                f"INSERT INTO notifications (user_email, type, message, practice_session_id) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
                (user_email, notif_type, message, practice_session_id)
            )
        else:
            cur.execute(
                f"INSERT INTO notifications (user_email, type, message) VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})",
                (user_email, notif_type, message)
            )
        conn.commit()

def notify_all_users(notif_type: str, message: str, exclude_email: str = None, related_date: str = None, practice_session_id: int = None):
    """Create notification for all users except the one who triggered it"""
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT email FROM users WHERE (is_deleted = FALSE OR is_deleted IS NULL)")
        users = cur.fetchall()
        for user in users:
            email = user["email"] if USE_POSTGRES else user[0]
            if email != exclude_email:
                create_notification(email, notif_type, message, related_date, practice_session_id)

@app.get("/api/admin/notification-settings", response_model=List[NotificationSettingOut])
def list_notification_settings(current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    settings_map = get_notification_settings_map()
    return [serialize_notification_setting(setting) for setting in settings_map.values()]

@app.get("/api/admin/notification-settings/meta")
def notification_settings_meta(current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    return {
        "target_audiences": [
            {"value": "all_active_users", "label": "All active users"},
            {"value": "admin_users", "label": "Admin users"},
            {"value": "available_players", "label": "Only players marked available for the related practice"},
            {"value": "direct_user", "label": "Only the directly related user"},
        ],
        "notification_types": [
            {"value": notif_type, "label": defaults["display_name"]}
            for notif_type, defaults in NOTIFICATION_TYPE_DEFAULTS.items()
        ],
        "template_variables": [
            "{{event_name}}",
            "{{event_type}}",
            "{{event_type_label}}",
            "{{event_title}}",
            "{{date}}",
            "{{date_iso}}",
            "{{time}}",
            "{{location}}",
            "{{session_id}}",
            "{{maximum_capacity}}",
            "{{available_count}}",
            "{{remaining_slots}}",
            "{{member_name}}",
            "{{full_name}}",
            "{{author_name}}",
            "{{content}}",
            "{{content_preview}}",
            "{{payments_link}}",
            "{{club_name}}",
            "{{time_suffix}}",
            "{{location_suffix}}",
            "{{location_comma_suffix}}",
            "{{time_line}}",
            "{{location_line}}",
        ],
    }

@app.put("/api/admin/notification-settings/{notif_type}", response_model=NotificationSettingOut)
def update_notification_setting(notif_type: str, payload: NotificationSettingUpdate, current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    if notif_type not in NOTIFICATION_TYPE_DEFAULTS:
        raise HTTPException(status_code=404, detail="Notification type not found")
    if payload.target_audience not in NOTIFICATION_TARGET_OPTIONS:
        raise HTTPException(status_code=400, detail="Invalid target audience")

    with get_connection() as conn:
        cur = conn.cursor()
        if USE_POSTGRES:
            cur.execute(
                f"""
                UPDATE notification_settings
                SET display_name = {PLACEHOLDER},
                    description = {PLACEHOLDER},
                    app_enabled = {PLACEHOLDER},
                    email_enabled = {PLACEHOLDER},
                    whatsapp_enabled = {PLACEHOLDER},
                    target_audience = {PLACEHOLDER},
                    app_template = {PLACEHOLDER},
                    email_subject = {PLACEHOLDER},
                    email_template = {PLACEHOLDER},
                    whatsapp_template = {PLACEHOLDER},
                    updated_at = CURRENT_TIMESTAMP
                WHERE notif_type = {PLACEHOLDER}
                """,
                (
                    payload.display_name,
                    payload.description,
                    payload.app_enabled,
                    payload.email_enabled,
                    payload.whatsapp_enabled,
                    payload.target_audience,
                    payload.app_template,
                    payload.email_subject,
                    payload.email_template,
                    payload.whatsapp_template,
                    notif_type,
                ),
            )
        else:
            cur.execute(
                """
                UPDATE notification_settings
                SET display_name = ?,
                    description = ?,
                    app_enabled = ?,
                    email_enabled = ?,
                    whatsapp_enabled = ?,
                    target_audience = ?,
                    app_template = ?,
                    email_subject = ?,
                    email_template = ?,
                    whatsapp_template = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE notif_type = ?
                """,
                (
                    payload.display_name,
                    payload.description,
                    1 if payload.app_enabled else 0,
                    1 if payload.email_enabled else 0,
                    1 if payload.whatsapp_enabled else 0,
                    payload.target_audience,
                    payload.app_template,
                    payload.email_subject,
                    payload.email_template,
                    payload.whatsapp_template,
                    notif_type,
                ),
            )
        conn.commit()

    return serialize_notification_setting(get_notification_setting(notif_type))

@app.post("/api/admin/notification-settings/{notif_type}/reset", response_model=NotificationSettingOut)
def reset_notification_setting(notif_type: str, current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    defaults = NOTIFICATION_TYPE_DEFAULTS.get(notif_type)
    if not defaults:
        raise HTTPException(status_code=404, detail="Notification type not found")

    with get_connection() as conn:
        cur = conn.cursor()
        if USE_POSTGRES:
            cur.execute(
                f"""
                UPDATE notification_settings
                SET display_name = {PLACEHOLDER},
                    description = {PLACEHOLDER},
                    app_enabled = {PLACEHOLDER},
                    email_enabled = {PLACEHOLDER},
                    whatsapp_enabled = {PLACEHOLDER},
                    target_audience = {PLACEHOLDER},
                    app_template = {PLACEHOLDER},
                    email_subject = {PLACEHOLDER},
                    email_template = {PLACEHOLDER},
                    whatsapp_template = {PLACEHOLDER},
                    updated_at = CURRENT_TIMESTAMP
                WHERE notif_type = {PLACEHOLDER}
                """,
                (
                    defaults["display_name"],
                    defaults["description"],
                    defaults["app_enabled"],
                    defaults["email_enabled"],
                    defaults["whatsapp_enabled"],
                    defaults["target_audience"],
                    defaults["app_template"],
                    defaults["email_subject"],
                    defaults["email_template"],
                    defaults["whatsapp_template"],
                    notif_type,
                ),
            )
        else:
            cur.execute(
                """
                UPDATE notification_settings
                SET display_name = ?,
                    description = ?,
                    app_enabled = ?,
                    email_enabled = ?,
                    whatsapp_enabled = ?,
                    target_audience = ?,
                    app_template = ?,
                    email_subject = ?,
                    email_template = ?,
                    whatsapp_template = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE notif_type = ?
                """,
                (
                    defaults["display_name"],
                    defaults["description"],
                    1 if defaults["app_enabled"] else 0,
                    1 if defaults["email_enabled"] else 0,
                    1 if defaults["whatsapp_enabled"] else 0,
                    defaults["target_audience"],
                    defaults["app_template"],
                    defaults["email_subject"],
                    defaults["email_template"],
                    defaults["whatsapp_template"],
                    notif_type,
                ),
            )
        conn.commit()

    return serialize_notification_setting(get_notification_setting(notif_type))

@app.get("/api/admin/whatsapp/status")
def whatsapp_status(current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    state = get_instance_state()
    target_group = resolve_group_chat_id()
    return {
        "configured": whatsapp_is_configured(),
        "enabled": WHATSAPP_NOTIFICATIONS_ENABLED,
        "state": state,
        "target_group": target_group,
    }

@app.post("/api/admin/whatsapp/test")
def send_test_whatsapp_message(data: WhatsAppMessageRequest, current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    message = data.message.strip()
    if not message:
        raise HTTPException(status_code=400, detail="Message is required")
    result = send_group_message(message)
    if not result.get("success"):
        raise HTTPException(status_code=503, detail=result.get("error", "Failed to send WhatsApp message"))
    return result

@app.post("/api/admin/whatsapp/find-group")
def lookup_whatsapp_group(data: WhatsAppGroupLookupRequest, current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    group_name = data.group_name.strip()
    if not group_name:
        raise HTTPException(status_code=400, detail="Group name is required")
    result = find_group_chat_id(group_name)
    if not result.get("success"):
        raise HTTPException(status_code=503, detail=result.get("error", "Failed to fetch WhatsApp chats"))
    return result

@app.get("/api/notifications")
def get_notifications(current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        
        # Check if related_date column exists before querying
        if USE_POSTGRES:
            cur.execute("""
                SELECT column_name FROM information_schema.columns 
                WHERE table_name='notifications' AND column_name='related_date'
            """)
            has_related_date = cur.fetchone() is not None
            cur.execute("""
                SELECT column_name FROM information_schema.columns 
                WHERE table_name='notifications' AND column_name='practice_session_id'
            """)
            has_practice_session_id = cur.fetchone() is not None
        else:
            cur.execute("PRAGMA table_info(notifications)")
            columns = [col[1] for col in cur.fetchall()]
            has_related_date = 'related_date' in columns
            has_practice_session_id = 'practice_session_id' in columns
        
        # Build query based on column existence
        if has_related_date and has_practice_session_id:
            cur.execute(
                f"SELECT id, type, message, read, related_date, practice_session_id, created_at FROM notifications WHERE user_email = {PLACEHOLDER} ORDER BY created_at DESC LIMIT 50",
                (current_user["email"],)
            )
        elif has_related_date:
            cur.execute(
                f"SELECT id, type, message, read, related_date, created_at FROM notifications WHERE user_email = {PLACEHOLDER} ORDER BY created_at DESC LIMIT 50",
                (current_user["email"],)
            )
        else:
            cur.execute(
                f"SELECT id, type, message, read, created_at FROM notifications WHERE user_email = {PLACEHOLDER} ORDER BY created_at DESC LIMIT 50",
                (current_user["email"],)
            )
        
        rows = cur.fetchall()
        notifications = []
        for row in rows:
            # Extract related_date if column exists
            if has_related_date:
                if USE_POSTGRES:
                    related_date = row["related_date"] if USE_POSTGRES else row[4]
                else:
                    related_date = row[4]
                
                if related_date and isinstance(related_date, str):
                    # If it's a string with timestamp, extract just the date part
                    related_date = related_date.split(' ')[0] if ' ' in related_date else related_date
            else:
                related_date = None

            if has_practice_session_id:
                if USE_POSTGRES:
                    practice_session_id = row["practice_session_id"]
                else:
                    practice_session_id = row[5] if has_related_date else None
            else:
                practice_session_id = None
            
            if USE_POSTGRES:
                notifications.append({
                    "id": row["id"],
                    "type": row["type"],
                    "message": row["message"],
                    "read": row["read"],
                    "related_date": related_date,
                    "practice_session_id": practice_session_id,
                    "created_at": row["created_at"],
                })
            else:
                notifications.append({
                    "id": row[0],
                    "type": row[1],
                    "message": row[2],
                    "read": bool(row[3]),
                    "related_date": related_date,
                    "practice_session_id": practice_session_id,
                    "created_at": row[6] if has_related_date and has_practice_session_id else (row[5] if has_related_date else row[4]),
                })
        return notifications

@app.post("/api/notifications/mark-read")
def mark_notifications_read(current_user: dict = Depends(get_current_user)):
    with get_connection() as conn:
        cur = conn.cursor()
        if USE_POSTGRES:
            cur.execute(
                f"UPDATE notifications SET read = TRUE WHERE user_email = {PLACEHOLDER}",
                (current_user["email"],)
            )
        else:
            cur.execute(
                f"UPDATE notifications SET read = 1 WHERE user_email = {PLACEHOLDER}",
                (current_user["email"],)
            )
        conn.commit()
        return {"message": "Notifications marked as read"}

# --- Reports ---
@app.get("/api/reports/booking")
def generate_booking_report(from_date: str, to_date: str, current_user: dict = Depends(get_current_user)):
    """Generate Booking Report Excel file"""
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    
    with get_connection() as conn:
        cur = conn.cursor()
        
        # Get practice sessions with payment info in date range
        cur.execute(f"""
            SELECT 
                ps.date,
                ps.event_type,
                ps.event_title,
                ps.time,
                ps.location,
                ps.session_cost,
                ps.paid_by,
                u.full_name as paid_by_name
            FROM practice_sessions ps
            LEFT JOIN users u ON ps.paid_by = u.email AND (u.is_deleted = FALSE OR u.is_deleted IS NULL)
            WHERE ps.date >= {PLACEHOLDER} AND ps.date <= {PLACEHOLDER}
            ORDER BY ps.date ASC
        """, (from_date, to_date))
        
        sessions = cur.fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Booking Report"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ["Event Date", "Event Type", "Event Title", "Time", "Place", "Total Cost (£)", "Paid By"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        for row_num, session in enumerate(sessions, 2):
            if USE_POSTGRES:
                event_type = normalize_event_type(session.get("event_type")) if session.get("event_type") else "practice"
                event_title = normalize_event_title(session.get("event_title"), event_type)
                ws.cell(row=row_num, column=1, value=session["date"])
                ws.cell(row=row_num, column=2, value=default_event_type_label(event_type))
                ws.cell(row=row_num, column=3, value=event_title)
                ws.cell(row=row_num, column=4, value=session["time"] or "TBD")
                ws.cell(row=row_num, column=5, value=session["location"] or "TBD")
                ws.cell(row=row_num, column=6, value=float(session["session_cost"]) if session["session_cost"] else 0.0)
                ws.cell(row=row_num, column=7, value=session["paid_by_name"] or session["paid_by"] or "Not Set")
            else:
                event_type = normalize_event_type(session[1]) if session[1] else "practice"
                event_title = normalize_event_title(session[2], event_type)
                ws.cell(row=row_num, column=1, value=session[0])
                ws.cell(row=row_num, column=2, value=default_event_type_label(event_type))
                ws.cell(row=row_num, column=3, value=event_title)
                ws.cell(row=row_num, column=4, value=session[3] or "TBD")
                ws.cell(row=row_num, column=5, value=session[4] or "TBD")
                ws.cell(row=row_num, column=6, value=float(session[5]) if session[5] else 0.0)
                ws.cell(row=row_num, column=7, value=session[7] or session[6] or "Not Set")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Booking_Report_{from_date}_to_{to_date}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

@app.get("/api/reports/expense")
def generate_expense_report(from_date: str, to_date: str, current_user: dict = Depends(get_current_user)):
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT e.expense_date, e.title, e.category, e.amount, e.payment_method, e.description,
                   e.paid_by, u.full_name as paid_by_name, e.created_at, 'Expense' as source
            FROM expenses e
            LEFT JOIN users u ON e.paid_by = u.email AND (u.is_deleted = FALSE OR u.is_deleted IS NULL)
            WHERE e.expense_date >= {PLACEHOLDER} AND e.expense_date <= {PLACEHOLDER}
            """,
            (from_date, to_date),
        )
        expense_rows = [dict(row) for row in cur.fetchall()]

        cur.execute(
            f"""
            SELECT ps.date as expense_date,
                   ps.event_type,
                   ps.event_title,
                   ps.session_cost as amount,
                   NULL as payment_method,
                   ps.time,
                   ps.location,
                   ps.paid_by,
                   u.full_name as paid_by_name,
                   ps.payment_requested_at as created_at,
                   'Event Booking' as source
            FROM practice_sessions ps
            LEFT JOIN users u ON ps.paid_by = u.email AND (u.is_deleted = FALSE OR u.is_deleted IS NULL)
            WHERE ps.session_cost IS NOT NULL AND ps.date >= {PLACEHOLDER} AND ps.date <= {PLACEHOLDER}
            """,
            (from_date, to_date),
        )
        booking_rows = []
        for row in cur.fetchall():
            row_dict = dict(row)
            event_type = normalize_event_type(row_dict.get("event_type")) if row_dict.get("event_type") else "practice"
            event_title = normalize_event_title(row_dict.get("event_title"), event_type)
            event_name = f"{default_event_type_label(event_type)} - {event_title}"
            time_value = row_dict.get("time")
            location_value = row_dict.get("location")
            if time_value and location_value:
                description = f"{event_name} booking at {time_value} - {location_value}"
            elif time_value:
                description = f"{event_name} booking at {time_value}"
            elif location_value:
                description = f"{event_name} booking - {location_value}"
            else:
                description = f"{event_name} booking cost"
            booking_rows.append({
                "expense_date": row_dict.get("expense_date"),
                "title": event_name,
                "category": "Event Booking",
                "amount": row_dict.get("amount"),
                "payment_method": row_dict.get("payment_method"),
                "description": description,
                "paid_by": row_dict.get("paid_by"),
                "paid_by_name": row_dict.get("paid_by_name"),
                "created_at": row_dict.get("created_at"),
                "source": row_dict.get("source"),
            })

        rows = expense_rows + booking_rows
        rows.sort(
            key=lambda item: (
                item.get("expense_date") or "",
                item.get("created_at") or "",
                item.get("title") or "",
            ),
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Expense Report"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = ["Expense Date", "Title", "Category", "Amount (£)", "Paid By", "Payment Method", "Description", "Created Date"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row_num, row in enumerate(rows, 2):
            row_dict = dict(row)
            ws.cell(row=row_num, column=1, value=row_dict.get("expense_date"))
            ws.cell(row=row_num, column=2, value=row_dict.get("title"))
            ws.cell(row=row_num, column=3, value=row_dict.get("category") or "")
            ws.cell(row=row_num, column=4, value=float(row_dict.get("amount") or 0))
            ws.cell(row=row_num, column=5, value=row_dict.get("paid_by_name") or row_dict.get("paid_by") or "")
            ws.cell(row=row_num, column=6, value=row_dict.get("payment_method") or "")
            ws.cell(row=row_num, column=7, value=row_dict.get("description") or "")
            created_at = row_dict.get("created_at")
            created_date_value = None
            if isinstance(created_at, datetime):
                created_date_value = created_at.date()
            elif isinstance(created_at, date):
                created_date_value = created_at
            elif created_at:
                try:
                    created_date_value = datetime.fromisoformat(str(created_at).replace("Z", "+00:00")).date()
                except ValueError:
                    try:
                        created_date_value = datetime.strptime(str(created_at)[:10], "%Y-%m-%d").date()
                    except ValueError:
                        created_date_value = None
            created_date_cell = ws.cell(row=row_num, column=8, value=created_date_value if created_date_value else (str(created_at)[:10] if created_at else ""))
            if created_date_value:
                created_date_cell.number_format = "yyyy-mm-dd"

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 24
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 36
        ws.column_dimensions['H'].width = 24

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Expense_Report_{from_date}_to_{to_date}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

@app.get("/api/reports/player-payment")
def generate_player_payment_report(from_date: str, to_date: str, current_user: dict = Depends(get_current_user)):
    """Generate Player Payment Report Excel file"""
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admins only")
    
    with get_connection() as conn:
        cur = conn.cursor()
        
        # Get all bookable events in date range with availability and payment data
        cur.execute(f"""
            SELECT 
                ps.id,
                ps.date,
                ps.event_type,
                ps.event_title,
                ps.time,
                ps.location,
                ps.session_cost,
                ps.paid_by,
                ps.payment_requested_at,
                pa.user_email,
                u.full_name,
                payer.full_name as paid_by_name,
                pa.status,
                pp.paid,
                pp.created_at as payment_date
            FROM practice_sessions ps
            LEFT JOIN practice_availability pa ON ps.id = pa.practice_session_id
            LEFT JOIN users u ON pa.user_email = u.email AND (u.is_deleted = FALSE OR u.is_deleted IS NULL)
            LEFT JOIN users payer ON ps.paid_by = payer.email AND (payer.is_deleted = FALSE OR payer.is_deleted IS NULL)
            LEFT JOIN practice_payments pp ON ps.id = pp.practice_session_id AND pa.user_email = pp.user_email
            WHERE ps.date >= {PLACEHOLDER} AND ps.date <= {PLACEHOLDER}
                AND pa.status IS NOT NULL
            ORDER BY ps.date ASC, ps.time ASC, ps.payment_requested_at ASC, u.full_name ASC
        """, (from_date, to_date))
        
        rows = cur.fetchall()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Event Payment Report"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ["Event Date", "Event Type", "Event Title", "Time", "Place", "Total Cost (£)", "Paid By", "Payment Requested Date", "Player Name", "Availability", 
                   "Individual Amount (£)", "Paid", "Payment Acknowledgement Date"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        row_num = 2
        for row in rows:
            if USE_POSTGRES:
                event_date = row["date"]
                event_type = normalize_event_type(row.get("event_type")) if row.get("event_type") else "practice"
                event_title = normalize_event_title(row.get("event_title"), event_type)
                time = row["time"] or "TBD"
                location = row["location"] or "TBD"
                session_cost = row["session_cost"]
                paid_by = row["paid_by_name"] or row["paid_by"]
                payment_requested_at = row["payment_requested_at"]
                user_email = row["user_email"]
                full_name = row["full_name"] or user_email
                status = row["status"]
                paid = row["paid"]
                payment_date = row["payment_date"]
            else:
                event_date = row[1]  # shifted because ps.id is now row[0]
                event_type = normalize_event_type(row[2]) if row[2] else "practice"
                event_title = normalize_event_title(row[3], event_type)
                time = row[4] or "TBD"
                location = row[5] or "TBD"
                session_cost = row[6]
                paid_by = row[11] or row[7]
                payment_requested_at = row[8]
                user_email = row[9]
                full_name = row[10] or user_email
                status = row[12]
                paid = row[13]
                payment_date = row[14]
            
            ws.cell(row=row_num, column=1, value=event_date)
            ws.cell(row=row_num, column=2, value=default_event_type_label(event_type))
            ws.cell(row=row_num, column=3, value=event_title)
            ws.cell(row=row_num, column=4, value=time)
            ws.cell(row=row_num, column=5, value=location)
            ws.cell(row=row_num, column=6, value=float(session_cost) if session_cost else 0)
            ws.cell(row=row_num, column=7, value=paid_by)
            if payment_requested_at:
                payment_requested_date_value = None
                if isinstance(payment_requested_at, datetime):
                    payment_requested_date_value = payment_requested_at.date()
                elif isinstance(payment_requested_at, date):
                    payment_requested_date_value = payment_requested_at
                else:
                    try:
                        payment_requested_date_value = datetime.fromisoformat(str(payment_requested_at).replace("Z", "+00:00")).date()
                    except ValueError:
                        try:
                            payment_requested_date_value = datetime.strptime(str(payment_requested_at)[:10], "%Y-%m-%d").date()
                        except ValueError:
                            payment_requested_date_value = None
                payment_requested_cell = ws.cell(row=row_num, column=8, value=payment_requested_date_value if payment_requested_date_value else str(payment_requested_at)[:10])
                if payment_requested_date_value:
                    payment_requested_cell.number_format = "yyyy-mm-dd"
            else:
                ws.cell(row=row_num, column=8, value="")
            ws.cell(row=row_num, column=9, value=full_name)
            ws.cell(row=row_num, column=10, value=status.capitalize() if status else "")

            individual_amount = 0.0
            if status == "available" and session_cost:
                # Get count of available users for this specific session
                if USE_POSTGRES:
                    cur.execute(f"""
                        SELECT COUNT(*) as count 
                        FROM practice_availability 
                        WHERE practice_session_id = {PLACEHOLDER} AND status = 'available'
                    """, (row["id"],))
                else:
                    cur.execute(f"""
                        SELECT COUNT(*) as count 
                        FROM practice_availability 
                        WHERE practice_session_id = ? AND status = 'available'
                    """, (row[0],))  # row[0] is the session ID in SQLite
                
                count_row = cur.fetchone()
                available_count = count_row["count"] if USE_POSTGRES else count_row[0]
                
                if available_count > 0:
                    individual_amount = float(session_cost) / available_count
            ws.cell(row=row_num, column=11, value=round(individual_amount, 2))
            ws.cell(row=row_num, column=12, value="Yes" if (paid if USE_POSTGRES else bool(paid)) else "No")
            payment_ack_date_value = None
            if isinstance(payment_date, datetime):
                payment_ack_date_value = payment_date.date()
            elif isinstance(payment_date, date):
                payment_ack_date_value = payment_date
            elif payment_date:
                try:
                    payment_ack_date_value = datetime.fromisoformat(str(payment_date).replace("Z", "+00:00")).date()
                except ValueError:
                    try:
                        payment_ack_date_value = datetime.strptime(str(payment_date)[:10], "%Y-%m-%d").date()
                    except ValueError:
                        payment_ack_date_value = None
            payment_ack_cell = ws.cell(row=row_num, column=13, value=payment_ack_date_value if payment_ack_date_value else ((str(payment_date)[:10]) if payment_date else ''))
            if payment_ack_date_value:
                payment_ack_cell.number_format = "yyyy-mm-dd"
            row_num += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 24
        ws.column_dimensions['H'].width = 22
        ws.column_dimensions['I'].width = 24
        ws.column_dimensions['J'].width = 16
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 24
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Player_Payment_Report_{from_date}_to_{to_date}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

@app.get("/api/user-actions/upcoming-sessions")
def get_upcoming_sessions(current_user: dict = Depends(get_current_user)):
    """Get all upcoming bookable events for user to set availability"""
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT ps.id, ps.date, ps.time, ps.location, ps.event_type, ps.event_title, ps.session_cost, ps.paid_by,
                   COALESCE(ps.maximum_capacity, 100) as maximum_capacity,
                   pa.status as user_status
            FROM practice_sessions ps
            LEFT JOIN practice_availability pa 
                ON ps.id = pa.practice_session_id AND pa.user_email = {PLACEHOLDER}
            ORDER BY ps.date ASC, COALESCE(ps.time, '') ASC, ps.id ASC
            """,
            (current_user["email"],)
        )
        
        rows = cur.fetchall()
        sessions = []
        for row in rows:
            row_dict = dict(row)
            normalized_time = get_practice_effective_time(row_dict.get("time")) if row_dict.get("time") else "21:00"
            if is_practice_datetime_in_past(row_dict["date"], normalized_time):
                continue
            available_count = get_available_count_for_session_id(cur, row_dict["id"])
            maximum_capacity = normalize_maximum_capacity(row_dict["maximum_capacity"])
            sessions.append({
                "id": row_dict["id"],
                "date": row_dict["date"],
                "time": normalized_time,
                "location": row_dict["location"],
                "event_type": normalize_event_type(row_dict.get("event_type")) if row_dict.get("event_type") else "practice",
                "event_title": normalize_event_title(row_dict.get("event_title"), normalize_event_type(row_dict.get("event_type")) if row_dict.get("event_type") else "practice"),
                "session_cost": row_dict["session_cost"],
                "paid_by": row_dict["paid_by"],
                "user_status": row_dict["user_status"],
                "maximum_capacity": maximum_capacity,
                "available_count": available_count,
                "remaining_slots": max(maximum_capacity - available_count, 0),
                "capacity_reached": available_count >= maximum_capacity,
            })
        
        return {"sessions": sessions}

@app.get("/api/user-actions/payments")
def get_pending_payments(current_user: dict = Depends(get_current_user)):
    """Get all events where user was available but hasn't confirmed payment"""
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT ps.id, ps.date, ps.time, ps.location, ps.event_type, ps.event_title, ps.session_cost, ps.paid_by,
                   u.full_name as paid_by_name,
                   u.bank_name as paid_by_bank_name,
                   u.sort_code as paid_by_sort_code,
                   u.account_number as paid_by_account_number,
                   pa.status,
                   COALESCE(pp.paid, {PLACEHOLDER}) as paid
            FROM practice_sessions ps
            INNER JOIN practice_availability pa 
                ON ps.id = pa.practice_session_id AND pa.user_email = {PLACEHOLDER}
            LEFT JOIN practice_payments pp 
                ON ps.id = pp.practice_session_id AND pp.user_email = {PLACEHOLDER}
            LEFT JOIN users u ON ps.paid_by = u.email
            WHERE pa.status = 'available' 
              AND ps.payment_requested = {PLACEHOLDER}
              AND (pp.paid IS NULL OR pp.paid = {PLACEHOLDER})
            ORDER BY ps.date DESC, COALESCE(ps.time, '') DESC, ps.id DESC
            """,
            (
                False if USE_POSTGRES else 0,
                current_user["email"],
                current_user["email"],
                True if USE_POSTGRES else 1,
                False if USE_POSTGRES else 0,
            )
        )
        
        rows = cur.fetchall()
        payments = []
        for row in rows:
            if USE_POSTGRES:
                if not is_practice_datetime_in_past(row["date"], row.get("time")):
                    continue
                available_count = get_available_count_for_session_id(cur, row["id"])
                session_cost = float(row["session_cost"]) if row["session_cost"] is not None else 0
                individual_amount = session_cost / available_count if available_count > 0 and session_cost > 0 else 0
                
                payments.append({
                    "id": row["id"],
                    "date": row["date"],
                    "time": row["time"],
                    "location": row["location"],
                    "event_type": normalize_event_type(row.get("event_type")) if row.get("event_type") else "practice",
                    "event_title": normalize_event_title(row.get("event_title"), normalize_event_type(row.get("event_type")) if row.get("event_type") else "practice"),
                    "session_cost": session_cost,
                    "individual_amount": round(individual_amount, 2),
                    "paid_by": row["paid_by"],
                    "paid_by_name": row["paid_by_name"],
                    "paid_by_bank_name": row["paid_by_bank_name"],
                    "paid_by_sort_code": row["paid_by_sort_code"],
                    "paid_by_account_number": row["paid_by_account_number"],
                    "paid": row["paid"]
                })
            else:
                if not is_practice_datetime_in_past(row[1], row[2]):
                    continue
                available_count = get_available_count_for_session_id(cur, row[0])
                session_cost = float(row[6]) if row[6] is not None else 0
                individual_amount = session_cost / available_count if available_count > 0 and session_cost > 0 else 0
                
                payments.append({
                    "id": row[0],
                    "date": row[1],
                    "time": row[2],
                    "location": row[3],
                    "event_type": normalize_event_type(row[4]) if row[4] else "practice",
                    "event_title": normalize_event_title(row[5], normalize_event_type(row[4]) if row[4] else "practice"),
                    "session_cost": session_cost,
                    "individual_amount": round(individual_amount, 2),
                    "paid_by": row[7],
                    "paid_by_name": row[8],
                    "paid_by_bank_name": row[9],
                    "paid_by_sort_code": row[10],
                    "paid_by_account_number": row[11],
                    "paid": row[13],
                })
        
        return {"payments": payments}

@app.get("/api/user-actions/pending-payments")
def get_pending_payments_alias(current_user: dict = Depends(get_current_user)):
    return get_pending_payments(current_user)

# --- About Us ---
@app.get("/api/about")
def get_about():
    # Static data for now; can be moved to DB later
    return {
        "club_name": "Glasgow Bengali Football Club",
        "summary": "A community-driven Bengali football club in Glasgow, bringing players, families, and fans together every week.",
        "committee": [
            {"name": "Arif Rahman", "role": "President", "image": "https://images.pexels.com/photos/2182970/pexels-photo-2182970.jpeg"},
            {"name": "Sara Khan", "role": "Treasurer", "image": "https://images.pexels.com/photos/3184418/pexels-photo-3184418.jpeg"},
            {"name": "Jamal Ahmed", "role": "Coach", "image": "https://images.pexels.com/photos/2379004/pexels-photo-2379004.jpeg"},
        ],
        "members": [
            {"name": "Riaz Uddin", "intro": "Midfielder, joined 2022"},
            {"name": "Nadia Islam", "intro": "Forward, joined 2023"},
            {"name": "Tariq Mehmood", "intro": "Defender, joined 2021"},
        ],
    }

if __name__ == "__main__":
    init_db()
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
